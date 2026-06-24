import os
from pathlib import Path

from dotenv import load_dotenv

_ROOT = Path(__file__).resolve().parent
_LOCAL_DB_FILE = _ROOT / "data" / "local_database.url"
_LOCAL_DB_DEFAULTS = (
    "postgresql://postgres:postgres@localhost:5432/postgres",
    "postgresql://postgres@localhost:5432/postgres",
)


def _load_env_files():
    # .env base; .env.local sobrescreve só em desenvolvimento (não usado no Render).
    load_dotenv(dotenv_path=_ROOT / ".env", encoding="utf-8-sig")
    if not _is_render_deploy():
        load_dotenv(dotenv_path=_ROOT / ".env.local", encoding="utf-8-sig", override=True)


def _is_render_deploy():
    return bool(
        os.getenv("RENDER")
        or os.getenv("RENDER_SERVICE_ID")
        or os.getenv("RENDER_EXTERNAL_URL")
    )


def _is_remote_db_url(url):
    u = (url or "").lower()
    return "localhost" not in u and "127.0.0.1" not in u


def _apply_sslmode(url):
    if not url or not _is_remote_db_url(url) or "sslmode=" in url:
        return url
    return url + ("&" if "?" in url else "?") + "sslmode=require"


def _read_db_url_file():
    if not _LOCAL_DB_FILE.is_file():
        return ""
    return _LOCAL_DB_FILE.read_text(encoding="utf-8-sig").strip()


def _try_connect(url):
    try:
        import psycopg2

        conn = psycopg2.connect(url, connect_timeout=2)
        conn.close()
        return True
    except Exception:
        return False


def _resolve_database_url():
    explicit = os.getenv("DATABASE_URL", "").strip()
    if explicit:
        return _apply_sslmode(explicit)

    if _is_render_deploy():
        return ""

    candidates = []
    local = os.getenv("LOCAL_DATABASE_URL", "").strip()
    if local:
        candidates.append(local)
    file_url = _read_db_url_file()
    if file_url:
        candidates.append(file_url)
    candidates.extend(_LOCAL_DB_DEFAULTS)

    seen = set()
    for candidate in candidates:
        url = (candidate or "").strip()
        if not url or url in seen:
            continue
        seen.add(url)
        if _try_connect(url):
            return url

    return candidates[0] if candidates else ""


_load_env_files()
DATABASE_URL = _resolve_database_url()

import urllib
import urllib.request
import urllib.parse
import json
from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from datetime import datetime
from zoneinfo import ZoneInfo
import pandas as pd
import io, socket, shutil, uuid, json, math, re, base64, html
import psycopg2
from psycopg2.extras import RealDictCursor, Json

app = FastAPI(title="Sistema Interno de Reboque V13 - Faturamento")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount('/static', StaticFiles(directory='static'), name='static')
templates = Jinja2Templates(directory='templates')
UPLOAD_DIR = os.path.join('static', 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
CADASTROS_UPLOAD_DIR = os.path.join('static', 'uploads', 'cadastros')
os.makedirs(CADASTROS_UPLOAD_DIR, exist_ok=True)
FINANCEIRO_UPLOAD_DIR = os.path.join('static', 'uploads', 'financeiro')
os.makedirs(FINANCEIRO_UPLOAD_DIR, exist_ok=True)
CHECKLIST_UPLOAD_DIR = os.path.join('static', 'uploads', 'checklist')
os.makedirs(CHECKLIST_UPLOAD_DIR, exist_ok=True)

MODULOS_CONTROLE_HUB = [
    {"titulo": "Danos", "icone": "💥", "descricao": "Registro e acompanhamento de danos em viaturas, objetos e operações.", "indicador": "Frota", "rota": "/controle/danos"},
    {"titulo": "Abastecimentos", "icone": "⛽", "descricao": "Controle de combustível, postos, litros e custos por viatura.", "indicador": "Combustível", "rota": "/controle/abastecimentos"},
    {"titulo": "Checklist Viatura", "icone": "✅", "descricao": "Checklists de saída, retorno e inspeção das viaturas da frota.", "indicador": "Inspeção", "rota": "/controle/checklist-viatura"},
    {"titulo": "Manutenções", "icone": "🔧", "descricao": "Ordens de serviço, fornecedores e custos de manutenção.", "indicador": "Oficina", "rota": "/controle/manutencoes"},
    {"titulo": "Multas", "icone": "🚨", "descricao": "Infrações, condutores, vencimentos e situação de pagamento.", "indicador": "Compliance", "rota": "/controle/multas"},
    {"titulo": "Seguros", "icone": "🛡️", "descricao": "Apólices, vigências, parcelas e situação dos seguros da frota.", "indicador": "Apólices", "rota": "/controle/seguros"},
]


@app.get("/controle", response_class=HTMLResponse)
def controle_home(request: Request):
    return templates.TemplateResponse(
        "controle.html",
        {
            "request": request,
            "nav_ativo": "controle",
            "nav_som": False,
            "modulos": MODULOS_CONTROLE_HUB,
        },
    )

MODULOS_CADASTROS_HUB = [
    {
        "titulo": "Viaturas",
        "icone": "🚚",
        "descricao": "Frota, documentos, configurações e histórico operacional por viatura.",
        "indicador": "Frota",
        "rota": "/cadastros/viaturas",
        "ativos": 0,
    },
    {
        "titulo": "Profissionais",
        "icone": "👤",
        "descricao": "Motoristas, prestadores e equipe com documentos e permissões de operação.",
        "indicador": "Equipe",
        "rota": "/cadastros/profissionais",
        "ativos": 0,
    },
    {
        "titulo": "Clientes",
        "icone": "🤝",
        "descricao": "Cadastro mestre de contatos classificados como clientes.",
        "indicador": "Receber",
        "rota": "/cadastros/clientes",
        "ativos": 0,
    },
    {
        "titulo": "Fornecedores",
        "icone": "🏭",
        "descricao": "Cadastro mestre de contatos classificados como fornecedores.",
        "indicador": "Pagar",
        "rota": "/cadastros/fornecedores",
        "ativos": 0,
    },
]


def _hub_cadastros_modulos():
    v = resumo_viaturas_cadastro()
    p = resumo_profissionais_cadastro()
    c = resumo_contatos_cadastro(cliente=True)
    f = resumo_contatos_cadastro(fornecedor=True)
    return [
        {**MODULOS_CADASTROS_HUB[0], "ativos": v.get("ativas", 0)},
        {**MODULOS_CADASTROS_HUB[1], "ativos": p.get("ativos", 0)},
        {**MODULOS_CADASTROS_HUB[2], "ativos": c.get("ativos", 0)},
        {**MODULOS_CADASTROS_HUB[3], "ativos": f.get("ativos", 0)},
    ]


@app.get("/cadastros", response_class=HTMLResponse)
def cadastros_home(request: Request):
    return templates.TemplateResponse(
        "cadastros/hub.html",
        {
            "request": request,
            "nav_ativo": "cadastros",
            "nav_som": False,
            "modulos": _hub_cadastros_modulos(),
        },
    )


@app.get("/cadastros/viaturas", response_class=HTMLResponse)
def cadastros_pagina_viaturas(request: Request):
    return templates.TemplateResponse(
        "cadastros/viaturas.html",
        {
            "request": request,
            "nav_ativo": "cadastros",
            "nav_som": False,
            "kpis": resumo_viaturas_cadastro(),
        },
    )


@app.get("/cadastros/profissionais", response_class=HTMLResponse)
def cadastros_pagina_profissionais(request: Request):
    return templates.TemplateResponse(
        "cadastros/profissionais.html",
        {
            "request": request,
            "nav_ativo": "cadastros",
            "nav_som": False,
            "kpis": resumo_profissionais_cadastro(),
        },
    )


def _ctx_pagina_contatos(request: Request, modo: str):
    cliente = modo == "cliente"
    fornecedor = modo == "fornecedor"
    titulos = {
        "cliente": ("Clientes", "🤝", "Contatos classificados como clientes — usados em Contas a Receber."),
        "fornecedor": ("Fornecedores", "🏭", "Contatos classificados como fornecedores — usados em Contas a Pagar."),
    }
    t, icone, desc = titulos.get(modo, titulos["cliente"])
    return {
        "request": request,
        "nav_ativo": "cadastros",
        "nav_som": False,
        "modo": modo,
        "titulo_modulo": t,
        "icone_modulo": icone,
        "descricao_modulo": desc,
        "kpis": resumo_contatos_cadastro(cliente=cliente, fornecedor=fornecedor),
        "filtro_cliente": cliente,
        "filtro_fornecedor": fornecedor,
    }


@app.get("/cadastros/clientes", response_class=HTMLResponse)
def cadastros_pagina_clientes(request: Request):
    return templates.TemplateResponse(
        "cadastros/contatos.html",
        _ctx_pagina_contatos(request, "cliente"),
    )


@app.get("/cadastros/fornecedores", response_class=HTMLResponse)
def cadastros_pagina_fornecedores(request: Request):
    return templates.TemplateResponse(
        "cadastros/contatos.html",
        _ctx_pagina_contatos(request, "fornecedor"),
    )


PRICE_DATA = [
  {
    "tipo": "C. MEC. LEVE",
    "nome": "SAIDA",
    "valor": 105.0
  },
  {
    "tipo": "E. PATIO",
    "nome": "ESTADIA",
    "valor": 45.0
  },
  {
    "tipo": "GUINDAUTO",
    "nome": "SAIDA",
    "valor": 450.89
  },
  {
    "tipo": "GUINDAUTO",
    "nome": "HORA TRABALHADA",
    "valor": 133.0
  },
  {
    "tipo": "R. 5 RODA",
    "nome": "HORA TRABALHADA",
    "valor": 145.0
  },
  {
    "tipo": "R. 5 RODA",
    "nome": "SAIDA",
    "valor": 530.0
  },
  {
    "tipo": "R. 5 RODA",
    "nome": "HORA PARADA",
    "valor": 120.0
  },
  {
    "tipo": "R. 5 RODA",
    "nome": "KM VIAGEM",
    "valor": 5.6
  },
  {
    "tipo": "R. E. PESADO",
    "nome": "ESTADIA",
    "valor": 100.0
  },
  {
    "tipo": "R. E. PESADO",
    "nome": "PEDAGIO",
    "valor": 1.0
  },
  {
    "tipo": "R. E. PESADO",
    "nome": "HORA TRABALHADA",
    "valor": 145.0
  },
  {
    "tipo": "R. E. PESADO",
    "nome": "HORA PARADA",
    "valor": 120.0
  },
  {
    "tipo": "R. E. PESADO",
    "nome": "KM DESLOCAMENTO",
    "valor": 6.36
  },
  {
    "tipo": "R. E. PESADO",
    "nome": "KM VIAGEM",
    "valor": 6.36
  },
  {
    "tipo": "R. E. PESADO",
    "nome": "SAIDA",
    "valor": 570.48
  },
  {
    "tipo": "R. GARAGEM",
    "nome": "PEDAGIO",
    "valor": 1.0
  },
  {
    "tipo": "R. GARAGEM",
    "nome": "HORA PARADA",
    "valor": 60.0
  },
  {
    "tipo": "R. GARAGEM",
    "nome": "HORA TRABALHADA",
    "valor": 60.0
  },
  {
    "tipo": "R. GARAGEM",
    "nome": "KM RETORNO",
    "valor": 3.24
  },
  {
    "tipo": "R. GARAGEM",
    "nome": "KM DESLOCAMENTO",
    "valor": 3.24
  },
  {
    "tipo": "R. GARAGEM",
    "nome": "KM VIAGEM",
    "valor": 3.24
  },
  {
    "tipo": "R. GARAGEM",
    "nome": "SAIDA",
    "valor": 215.0
  },
  {
    "tipo": "R. LEVE",
    "nome": "KM VIAGEM",
    "valor": 3.24
  },
  {
    "tipo": "R. LEVE",
    "nome": "PEDAGIO",
    "valor": 1.0
  },
  {
    "tipo": "R. LEVE",
    "nome": "HORA TRABALHADA",
    "valor": 60.0
  },
  {
    "tipo": "R. LEVE",
    "nome": "SAIDA",
    "valor": 177.6
  },
  {
    "tipo": "R. LEVE",
    "nome": "HORA PARADA",
    "valor": 60.0
  },
  {
    "tipo": "R. LEVE",
    "nome": "ESTADIA",
    "valor": 50.0
  },
  {
    "tipo": "R. MEGA PESADO",
    "nome": "HORA PARADA",
    "valor": 150.0
  },
  {
    "tipo": "R. MEGA PESADO",
    "nome": "KM RETORNO",
    "valor": 7.5
  },
  {
    "tipo": "R. MEGA PESADO",
    "nome": "KM DESLOCAMENTO",
    "valor": 7.5
  },
  {
    "tipo": "R. MEGA PESADO",
    "nome": "KM VIAGEM",
    "valor": 7.5
  },
  {
    "tipo": "R. MEGA PESADO",
    "nome": "SAIDA",
    "valor": 800.0
  },
  {
    "tipo": "R. MEGA PESADO",
    "nome": "PEDAGIO",
    "valor": 1.0
  },
  {
    "tipo": "R. MEGA PESADO",
    "nome": "HORA TRABALHADA",
    "valor": 200.0
  },
  {
    "tipo": "R. MOTO",
    "nome": "SAIDA",
    "valor": 191.0
  },
  {
    "tipo": "R. MOTO",
    "nome": "KM DESLOCAMENTO",
    "valor": 3.24
  },
  {
    "tipo": "R. MOTO",
    "nome": "KM RETORNO",
    "valor": 3.24
  },
  {
    "tipo": "R. MOTO",
    "nome": "HORA PARADA",
    "valor": 60.0
  },
  {
    "tipo": "R. MOTO",
    "nome": "HORA TRABALHADA",
    "valor": 60.0
  },
  {
    "tipo": "R. MOTO",
    "nome": "PEDAGIO",
    "valor": 1.0
  },
  {
    "tipo": "R. MOTO",
    "nome": "KM VIAGEM",
    "valor": 3.24
  },
  {
    "tipo": "R. MOTO",
    "nome": "ESTADIA",
    "valor": 45.0
  },
  {
    "tipo": "R. MOTO ESP.",
    "nome": "HORA PARADA",
    "valor": 60.0
  },
  {
    "tipo": "R. MOTO ESP.",
    "nome": "SAIDA",
    "valor": 308.0
  },
  {
    "tipo": "R. MOTO ESP.",
    "nome": "PEDAGIO",
    "valor": 1.0
  },
  {
    "tipo": "R. MOTO ESP.",
    "nome": "KM VIAGEM",
    "valor": 3.24
  },
  {
    "tipo": "R. MOTO ESP.",
    "nome": "HORA TRABALHADA",
    "valor": 60.0
  },
  {
    "tipo": "R. PATINS",
    "nome": "ESTADIA",
    "valor": 45.0
  },
  {
    "tipo": "R. PATINS",
    "nome": "PEDAGIO",
    "valor": 1.0
  },
  {
    "tipo": "R. PATINS",
    "nome": "HORA TRABALHADA",
    "valor": 60.0
  },
  {
    "tipo": "R. PATINS",
    "nome": "HORA PARADA",
    "valor": 60.0
  },
  {
    "tipo": "R. PATINS",
    "nome": "KM RETORNO",
    "valor": 3.24
  },
  {
    "tipo": "R. PATINS",
    "nome": "KM DESLOCAMENTO",
    "valor": 3.24
  },
  {
    "tipo": "R. PATINS",
    "nome": "KM VIAGEM",
    "valor": 3.24
  },
  {
    "tipo": "R. PATINS",
    "nome": "SAIDA",
    "valor": 305.0
  },
  {
    "tipo": "R. PESADO",
    "nome": "PEDAGIO",
    "valor": 1.0
  },
  {
    "tipo": "R. PESADO",
    "nome": "KM DESLOCAMENTO",
    "valor": 5.17
  },
  {
    "tipo": "R. PESADO",
    "nome": "SAIDA",
    "valor": 462.84
  },
  {
    "tipo": "R. PESADO",
    "nome": "KM VIAGEM",
    "valor": 5.17
  },
  {
    "tipo": "R. PESADO",
    "nome": "HORA TRABALHADA",
    "valor": 145.0
  },
  {
    "tipo": "R. PESADO",
    "nome": "ESTADIA",
    "valor": 100.0
  },
  {
    "tipo": "R. PESADO",
    "nome": "HORA PARADA",
    "valor": 120.0
  },
  {
    "tipo": "R. UTILITARIO",
    "nome": "ESTADIA",
    "valor": 50.0
  },
  {
    "tipo": "R. UTILITARIO",
    "nome": "PEDAGIO",
    "valor": 1.0
  },
  {
    "tipo": "R. UTILITARIO",
    "nome": "SAIDA",
    "valor": 269.1
  },
  {
    "tipo": "R. UTILITARIO",
    "nome": "KM VIAGEM",
    "valor": 3.52
  },
  {
    "tipo": "R. UTILITARIO",
    "nome": "HORA PARADA",
    "valor": 60.0
  },
  {
    "tipo": "R. UTILITARIO",
    "nome": "HORA TRABALHADA",
    "valor": 70.0
  }
]

DEFAULT_FINANCE_ITEMS = [
    "SAIDA",
    "KM DESLOCAMENTO",
    "KM RETORNO",
    "KM VIAGEM",
    "PEDAGIO",
    "HORA PARADA",
    "HORA TRABALHADA",
    "ESTADIA",
]

def normalizar_tipo_importado(tipo):
    """Ajusta nomes vindos do AutEM/Excel para bater com a tabela de preços."""
    t = (tipo or "").strip().upper()
    mapa = {
        "REBOQUE LEVE": "R. LEVE",
        "R LEVE": "R. LEVE",
        "LEVE": "R. LEVE",
        "REBOQUE MOTO": "R. MOTO",
        "R MOTO": "R. MOTO",
        "MOTO": "R. MOTO",
        "REBOQUE MOTO ESPECIAL": "R. MOTO ESP.",
        "MOTO ESPECIAL": "R. MOTO ESP.",
        "REBOQUE PESADO": "R. PESADO",
        "R PESADO": "R. PESADO",
        "PESADO": "R. PESADO",
        "REBOQUE PATINS": "R. PATINS",
        "PATINS": "R. PATINS",
        "REBOQUE UTILITARIO": "R. UTILITARIO",
        "REBOQUE UTILITÁRIO": "R. UTILITARIO",
        "UTILITARIO": "R. UTILITARIO",
        "UTILITÁRIO": "R. UTILITARIO",
        "REBOQUE GARAGEM": "R. GARAGEM",
        "GARAGEM": "R. GARAGEM",
    }
    return mapa.get(t, tipo)

def moeda(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0

def fmt_moeda(v):
    try:
        return f"R$ {float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


TZ_BR = ZoneInfo("America/Sao_Paulo")


def agora_dt():
    """Datetime naive no fuso America/Sao_Paulo (horário de Brasília)."""
    return datetime.now(TZ_BR).replace(tzinfo=None)


def agora(): return agora_dt().strftime('%d/%m/%Y %H:%M:%S')


def parse_datetime_br(valor):
    """Converte dd/mm/yyyy HH:mm(:ss) — horário de Brasília, sem UTC."""
    s = (valor or "").strip()
    if not s or s == "-":
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parse_dt_valor(v):
    """Interpreta datetime do banco, ISO ou formato BR."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s or s == "-":
        return None

    dt = parse_datetime_br(s)
    if dt:
        return dt

    s_iso = s.replace("Z", "+00:00")
    if re.search(r"\+\d{2}$", s_iso):
        s_iso = s_iso + ":00"
    if " " in s_iso and "T" not in s_iso:
        s_iso = s_iso.replace(" ", "T", 1)
    try:
        return datetime.fromisoformat(s_iso)
    except ValueError:
        pass

    m = re.match(r"(\d{4}-\d{2}-\d{2})[\sT](\d{2}:\d{2})", s)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)}", "%Y-%m-%d %H:%M")
        except ValueError:
            pass
    return None


def _dt_para_local_br(dt):
    if dt is None:
        return None
    if dt.tzinfo is not None:
        return dt.astimezone(TZ_BR).replace(tzinfo=None)
    return dt


def formatar_data_hora_central(v):
    """Coluna Data e Hora da Central: dd/mm/yyyy HH:mm (sem segundos, sem +00)."""
    dt = _parse_dt_valor(v)
    if not dt:
        return "-"
    dt = _dt_para_local_br(dt)
    return dt.strftime("%d/%m/%Y %H:%M")


def dt_str(v):
    """Formata timestamp para exibição geral."""
    dt = _parse_dt_valor(v)
    if not dt:
        return "-"
    dt = _dt_para_local_br(dt)
    return dt.strftime("%d/%m/%Y %H:%M:%S")
def get_conn():
    if not DATABASE_URL:
        if _is_render_deploy():
            raise RuntimeError("DATABASE_URL não configurada no Render.")
        raise RuntimeError(
            "Banco local não configurado. Defina DATABASE_URL ou LOCAL_DATABASE_URL no .env "
            f"ou crie {_LOCAL_DB_FILE} com a URL do PostgreSQL."
        )
    try:
        return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    except psycopg2.OperationalError as exc:
        if _is_render_deploy():
            raise
        raise RuntimeError(
            "Não foi possível conectar ao PostgreSQL. "
            "Restaure DATABASE_URL no .env (foi removida ao adicionar GOOGLE_MAPS_API_KEY) "
            f"ou verifique se o servidor está ativo. Detalhe: {exc}"
        ) from exc
def q(sql, params=None, fetch=False):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params or ())
            rows = cur.fetchall() if fetch else None
        conn.commit()
    return rows
def one(sql, params=None):
    rows=q(sql, params, True)
    return rows[0] if rows else None


def _backup_controle_multas_frota_if_needed():
    """Backup opcional de multas — não chamar no startup."""
    marker = Path("data/backup/.multas_frota_v1")
    if marker.exists():
        return
    try:
        Path("data/backup").mkdir(parents=True, exist_ok=True)
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SET LOCAL statement_timeout = '5s'")
                cur.execute(
                    """
                    select id, viatura, placa, auto_infracao, data_infracao, valor, vencimento, created_at
                      from controle_multas
                     order by created_at asc nulls last
                     limit 5000
                    """
                )
                rows = cur.fetchall() or []
            conn.commit()
        payload = [dict(r) for r in rows]
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = Path(f"data/backup/multas_backup_{ts}.json")
        out.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        marker.write_text(ts, encoding="utf-8")
    except Exception as exc:
        print(f"[warn] backup controle_multas ignorado: {exc}")


def init_db():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute('create extension if not exists "uuid-ossp";')
            cur.execute("""
            create table if not exists motoristas (
              id uuid primary key default uuid_generate_v4(),
              nome text, telefone text, veiculo text, placa text, tipo text,
              online boolean default false, lat double precision, lng double precision,
              ultima_atualizacao timestamp, ativo boolean default true,
              created_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists servicos (
              id uuid primary key default uuid_generate_v4(),
              protocolo text, seguradora text, tipo text, origem text, destino text,
              observacao text, status text default 'novo', motorista_id uuid,
              motorista_nome text, placa_removida text, placa_veiculo_removido text,
              ultimo_evento text, historico jsonb default '[]'::jsonb,
              criado_em timestamp default now(), atualizado_em timestamp default now(),
              finalizado_em timestamp, created_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists fotos (
              id uuid primary key default uuid_generate_v4(),
              servico_id uuid, url text, filename text, created_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists checklist_avarias (
            id uuid primary key default uuid_generate_v4(),
            servico_id uuid not null,
            parte text,
            marcacoes jsonb default '[]'::jsonb,
            fotos jsonb default '[]'::jsonb,
            created_at timestamp default now()
        );
          """)
            cur.execute("""
            create table if not exists checklist_assinaturas (
              id uuid primary key default uuid_generate_v4(),
              servico_id uuid not null unique,
              assinatura_origem_base64 text,
              assinatura_destino_base64 text,
              assinatura_base64 text,
              origem_atualizada_em timestamp,
              destino_atualizada_em timestamp,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists tabela_precos (
              id uuid primary key default uuid_generate_v4(),
              tipo_servico text not null,
              nome_item text not null,
              valor_padrao numeric(12,2) default 0,
              ativo boolean default true,
              created_at timestamp default now(),
              unique(tipo_servico, nome_item)
            );""")
            cur.execute("""
            create table if not exists servico_itens (
              id uuid primary key default uuid_generate_v4(),
              servico_id uuid not null,
              tipo_servico text,
              nome_item text not null,
              quantidade numeric(12,2) default 0,
              valor_unitario numeric(12,2) default 0,
              valor_total numeric(12,2) default 0,
              observacao text,
              created_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists veiculos (
              id uuid primary key default uuid_generate_v4(),
              placa text, modelo text, tipo text, ano text, renavam text, observacao text,
              ativo boolean default true,
              created_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists usuarios_sistema (
              id uuid primary key default uuid_generate_v4(),
              nome text, cpf text, telefone text, email text, senha text, perfil text,
              ativo boolean default true,
              created_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists comentarios_servico (
              id uuid primary key default uuid_generate_v4(),
              servico_id uuid not null,
              texto text not null,
              criado_em timestamp default now()
            );""")
            cur.execute("""
            create table if not exists servico_mensagens_mobile (
              id uuid primary key default uuid_generate_v4(),
              servico_id uuid not null,
              remetente_tipo text not null,
              remetente_id text,
              remetente_nome text,
              mensagem text not null,
              lida_central boolean default false,
              lida_motorista boolean default false,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create index if not exists idx_servico_mensagens_mobile_servico
              on servico_mensagens_mobile (servico_id, created_at);
            """)
            cur.execute("""
            create table if not exists central_alertas_operacionais (
              id uuid primary key default uuid_generate_v4(),
              servico_id uuid not null,
              tipo text not null,
              titulo text,
              descricao text,
              visto boolean default false,
              created_at timestamp default now(),
              visto_em timestamp
            );""")
            cur.execute("""
            create index if not exists idx_central_alertas_servico
              on central_alertas_operacionais (servico_id, tipo, visto);
            """)
            cur.execute("""
            create table if not exists controle_danos (
              id uuid primary key default uuid_generate_v4(),
              data_dano date,
              profissional text,
              viatura text,
              origem text,
              tipo_dano text,
              objeto text,
              valor numeric(12,2) default 0,
              situacao text,
              status text default 'ativo',
              observacoes text,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists controle_abastecimentos (
              id uuid primary key default uuid_generate_v4(),
              data_abastecimento date,
              viatura text,
              posto text,
              combustivel text,
              litros numeric(12,3) default 0,
              valor_unitario numeric(12,2) default 0,
              valor_total numeric(12,2) default 0,
              hodometro numeric(12,0),
              profissional text,
              status text default 'ativo',
              observacoes text,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists controle_checklists_viatura (
              id uuid primary key default uuid_generate_v4(),
              data_checklist date,
              viatura text,
              motorista text,
              tipo_checklist text,
              km numeric(12,0),
              status_checklist text,
              status text default 'ativo',
              observacoes text,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists controle_manutencoes (
              id uuid primary key default uuid_generate_v4(),
              data_manutencao date,
              documento text,
              fornecedor text,
              viatura text,
              hodometro numeric(12,0),
              total numeric(12,2) default 0,
              status text default 'ativo',
              observacoes text,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists controle_multas (
              id uuid primary key default uuid_generate_v4(),
              data_infracao date,
              viatura text,
              auto_infracao text,
              municipio text,
              condutor text,
              valor numeric(12,2) default 0,
              vencimento date,
              situacao text,
              status text default 'ativo',
              observacoes text,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists controle_seguros (
              id uuid primary key default uuid_generate_v4(),
              viatura text,
              seguradora text,
              apolice text,
              vigencia_inicial date,
              vigencia_final date,
              valor numeric(12,2) default 0,
              parcelas integer default 1,
              situacao text,
              status text default 'ativo',
              observacoes text,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists controle_patio (
              id uuid primary key default uuid_generate_v4(),
              data_entrada date,
              hora_chegada time,
              data_hora_chegada timestamp,
              motorista text,
              placa text not null,
              modelo text,
              cor text,
              seguradora text,
              data_saida date,
              hora_saida time,
              data_hora_saida timestamp,
              responsavel_entrada text,
              responsavel_saida text,
              observacao text,
              observacao_saida text,
              status text default 'no_patio',
              tempo_no_patio_minutos integer default 0,
              anexos jsonb default '[]'::jsonb,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            alters=[
            "alter table motoristas add column if not exists tipo text;",
            "alter table motoristas add column if not exists cpf text;",
            "alter table motoristas add column if not exists cnh text;",
            "alter table motoristas add column if not exists vencimento_cnh text;",
            "alter table motoristas add column if not exists nascimento text;",
            "alter table motoristas add column if not exists estado_civil text;",
            "alter table motoristas add column if not exists endereco text;",
            "alter table motoristas add column if not exists online boolean default false;",
            "alter table motoristas add column if not exists lat double precision;",
            "alter table motoristas add column if not exists lng double precision;",
            "alter table motoristas add column if not exists ultima_atualizacao timestamp;",
            "alter table motoristas add column if not exists ativo boolean default true;",
            "alter table motoristas add column if not exists login text;",
            "alter table motoristas add column if not exists senha text;",
            "alter table motoristas add column if not exists placa_atual text;",
            "alter table motoristas add column if not exists ultimo_login timestamp;",
            "alter table motoristas add column if not exists push_token text;",
            "alter table motoristas add column if not exists push_platform text;",
            "alter table motoristas add column if not exists fcm_token text;",

            "alter table servicos add column if not exists observacao text;",
            "alter table servicos add column if not exists motorista_nome text;",
            "alter table servicos add column if not exists placa_removida text;",
            "alter table servicos add column if not exists placa_veiculo_removido text;",
            "alter table servicos add column if not exists ultimo_evento text;",
            "alter table servicos add column if not exists historico jsonb default '[]'::jsonb;",
            "alter table servicos add column if not exists criado_em timestamp default now();",
            "alter table servicos add column if not exists atualizado_em timestamp default now();",
            "alter table servicos add column if not exists finalizado_em timestamp;",
            "alter table fotos add column if not exists filename text;",
            "alter table servicos add column if not exists status_faturamento text default 'para_conferir';",
            "alter table servicos add column if not exists valor_total numeric(12,2) default 0;",
            "alter table servicos add column if not exists beneficiario text;",
            "alter table servicos add column if not exists telefone_cliente text;",
            "alter table servicos add column if not exists veiculo_cliente text;",
            "alter table servicos add column if not exists cor_cliente text;",
            "alter table servicos add column if not exists cnpj_cliente text;",
            "alter table servicos add column if not exists referencia_origem text;",
            "alter table servicos add column if not exists referencia_destino text;",
            "alter table servicos add column if not exists solicitante text;",
            "alter table servicos add column if not exists problema text;",
            "alter table servicos add column if not exists fonte_importacao text;",
            "alter table servicos add column if not exists origem_lat double precision;",
            "alter table servicos add column if not exists origem_lng double precision;",
            "alter table checklist_assinaturas add column if not exists assinatura_origem_base64 text;",
            "alter table checklist_assinaturas add column if not exists assinatura_destino_base64 text;",
            "alter table checklist_assinaturas add column if not exists origem_atualizada_em timestamp;",
            "alter table checklist_assinaturas add column if not exists destino_atualizada_em timestamp;",
            "alter table checklist_assinaturas alter column assinatura_base64 drop not null;",
            "alter table checklist_avarias add column if not exists observacao text;"]
            controle_alters = [
                "alter table controle_danos add column if not exists data_hora_dano timestamp;",
                "alter table controle_danos add column if not exists aviso text;",
                "alter table controle_danos add column if not exists assistencia text;",
                "alter table controle_danos add column if not exists identificacao_objeto text;",
                "alter table controle_danos add column if not exists parcelas integer default 1;",
                "alter table controle_danos add column if not exists data_desconto date;",
                "alter table controle_danos add column if not exists descricao_dano text;",
                "alter table controle_danos add column if not exists analise_interna text;",
                "alter table controle_abastecimentos add column if not exists data_hora timestamp;",
                "alter table controle_abastecimentos add column if not exists desconsiderar_ultimo_km boolean default false;",
                "alter table controle_abastecimentos add column if not exists gerar_contas_pagar boolean default false;",
                "alter table controle_abastecimentos add column if not exists viatura_id uuid;",
                "alter table controle_abastecimentos add column if not exists profissional_id uuid;",
                "alter table controle_abastecimentos add column if not exists placa text;",
                "alter table controle_abastecimentos add column if not exists km_rodado numeric(12,0);",
                "alter table controle_abastecimentos add column if not exists media_km_litro numeric(12,3);",
                "alter table controle_abastecimentos add column if not exists contas_pagar_id uuid;",
                "alter table controle_abastecimentos add column if not exists alerta_hodometro text;",
                "alter table controle_manutencoes add column if not exists itens jsonb default '[]'::jsonb;",
                "alter table controle_manutencoes add column if not exists total_produtos numeric(12,2) default 0;",
                "alter table controle_manutencoes add column if not exists total_servicos numeric(12,2) default 0;",
                "alter table controle_manutencoes add column if not exists viatura_id uuid;",
                "alter table controle_manutencoes add column if not exists placa text;",
                "alter table controle_manutencoes add column if not exists fornecedor_id uuid;",
                "alter table controle_manutencoes add column if not exists contas_pagar_id uuid;",
                "alter table controle_manutencoes add column if not exists gerar_contas_pagar boolean default false;",
                "alter table controle_manutencoes add column if not exists alerta_hodometro text;",
                "alter table controle_manutencoes add column if not exists hodometro_menor_ultimo boolean default false;",
                "alter table controle_multas add column if not exists data_hora_infracao timestamp;",
                "alter table controle_multas add column if not exists data_limite_indicacao date;",
                "alter table controle_multas add column if not exists endereco text;",
                "alter table controle_multas add column if not exists descricao_multa text;",
                "alter table controle_multas add column if not exists natureza text;",
                "alter table controle_multas add column if not exists condutor_responsavel text;",
                "alter table controle_multas add column if not exists parcelas integer default 1;",
                "alter table controle_multas add column if not exists valor_pago numeric(12,2) default 0;",
                "alter table controle_multas add column if not exists contas_pagar boolean default false;",
                "alter table controle_multas add column if not exists extrato_profissional boolean default false;",
                "alter table controle_multas add column if not exists viatura_id uuid;",
                "alter table controle_multas add column if not exists placa text;",
                "alter table controle_multas add column if not exists condutor_id uuid;",
                "alter table controle_multas add column if not exists condutor_responsavel_id uuid;",
                "alter table controle_multas add column if not exists valor_aberto numeric(12,2) default 0;",
                "alter table controle_multas add column if not exists gerar_contas_pagar boolean default false;",
                "alter table controle_multas add column if not exists gerar_extrato_profissional boolean default false;",
                "alter table controle_multas add column if not exists contas_pagar_id uuid;",
                "alter table controle_multas add column if not exists contas_pagar_ids jsonb default '[]'::jsonb;",
                "alter table controle_multas add column if not exists extrato_profissional_id text;",
                "alter table controle_seguros add column if not exists corretor text;",
                "alter table controle_seguros add column if not exists telefone text;",
                "alter table controle_seguros add column if not exists vencimento date;",
                "alter table controle_checklists_viatura add column if not exists data_hora timestamp;",
                "alter table controle_checklists_viatura add column if not exists itens_conferidos text;",
                "alter table controle_checklists_viatura add column if not exists itens_problema text;",
                "alter table controle_patio add column if not exists servico_id uuid;",
            ]
            for a in alters:
                cur.execute(a)
            for a in controle_alters:
                cur.execute(a)
            viaturas_alters = [
                "alter table cadastro_viaturas add column if not exists classificacao_original text;",
                "alter table cadastro_viaturas add column if not exists origem text default 'propria';",
            ]
            for a in viaturas_alters:
                cur.execute(a)
            cur.execute("""
            create table if not exists cadastro_viaturas (
              id uuid primary key default uuid_generate_v4(),
              placa text,
              marca text,
              modelo text,
              renavam text,
              chassi text,
              ano_fabricacao text,
              ano_modelo text,
              combustivel text,
              capacidade_litros numeric(12,2),
              cor text,
              estado_placa text,
              tipo_viatura text,
              observacoes text,
              status text default 'ativo',
              cpf_cnpj_crlv text,
              exibicao text,
              telefone text,
              personalizacao text,
              consumo_km_l numeric(8,2),
              hodometro numeric(12,0),
              terceiro boolean default false,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists cadastro_viatura_arquivos (
              id uuid primary key default uuid_generate_v4(),
              viatura_id uuid not null,
              nome text,
              tipo_documento text,
              data_documento date,
              filename text,
              url text,
              created_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists cadastro_profissionais (
              id uuid primary key default uuid_generate_v4(),
              nome text,
              cpf text,
              rg text,
              cnh text,
              categoria_cnh text,
              validade_cnh date,
              telefone text,
              email text,
              endereco text,
              funcao text,
              observacoes text,
              status text default 'ativo',
              tipo_profissional text,
              comissao_padrao numeric(12,2) default 0,
              pode_receber_servicos boolean default true,
              pode_aparecer_controle boolean default true,
              motorista_id uuid,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists cadastro_profissional_arquivos (
              id uuid primary key default uuid_generate_v4(),
              profissional_id uuid not null,
              nome text,
              tipo_documento text,
              data_documento date,
              filename text,
              url text,
              created_at timestamp default now()
            );""")
            profissionais_alters = [
                "alter table cadastro_profissionais add column if not exists filial_cnpj text;",
                "alter table cadastro_profissionais add column if not exists nome_completo text;",
                "alter table cadastro_profissionais add column if not exists nome_trabalho text;",
                "alter table cadastro_profissionais add column if not exists data_nascimento date;",
                "alter table cadastro_profissionais add column if not exists telefone_fixo text;",
                "alter table cadastro_profissionais add column if not exists telefone_movel text;",
                "alter table cadastro_profissionais add column if not exists estado_civil text;",
                "alter table cadastro_profissionais add column if not exists cep text;",
                "alter table cadastro_profissionais add column if not exists logradouro text;",
                "alter table cadastro_profissionais add column if not exists bairro text;",
                "alter table cadastro_profissionais add column if not exists cidade text;",
                "alter table cadastro_profissionais add column if not exists uf text;",
                "alter table cadastro_profissionais add column if not exists terceiro boolean default false;",
                "alter table cadastro_profissionais add column if not exists cnpj text;",
                "alter table cadastro_profissionais add column if not exists remuneracao text;",
                "alter table cadastro_profissionais add column if not exists forma_pagamento text;",
                "alter table cadastro_profissionais add column if not exists data_admissao date;",
                "alter table cadastro_profissionais add column if not exists data_demissao date;",
                "alter table cadastro_profissionais add column if not exists hora_inicio text;",
                "alter table cadastro_profissionais add column if not exists hora_termino text;",
                "alter table cadastro_profissionais add column if not exists carga_horaria text;",
                "alter table cadastro_profissionais add column if not exists intervalo text;",
                "alter table cadastro_profissionais add column if not exists escala text;",
                "alter table cadastro_profissionais add column if not exists registro_ctps text;",
                "alter table cadastro_profissionais add column if not exists cnh_numero text;",
                "alter table cadastro_profissionais add column if not exists cnh_vencimento date;",
                "alter table cadastro_profissionais add column if not exists cnh_categoria text;",
                "alter table cadastro_profissional_arquivos add column if not exists nome_arquivo text;",
                "alter table cadastro_profissional_arquivos add column if not exists extensao text;",
                "alter table cadastro_profissional_arquivos add column if not exists tipo text;",
                "alter table cadastro_profissional_arquivos add column if not exists caminho_arquivo text;",
                "alter table cadastro_profissionais add column if not exists classificacao_vinculo text;",
            ]
            for a in profissionais_alters:
                cur.execute(a)
            cur.execute("""
                update cadastro_profissionais
                   set classificacao_vinculo = case when coalesce(terceiro, false) then 'terceiro' else 'contratado' end
                 where coalesce(trim(classificacao_vinculo), '') = ''
            """)
            cur.execute("""
            create table if not exists cadastro_contatos (
              id uuid primary key default uuid_generate_v4(),
              tipo_pessoa text default 'juridica',
              status text default 'ativo',
              razao_social text,
              nome_fantasia text,
              cnpj text,
              inscricao_estadual text,
              contribuinte_icms text,
              nome text,
              cpf text,
              rg text,
              data_nascimento date,
              cep text,
              logradouro text,
              numero text,
              complemento text,
              bairro text,
              cidade text,
              uf text,
              email text,
              email_financeiro text,
              telefone text,
              celular text,
              cliente boolean default false,
              fornecedor boolean default false,
              limite_credito numeric(14,2) default 0,
              prazo_recebimento text,
              observacoes_comercial_cliente text,
              prazo_pagamento text,
              categoria_fornecedor text,
              observacoes_comercial_fornecedor text,
              observacoes text,
              created_at timestamp default now(),
              updated_at timestamp default now()
            );""")
            cur.execute("""
            create table if not exists cadastro_contato_arquivos (
              id uuid primary key default uuid_generate_v4(),
              contato_id uuid not null,
              nome_arquivo text,
              extensao text,
              tipo text,
              caminho_arquivo text,
              nome text,
              tipo_documento text,
              data_documento date,
              filename text,
              url text,
              created_at timestamp default now()
            );""")
            contatos_alters = [
                "alter table cadastro_contatos add column if not exists codigo_cliente text;",
                "alter table cadastro_contatos add column if not exists fax text;",
                "alter table cadastro_contatos add column if not exists website text;",
                "alter table cadastro_contatos add column if not exists email_nfse text;",
                "alter table cadastro_contatos add column if not exists segmento text;",
                "alter table cadastro_contatos add column if not exists vendedor text;",
                "alter table cadastro_contatos add column if not exists condicao_pagamento text;",
                "alter table cadastro_contatos add column if not exists regime_tributario text;",
                "alter table cadastro_contatos add column if not exists cliente_desde date;",
            ]
            for a in contatos_alters:
                cur.execute(a)
            cur.execute("""
                update cadastro_profissionais set
                  nome_completo = coalesce(nullif(trim(nome_completo),''), nome),
                  nome_trabalho = coalesce(nullif(trim(nome_trabalho),''), nome),
                  telefone_movel = coalesce(nullif(trim(telefone_movel),''), telefone),
                  cnh_numero = coalesce(nullif(trim(cnh_numero),''), cnh),
                  cnh_vencimento = coalesce(cnh_vencimento, validade_cnh),
                  cnh_categoria = coalesce(nullif(trim(cnh_categoria),''), categoria_cnh)
            """)
            cur.execute("""
                update cadastro_profissional_arquivos set
                  nome_arquivo = coalesce(nome_arquivo, nome),
                  tipo = coalesce(tipo, tipo_documento),
                  caminho_arquivo = coalesce(caminho_arquivo, url),
                  extensao = coalesce(extensao, lower(regexp_replace(coalesce(filename,''), '^.*\\.', '')))
                where nome_arquivo is null or caminho_arquivo is null
            """)
            cur.execute("""
                insert into cadastro_viaturas (placa, modelo, tipo_viatura, status, exibicao, renavam, observacoes)
                select upper(trim(placa)), modelo, tipo,
                       case when coalesce(ativo,true) then 'ativo' else 'inativo' end,
                       coalesce(nullif(trim(placa),''), modelo, tipo), renavam, observacao
                  from veiculos v
                 where not exists (
                       select 1 from cadastro_viaturas c
                        where upper(trim(coalesce(c.placa,''))) = upper(trim(coalesce(v.placa,'')))
                          and coalesce(trim(c.placa),'') <> ''
                 )
                   and (coalesce(trim(v.placa),'') <> '' or coalesce(trim(v.modelo),'') <> '')
            """)
            cur.execute("""
                insert into cadastro_profissionais (
                  nome, telefone, funcao, status, motorista_id,
                  pode_receber_servicos, pode_aparecer_controle, cpf, cnh
                )
                select nome, telefone, coalesce(tipo,'Motorista'),
                       case when coalesce(ativo,true) then 'ativo' else 'inativo' end,
                       id, true, true, cpf, cnh
                  from motoristas m
                 where coalesce(trim(nome),'') <> ''
                   and not exists (
                     select 1 from cadastro_profissionais p
                      where p.motorista_id = m.id
                   )
            """)
            cur.execute("""
                update checklist_assinaturas
                   set assinatura_origem_base64 = assinatura_base64,
                       origem_atualizada_em = coalesce(origem_atualizada_em, updated_at, created_at, now())
                 where coalesce(assinatura_origem_base64, '') = ''
                   and coalesce(assinatura_base64, '') <> ''
            """)
            import financeiro_routes
            financeiro_routes.init_financeiro_tables(cur)
            import financeiro_notas_entrada
            financeiro_notas_entrada.init_notas_entrada_tables(cur)
            import financeiro_nfse
            financeiro_nfse.init_nfse_tables(cur)
            import financeiro_config_fiscal
            financeiro_config_fiscal.init_config_fiscal_tables(cur)
            cur.execute("select count(*) as total from tabela_precos")
            total_precos = cur.fetchone()["total"]
            if not total_precos:
                for item in PRICE_DATA:
                    cur.execute(
                        "insert into tabela_precos (tipo_servico,nome_item,valor_padrao,ativo) values (%s,%s,%s,true) on conflict (tipo_servico,nome_item) do update set valor_padrao=excluded.valor_padrao, ativo=true",
                        (item["tipo"], item["nome"], item["valor"])
                    )
        conn.commit()

@app.on_event("startup")
def startup_event(): init_db()


def lista_tipos_servico():
    rows = q("select tipo_servico from tabela_precos where coalesce(ativo,true)=true group by tipo_servico order by tipo_servico", fetch=True)
    return [r["tipo_servico"] for r in rows]

def itens_padrao_tipo(tipo_servico):
    tipo_servico = normalizar_tipo_importado(tipo_servico)
    if not tipo_servico:
        return []
    rows = q("select nome_item, valor_padrao from tabela_precos where tipo_servico=%s and coalesce(ativo,true)=true order by nome_item", (tipo_servico,), True)
    return [{"nome_item": r["nome_item"], "valor_padrao": float(r["valor_padrao"] or 0)} for r in rows]

def itens_do_servico(servico_id):
    rows = q("select * from servico_itens where servico_id=%s order by nome_item", (str(servico_id),), True)
    out = []
    for r in rows:
        r = dict(r)
        r["id"] = str(r["id"])
        r["servico_id"] = str(r["servico_id"])
        r["quantidade"] = float(r.get("quantidade") or 0)
        r["valor_unitario"] = float(r.get("valor_unitario") or 0)
        r["valor_total"] = float(r.get("valor_total") or 0)
        r["valor_unitario_fmt"] = fmt_moeda(r["valor_unitario"])
        r["valor_total_fmt"] = fmt_moeda(r["valor_total"])
        return_dummy = None
        out.append(r)
    return out

def atualizar_total_servico(servico_id):
    row = one("select coalesce(sum(valor_total),0) as total from servico_itens where servico_id=%s", (str(servico_id),))
    total = float(row["total"] or 0) if row else 0
    q("update servicos set valor_total=%s, atualizado_em=now() where id=%s", (total, str(servico_id)))
    return total

def criar_itens_para_servico(servico_id, tipo_servico, nomes=None, qtds=None, valores=None):
    tipo_servico = normalizar_tipo_importado(tipo_servico)
    nomes = nomes or []
    qtds = qtds or []
    valores = valores or []
    if not nomes:
        padrao = itens_padrao_tipo(tipo_servico)
        # Se o tipo importado não existir na tabela, cria os itens básicos zerados
        # para o faturamento conseguir editar manualmente.
        if not padrao:
            padrao = [{"nome_item": nome, "valor_padrao": 0} for nome in DEFAULT_FINANCE_ITEMS]
        nomes = [i["nome_item"] for i in padrao]
        valores = [i["valor_padrao"] for i in padrao]
        qtds = [1 if str(i["nome_item"]).upper() in ["SAIDA", "SAÍDA"] else 0 for i in padrao]
    q("delete from servico_itens where servico_id=%s", (str(servico_id),))
    for idx, nome in enumerate(nomes):
        nome = (nome or "").strip()
        if not nome:
            continue
        qtd = moeda(qtds[idx] if idx < len(qtds) else 0)
        valor = moeda(valores[idx] if idx < len(valores) else 0)
        total = qtd * valor
        q(
            "insert into servico_itens (servico_id,tipo_servico,nome_item,quantidade,valor_unitario,valor_total) values (%s,%s,%s,%s,%s,%s)",
            (str(servico_id), tipo_servico, nome, qtd, valor, total)
        )
    return atualizar_total_servico(servico_id)

def normalizar_motorista(m):
    if not m: return None
    m=dict(m); m["id"]=str(m["id"]); m["placa_atual"]=m.get("placa_atual") or m.get("placa"); m["ultima_atualizacao"]=dt_str(m.get("ultima_atualizacao")); m["created_at"]=dt_str(m.get("created_at")); return m
def fotos_do_servico(servico_id):
    return [dict(r) for r in q("select url, filename, created_at from fotos where servico_id=%s order by created_at asc",(str(servico_id),), True)]
def normalizar_servico(s, incluir_fotos=True):
    if not s: return None
    s=dict(s); s["id"]=str(s["id"])
    if s.get("motorista_id"): s["motorista_id"]=str(s["motorista_id"])
    if not s.get("placa_veiculo_removido") and s.get("placa_removida"): s["placa_veiculo_removido"]=s.get("placa_removida")
    if not s.get("placa_removida") and s.get("placa_veiculo_removido"): s["placa_removida"]=s.get("placa_veiculo_removido")
    for k in ["criado_em","atualizado_em","finalizado_em","created_at"]: s[k]=dt_str(s.get(k))
    h=s.get("historico") or []
    if isinstance(h,str):
        try: h=json.loads(h)
        except Exception: h=[]
    s["historico"]=h
    if incluir_fotos: s["fotos"]=fotos_do_servico(s["id"])
    s["valor_total"] = float(s.get("valor_total") or 0)
    s["valor_total_fmt"] = fmt_moeda(s["valor_total"])
    s["status_faturamento"] = s.get("status_faturamento") or "para_conferir"
    return s

PLACA_CLIENTE_RE = re.compile(r"^[A-Z]{3}(\d{4}|\d[A-Z]\d{2})$")


def normalizar_placa_cliente(valor):
    return (valor or "").upper().replace("-", "").replace(" ", "").strip()


def placa_cliente_valida(placa):
    placa = normalizar_placa_cliente(placa)
    return bool(placa) and bool(PLACA_CLIENTE_RE.match(placa))


def placa_cliente_do_servico(servico):
    if not servico:
        return ""
    return normalizar_placa_cliente(
        servico.get("placa_veiculo_removido") or servico.get("placa_removida") or ""
    )


def salvar_placa_cliente_servico(sid, placa_veiculo):
    placa = normalizar_placa_cliente(placa_veiculo)
    if not placa:
        return False, "Informe a placa"
    if not placa_cliente_valida(placa):
        return (
            False,
            "Placa inválida. Use o formato antigo (ABC1234) ou Mercosul (ABC1D23).",
        )

    existe = servico_by_id(sid)
    if not existe:
        return False, "Serviço não encontrado"

    status_atual = (existe.get("status") or "").strip().lower()
    if status_atual == "finalizado":
        return False, "Não é possível alterar a placa de um serviço finalizado."

    novo_status = existe.get("status") or "novo"
    if novo_status not in ["em transporte", "finalizado"]:
        novo_status = "na origem"

    q(
        """
        update servicos
           set placa_veiculo_removido=%s,
               placa_removida=%s,
               status=%s,
               atualizado_em=now()
         where id=%s
        """,
        (placa, placa, novo_status, str(sid)),
    )
    registrar_evento_db(sid, "placa lançada", f"Placa: {placa}")
    s_atual = servico_by_id(sid)
    protocolo = (s_atual.get("protocolo") or "").strip() if s_atual else ""
    registrar_alerta_operacional(
        sid,
        "placa",
        "Placa enviada",
        f"Motorista enviou placa no protocolo {protocolo}" if protocolo else f"Motorista enviou placa {placa}",
    )
    return True, None


def desmontar_endereco_servico(endereco):
    """Extrai local, bairro, cidade e UF de endereço salvo (formato local - bairro - cidade/UF)."""
    out = {"local": "", "bairro": "", "cidade": "", "uf": ""}
    if not endereco or str(endereco).strip() in ("-", ""):
        return out
    texto = str(endereco).strip()
    partes = [p.strip() for p in texto.split(" - ") if p.strip()]
    if partes:
        out["local"] = partes[0]
    if len(partes) >= 3:
        out["bairro"] = partes[1]
        ult = partes[2]
        if "/" in ult:
            cidade, uf = ult.split("/", 1)
            out["cidade"] = cidade.strip()
            out["uf"] = uf.strip()
        else:
            out["cidade"] = ult
    else:
        bairro, cidade = extrair_bairro_cidade(texto)
        if bairro and bairro != "-":
            out["bairro"] = bairro
        if cidade and cidade != "-":
            out["cidade"] = cidade
    return out


def listar_comentarios_servico(servico_id):
    rows = q(
        """
        select id, texto, criado_em
          from comentarios_servico
         where servico_id = %s
         order by criado_em desc
        """,
        (str(servico_id),),
        True,
    )
    return [
        {
            "id": str(r["id"]),
            "texto": r["texto"],
            "criado_em": formatar_data_hora_central(r.get("criado_em")),
        }
        for r in (rows or [])
    ]


MOBILE_MSG_MAX_LEN = 1000


def sanitizar_mensagem_mobile(texto):
    t = (texto or "").strip()
    if not t:
        return ""
    t = html.escape(t)
    if len(t) > MOBILE_MSG_MAX_LEN:
        t = t[:MOBILE_MSG_MAX_LEN]
    return t


def _serializar_mensagem_mobile(row):
    return {
        "id": str(row["id"]),
        "servico_id": str(row["servico_id"]),
        "remetente_tipo": row.get("remetente_tipo") or "",
        "remetente_id": row.get("remetente_id") or "",
        "remetente_nome": row.get("remetente_nome") or "",
        "mensagem": row.get("mensagem") or "",
        "lida_central": bool(row.get("lida_central")),
        "lida_motorista": bool(row.get("lida_motorista")),
        "created_at": formatar_data_hora_central(row.get("created_at")),
        "created_at_iso": row.get("created_at").isoformat() if row.get("created_at") else "",
    }


def listar_mensagens_mobile_servico(servico_id):
    rows = q(
        """
        select id, servico_id, remetente_tipo, remetente_id, remetente_nome,
               mensagem, lida_central, lida_motorista, created_at, updated_at
          from servico_mensagens_mobile
         where servico_id = %s
         order by created_at asc
        """,
        (str(servico_id),),
        True,
    )
    return [_serializar_mensagem_mobile(r) for r in (rows or [])]


def contar_mensagens_mobile_nao_lidas(servico_id, leitor):
    leitor = (leitor or "").strip().lower()
    if leitor == "central":
        row = one(
            """
            select count(*)::int as qtd
              from servico_mensagens_mobile
             where servico_id = %s
               and remetente_tipo = 'motorista'
               and lida_central = false
            """,
            (str(servico_id),),
        )
    elif leitor == "motorista":
        row = one(
            """
            select count(*)::int as qtd
              from servico_mensagens_mobile
             where servico_id = %s
               and remetente_tipo = 'central'
               and lida_motorista = false
            """,
            (str(servico_id),),
        )
    else:
        return 0
    return int((row or {}).get("qtd") or 0)


def marcar_mensagens_mobile_lidas(servico_id, leitor):
    leitor = (leitor or "").strip().lower()
    if leitor == "central":
        q(
            """
            update servico_mensagens_mobile
               set lida_central = true, updated_at = now()
             where servico_id = %s
               and remetente_tipo = 'motorista'
               and lida_central = false
            """,
            (str(servico_id),),
        )
    elif leitor == "motorista":
        q(
            """
            update servico_mensagens_mobile
               set lida_motorista = true, updated_at = now()
             where servico_id = %s
               and remetente_tipo = 'central'
               and lida_motorista = false
            """,
            (str(servico_id),),
        )


def criar_mensagem_mobile_servico(servico_id, remetente_tipo, remetente_id, remetente_nome, mensagem):
    sid = str(servico_id or "").strip()
    tipo = (remetente_tipo or "").strip().lower()
    texto = sanitizar_mensagem_mobile(mensagem)
    if not sid or tipo not in ("central", "motorista") or not texto:
        return None, "Dados inválidos ou mensagem vazia."

    s = servico_by_id(sid)
    if not s:
        return None, "Serviço não encontrado."

    nome = (remetente_nome or "").strip()
    if not nome:
        nome = "Central / Operação" if tipo == "central" else (s.get("motorista_nome") or "Motorista")

    lida_central = tipo == "central"
    lida_motorista = tipo == "motorista"

    row = one(
        """
        insert into servico_mensagens_mobile (
          servico_id, remetente_tipo, remetente_id, remetente_nome, mensagem,
          lida_central, lida_motorista, created_at, updated_at
        ) values (%s, %s, %s, %s, %s, %s, %s, now(), now())
        returning id, servico_id, remetente_tipo, remetente_id, remetente_nome,
                  mensagem, lida_central, lida_motorista, created_at, updated_at
        """,
        (
            sid,
            tipo,
            (remetente_id or "").strip() or None,
            nome,
            texto,
            lida_central,
            lida_motorista,
        ),
    )
    msg = _serializar_mensagem_mobile(row) if row else None

    if tipo == "motorista":
        print(f"[CENTRAL_ALERTAS] mensagem motorista servico_id={sid}")

    if tipo == "central":
        motorista_id = s.get("motorista_id")
        if motorista_id:
            try:
                enviar_push_expo_mensagem_central(str(motorista_id), s, texto)
            except Exception as exc:
                print(f"[PUSH] mobile mensagem: {exc}")

    return msg, None


def resumo_mensagens_mobile_central():
    rows = q(
        """
        select m.servico_id, s.protocolo, count(*)::int as qtd, max(m.created_at) as ultima_em
          from servico_mensagens_mobile m
          join servicos s on s.id = m.servico_id
         where m.remetente_tipo = 'motorista'
           and m.lida_central = false
         group by m.servico_id, s.protocolo
         order by max(m.created_at) desc
        """,
        fetch=True,
    )
    itens = []
    total = 0
    latest_at = None
    for r in rows or []:
        qtd = int(r.get("qtd") or 0)
        total += qtd
        ultima = r.get("ultima_em")
        if ultima and (latest_at is None or ultima > latest_at):
            latest_at = ultima
        itens.append({
            "servico_id": str(r["servico_id"]),
            "protocolo": r.get("protocolo") or "",
            "qtd": qtd,
        })
    return {
        "total": total,
        "itens": itens,
        "latest_at": latest_at.isoformat() if latest_at else "",
    }


def ensure_central_alertas_operacionais_table():
    """Garante tabela de alertas mesmo se init_db não rodou após deploy."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute('create extension if not exists "uuid-ossp";')
            cur.execute("""
            create table if not exists central_alertas_operacionais (
              id uuid primary key default uuid_generate_v4(),
              servico_id uuid not null,
              tipo text not null,
              titulo text,
              descricao text,
              visto boolean default false,
              created_at timestamp default now(),
              visto_em timestamp
            );""")
            cur.execute("""
            create index if not exists idx_central_alertas_servico
              on central_alertas_operacionais (servico_id, tipo, visto);
            """)
        conn.commit()


def registrar_alerta_operacional(servico_id, tipo, titulo="", descricao=""):
    sid = str(servico_id or "").strip()
    t = (tipo or "").strip().lower()
    if not sid or t not in ("mensagem", "placa", "recusa"):
        return None
    ensure_central_alertas_operacionais_table()
    existente = one(
        """
        select id from central_alertas_operacionais
         where servico_id=%s and tipo=%s and visto=false
         limit 1
        """,
        (sid, t),
    )
    if existente:
        q(
            """
            update central_alertas_operacionais
               set titulo=%s, descricao=%s, created_at=now()
             where id=%s
            """,
            ((titulo or "").strip(), (descricao or "").strip(), str(existente["id"])),
        )
        print(f"[CENTRAL_ALERTAS] atualizado tipo={t} servico_id={sid}")
        return str(existente["id"])
    row = one(
        """
        insert into central_alertas_operacionais (servico_id, tipo, titulo, descricao, visto, created_at)
        values (%s, %s, %s, %s, false, now())
        returning id
        """,
        (sid, t, (titulo or "").strip(), (descricao or "").strip()),
    )
    print(f"[CENTRAL_ALERTAS] criado tipo={t} servico_id={sid}")
    return str(row["id"]) if row else None


def listar_alertas_operacionais_pendentes(tipo=None):
    try:
        ensure_central_alertas_operacionais_table()
        params = []
        where = "where coalesce(visto, false) = false"
        if tipo:
            where += " and lower(tipo)=%s"
            params.append(str(tipo).strip().lower())
        rows = q(
            f"""
            select id, servico_id, tipo, titulo, descricao, created_at
              from central_alertas_operacionais
             {where}
             order by created_at desc
            """,
            tuple(params) if params else None,
            True,
        )
    except Exception as exc:
        print(f"[CENTRAL_ALERTAS] listar pendentes erro tipo={tipo}: {exc}")
        return []
    out = []
    for r in rows or []:
        s = servico_by_id(r.get("servico_id")) or {}
        out.append({
            "id": str(r["id"]),
            "servico_id": str(r["servico_id"]),
            "tipo": r.get("tipo") or "",
            "titulo": r.get("titulo") or "",
            "descricao": r.get("descricao") or "",
            "protocolo": s.get("protocolo") or "",
            "motorista_nome": s.get("motorista_nome") or "",
            "placa": s.get("placa_veiculo_removido") or s.get("placa_removida") or "",
            "status": s.get("status") or "",
            "created_at": formatar_data_hora_central(r.get("created_at")),
            "created_at_iso": r.get("created_at").isoformat() if r.get("created_at") else "",
        })
    print(f"[CENTRAL_ALERTAS] listar tipo={tipo or 'todos'} count={len(out)}")
    return out


def marcar_alerta_operacional_visto(servico_id, tipo):
    sid = str(servico_id or "").strip()
    t = (tipo or "").strip().lower()
    if not sid or t not in ("mensagem", "placa", "recusa"):
        return False
    if t == "mensagem":
        marcar_mensagens_mobile_lidas(sid, "central")
    ensure_central_alertas_operacionais_table()
    q(
        """
        update central_alertas_operacionais
           set visto=true, visto_em=now()
         where servico_id=%s and tipo=%s and visto=false
        """,
        (sid, t),
    )
    return True


def resumo_alertas_central():
    mensagens = resumo_mensagens_mobile_central()
    placas = listar_alertas_operacionais_pendentes("placa")
    recusas = listar_alertas_operacionais_pendentes("recusa")
    por_servico = {}

    def bump(sid, campo, qtd=1, protocolo=""):
        if not sid:
            return
        por_servico.setdefault(sid, {"mensagem": 0, "placa": 0, "recusa": 0, "protocolo": ""})
        por_servico[sid][campo] = por_servico[sid].get(campo, 0) + qtd
        if protocolo and not por_servico[sid].get("protocolo"):
            por_servico[sid]["protocolo"] = protocolo

    for item in mensagens.get("itens") or []:
        bump(str(item.get("servico_id")), "mensagem", int(item.get("qtd") or 0), item.get("protocolo") or "")

    for item in placas:
        bump(str(item.get("servico_id")), "placa", 1, item.get("protocolo") or "")
    for item in recusas:
        bump(str(item.get("servico_id")), "recusa", 1, item.get("protocolo") or "")

    for sid, info in por_servico.items():
        info["mensagens"] = info.get("mensagem", 0)
        info["placa_pendente"] = info.get("placa", 0) > 0
        info["recusa_pendente"] = info.get("recusa", 0) > 0

    latest_candidates = [mensagens.get("latest_at") or ""]
    latest_candidates += [a.get("created_at_iso") or "" for a in placas + recusas]
    latest_at = max([x for x in latest_candidates if x], default="")

    return {
        "mensagens": mensagens,
        "placas": placas,
        "recusas": recusas,
        "por_servico": por_servico,
        "latest_at": latest_at,
    }


def coletar_arquivos_servico(servico_id, servico=None):
    """Fotos gerais + fotos do checklist (URLs únicas)."""
    s = servico or servico_by_id(servico_id)
    vistos = set()
    arquivos = []

    def add(url, tipo, nome):
        u = (url or "").strip()
        if not u or u in vistos:
            return
        vistos.add(u)
        arquivos.append({"url": u, "tipo": tipo, "nome": nome or tipo})

    for f in s.get("fotos") or []:
        add(f.get("url"), "Foto do serviço", f.get("filename") or "Foto")
    for parte, dados in partes_checklist_dict(servico_id).items():
        for foto in dados.get('fotos') or []:
            info = resolver_url_foto_checklist(foto)
            add(info.get('url') or foto, f'Checklist — {parte}', parte)
    return arquivos


def enriquecer_foto_checklist(foto, request=None, protocolo='', parte='', index=0, servico_id=''):
    info = resolver_url_foto_checklist(foto, request)
    slug = _slug_parte_checklist(parte)
    prot = _protocolo_checklist_slug(protocolo, servico_id)
    indisponivel_local = info.get('erro') == 'uri_local_dispositivo'
    mensagem = ''
    if indisponivel_local:
        mensagem = 'Foto antiga do dispositivo — reenviar checklist para visualizar na Central.'
    return {
        'raw': info['raw'],
        'url': info['url'],
        'valid': info['valid'],
        'erro': info['erro'],
        'mensagem': mensagem,
        'indisponivel_local': indisponivel_local,
        'download_nome': f"checklist_{prot}_{slug}_{index + 1}.jpg",
        'parte': parte,
        'index': index + 1,
    }


def extrair_observacao_checklist_parte(dados):
    """Observação textual por parte (coluna DB ou chaves legadas do JSON)."""
    if not isinstance(dados, dict):
        return ''
    for chave in ('observacao', 'observacoes'):
        valor = dados.get(chave)
        if isinstance(valor, str):
            texto = valor.strip()
            if texto:
                return texto
    return ''


def checklist_resumo_servico(servico_id, request=None, servico=None):
    partes = partes_checklist_dict(servico_id)
    s = servico or servico_by_id(servico_id) or {}
    protocolo = s.get('protocolo') or servico_id
    tipo_veiculo = inferir_tipo_veiculo(s)
    mapa_imagens = mapa_imagens_checklist_central(tipo_veiculo)
    checklist = []
    for parte, dados in partes.items():
        fotos_raw = dados.get('fotos') or []
        fotos_resolvidas = [
            enriquecer_foto_checklist(f, request, protocolo, parte, i, servico_id)
            for i, f in enumerate(fotos_raw)
        ]
        checklist.append({
            'parte': parte,
            'marcacoes': dados.get('marcacoes') or [],
            'fotos': fotos_raw,
            'fotos_resolvidas': fotos_resolvidas,
            'observacao': extrair_observacao_checklist_parte(dados),
            'imagem_url': imagem_checklist_parte(parte, tipo_veiculo),
        })
    assinaturas = obter_assinaturas_checklist(servico_id)
    total_avarias = sum(len(c.get('marcacoes') or []) for c in checklist)
    total_fotos = sum(len(c.get('fotos') or []) for c in checklist)
    status = 'Concluído' if assinaturas.get('assinatura_origem_salva') else 'Pendente'
    if assinaturas.get('assinatura_destino_salva'):
        status = 'Origem e destino assinados'
    elif assinaturas.get('assinatura_origem_salva'):
        status = 'Origem assinada'
    return {
        'checklist': checklist,
        'assinaturas': assinaturas,
        'total_partes': len(checklist),
        'total_avarias': total_avarias,
        'total_fotos_checklist': total_fotos,
        'status': status,
        'protocolo': protocolo,
        'tipo_veiculo': tipo_veiculo,
        'tipo_veiculo_label': label_tipo_veiculo_checklist(tipo_veiculo),
        'mapa_imagens': mapa_imagens,
    }


def lista_motoristas():
    return [normalizar_motorista(r) for r in q("select * from motoristas where coalesce(ativo,true)=true order by nome asc", fetch=True)]
def lista_servicos(limit=None, ativos=False, filtros=None):
    filtros=filtros or {}; where=[]; params=[]
    if ativos: where.append("status not in ('finalizado','recusado')")
    for key,col,op in [("data_ini","date(created_at)",">="),("data_fim","date(created_at)","<=")]:
        if filtros.get(key): where.append(f"{col} {op} %s"); params.append(filtros[key])
    for key,col in [("seguradora","seguradora"),("tipo","tipo"),("motorista","coalesce(motorista_nome,'')")]:
        if filtros.get(key): where.append(f"{col} ilike %s"); params.append(f"%{filtros[key]}%")
    if filtros.get("status"): where.append("status=%s"); params.append(filtros["status"])
    sql="select * from servicos" + ((" where " + " and ".join(where)) if where else "") + " order by created_at desc"
    if limit: sql += f" limit {int(limit)}"
    return [normalizar_servico(r) for r in q(sql, tuple(params), True)]

def lista_servicos_hoje(limit=200):
    """
    Central operacional:
    - mostra somente serviços do dia atual
    - NÃO mostra serviços finalizados
    - histórico/relatórios continuam mostrando tudo
    """
    limit = int(limit or 200)
    rows = q(f"""
        select *
        from servicos
        where created_at::date = current_date
          and coalesce(status,'novo') <> 'finalizado'
        order by created_at desc
        limit {limit}
    """, fetch=True)
    return [normalizar_servico(r) for r in rows]

def motorista_by_id(mid): return normalizar_motorista(one("select * from motoristas where id=%s",(str(mid),)))

def distancia_km(lat1, lon1, lat2, lon2):
    try:
        lat1, lon1, lat2, lon2 = map(float, [lat1, lon1, lat2, lon2])
        r = 6371.0
        p1 = math.radians(lat1)
        p2 = math.radians(lat2)
        dp = math.radians(lat2-lat1)
        dl = math.radians(lon2-lon1)
        a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
        return round(2*r*math.atan2(math.sqrt(a), math.sqrt(1-a)), 1)
    except Exception:
        return None

def motorista_ocupado(mid):
    row = one("""
        select id, protocolo, status
        from servicos
        where motorista_id=%s
          and status not in ('finalizado','recusado')
        order by coalesce(created_at, now()) desc
        limit 1
    """, (str(mid),))
    return row


def servico_by_id(sid): return normalizar_servico(one("select * from servicos where id=%s",(str(sid),)))
def registrar_evento_db(sid, status, detalhe=''):
    s=one("select historico from servicos where id=%s",(str(sid),)); historico=s.get("historico") if s else []
    if isinstance(historico,str):
        try: historico=json.loads(historico)
        except Exception: historico=[]
    historico=historico or []
    evento={'status':status,'detalhe':detalhe,'data_hora':agora()}
    historico.append(evento)
    ultimo=f"{evento['data_hora']} - {status}{(' - '+detalhe) if detalhe else ''}"
    q("update servicos set historico=%s, ultimo_evento=%s, atualizado_em=now() where id=%s",(Json(historico),ultimo,str(sid)))
def get_lan_ip():
    try:
        s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM); s.connect(("8.8.8.8",80)); ip=s.getsockname()[0]; s.close(); return ip
    except Exception: return "SEU_IP"


def fmt_minutos_dashboard(valor):
    if valor is None:
        return "0 min"
    try:
        return f"{int(round(float(valor)))} min"
    except Exception:
        return "0 min"


# Colunas legadas (ex.: atualizado_em) podem estar como text no banco — só o Dashboard usa estes casts.
_DASH_ISO_TS_RE = r"^[0-9]{4}-[0-9]{2}-[0-9]{2}"


def _dash_sql_ts(coluna):
    """Converte text ou timestamp para timestamptz quando o valor parece data ISO."""
    return f"""(
      CASE
        WHEN coalesce({coluna}::text, '') ~ '{_DASH_ISO_TS_RE}'
        THEN ({coluna}::text)::timestamptz
        ELSE NULL
      END
    )"""


def _dash_sql_date(coluna):
    return f"""(
      CASE
        WHEN coalesce({coluna}::text, '') ~ '{_DASH_ISO_TS_RE}'
        THEN ({coluna}::text)::date
        ELSE NULL
      END
    )"""


def _dash_where_servico_hoje(prefixo=""):
    col = f"{prefixo}created_at" if prefixo else "created_at"
    return f"{_dash_sql_date(col)} = current_date"


def dados_dashboard():
    ts_created = _dash_sql_ts("created_at")
    row = one(
        f"""
        select
          count(*)::int as total_cadastrados,
          count(*) filter (where coalesce(status,'') = 'finalizado')::int as concluidos,
          count(*) filter (
            where coalesce(status,'') not in ('finalizado','recusado','cancelado')
          )::int as em_andamento,
          count(*) filter (
            where coalesce(status,'') in ('na origem', 'em transporte')
          )::int as em_origem_transporte,
          count(*) filter (where {_dash_where_servico_hoje()})::int as do_dia,
          count(*) filter (where coalesce(status,'') = 'recusado')::int as recusados,
          count(*) filter (where coalesce(status,'') = 'cancelado')::int as cancelados,
          coalesce(
            sum(valor_total) filter (
              where {ts_created} is not null
                and date_trunc('month', {ts_created}) = date_trunc('month', now())
            ), 0
          ) as receita_mes,
          (
            select count(*)::int
              from motoristas
             where coalesce(ativo, true) = true
               and coalesce(online, false) = true
          ) as motoristas_online
        from servicos
        """,
    ) or {}

    receita = float(row.get("receita_mes") or 0)
    return {
        "total_cadastrados": int(row.get("total_cadastrados") or 0),
        "concluidos": int(row.get("concluidos") or 0),
        "em_andamento": int(row.get("em_andamento") or 0),
        "em_origem_transporte": int(row.get("em_origem_transporte") or 0),
        "do_dia": int(row.get("do_dia") or 0),
        "recusados": int(row.get("recusados") or 0),
        "cancelados": int(row.get("cancelados") or 0),
        "motoristas_online": int(row.get("motoristas_online") or 0),
        "receita_mes": receita,
        "receita_mes_fmt": fmt_moeda(receita),
        "tmc_dia": "0 min",
        "tme_mes": "0 min",
        "tmc_mes": "0 min",
        "km_hoje": 0,
    }


def extrair_bairro_cidade(endereco):
    """Extrai bairro e cidade de endereço textual quando possível."""
    if not endereco or not str(endereco).strip():
        return "-", "-"
    texto = str(endereco).strip()
    partes = [p.strip() for p in re.split(r"\s*-\s*", texto) if p.strip()]
    if len(partes) >= 2:
        cidade = partes[-1].split("/")[0].strip()
        bairro = partes[-2] if len(partes) >= 3 else "-"
        if len(bairro) > 60:
            bairro = "-"
        if len(cidade) > 60:
            cidade = partes[-1][:60]
        return bairro or "-", cidade or "-"
    partes = [p.strip() for p in texto.split(",") if p.strip()]
    if len(partes) >= 2:
        cidade = partes[-1].split("/")[0].strip() or "-"
        bairro = partes[-2] if len(partes) >= 3 else "-"
        return bairro, cidade
    return "-", "-"


def enriquecer_servico_central(s, raw_row=None):
    if not s:
        return s
    ob, oc = extrair_bairro_cidade(s.get("origem"))
    db, dc = extrair_bairro_cidade(s.get("destino"))
    s["origem_bairro"] = ob
    s["origem_cidade"] = oc
    s["destino_bairro"] = db
    s["destino_cidade"] = dc
    if raw_row is not None:
        ts = raw_row.get("created_at") or raw_row.get("criado_em")
    else:
        ts = s.get("created_at") or s.get("criado_em")
    s["data_hora_central"] = formatar_data_hora_central(ts)
    return s


def _central_busca_avancada_ativa(**kwargs):
    avancados = (
        "data_ini", "data_fim", "protocolo", "placa", "seguradora",
        "motorista", "tipo", "origem", "destino", "checklist", "status_faturamento",
    )
    return any((kwargs.get(k) or "").strip() for k in avancados)


def _central_titulo_filtros(filtros):
    partes = []
    if filtros.get("statuses"):
        partes.append("Status: " + filtros["statuses"].replace(",", ", "))
    elif filtros.get("status"):
        partes.append("Status: " + filtros["status"])
    rotulos = {
        "data_ini": "De",
        "data_fim": "Até",
        "protocolo": "Protocolo",
        "placa": "Placa",
        "seguradora": "Empresa",
        "motorista": "Motorista",
        "tipo": "Tipo",
        "origem": "Origem",
        "destino": "Destino",
        "checklist": "Checklist",
        "status_faturamento": "Fat.",
    }
    for chave, rotulo in rotulos.items():
        val = (filtros.get(chave) or "").strip()
        if val:
            partes.append(f"{rotulo}: {val}")
    return " · ".join(partes)


def lista_servicos_central_filtrada(
    statuses: str = "",
    status: str = "",
    data_ini: str = "",
    data_fim: str = "",
    protocolo: str = "",
    placa: str = "",
    seguradora: str = "",
    motorista: str = "",
    tipo: str = "",
    origem: str = "",
    destino: str = "",
    checklist: str = "",
    status_faturamento: str = "",
    limit: int = 200,
):
    """Lista serviços para a Central com filtros opcionais."""
    limit = int(limit or 200)
    where = []
    params = []

    data_ini = (data_ini or "").strip()
    data_fim = (data_fim or "").strip()
    if data_ini:
        where.append("created_at::date >= %s")
        params.append(data_ini)
    if data_fim:
        where.append("created_at::date <= %s")
        params.append(data_fim)
    if not data_ini and not data_fim:
        where.append("created_at::date = current_date")

    statuses = (statuses or "").strip()
    status = (status or "").strip()
    if statuses:
        sts = [s.strip().lower() for s in statuses.split(",") if s.strip()]
        if sts:
            placeholders = ",".join(["%s"] * len(sts))
            where.append(f"lower(coalesce(status, 'novo')) in ({placeholders})")
            params.extend(sts)
    elif status:
        where.append("lower(coalesce(status, 'novo')) = %s")
        params.append(status.lower())
    elif not _central_busca_avancada_ativa(
        data_ini=data_ini, data_fim=data_fim, protocolo=protocolo, placa=placa,
        seguradora=seguradora, motorista=motorista, tipo=tipo, origem=origem,
        destino=destino, checklist=checklist, status_faturamento=status_faturamento,
    ):
        where.append("coalesce(status, 'novo') <> 'finalizado'")

    protocolo = (protocolo or "").strip()
    if protocolo:
        where.append("coalesce(protocolo, '') ilike %s")
        params.append(f"%{protocolo}%")

    placa = (placa or "").strip().upper().replace("-", "").replace(" ", "")
    if placa:
        where.append(
            """(
              upper(replace(replace(coalesce(placa_veiculo_removido, ''), '-', ''), ' ', '')) ilike %s
              or upper(replace(replace(coalesce(placa_removida, ''), '-', ''), ' ', '')) ilike %s
            )"""
        )
        params.extend([f"%{placa}%", f"%{placa}%"])

    seguradora = (seguradora or "").strip()
    if seguradora:
        where.append("coalesce(seguradora, '') ilike %s")
        params.append(f"%{seguradora}%")

    motorista = (motorista or "").strip()
    if motorista:
        where.append("coalesce(motorista_nome, '') ilike %s")
        params.append(f"%{motorista}%")

    tipo = (tipo or "").strip()
    if tipo:
        where.append("coalesce(tipo, '') ilike %s")
        params.append(f"%{tipo}%")

    origem = (origem or "").strip()
    if origem:
        where.append("coalesce(origem, '') ilike %s")
        params.append(f"%{origem}%")

    destino = (destino or "").strip()
    if destino:
        where.append("coalesce(destino, '') ilike %s")
        params.append(f"%{destino}%")

    status_faturamento = (status_faturamento or "").strip()
    if status_faturamento:
        where.append("coalesce(status_faturamento, 'para_conferir') = %s")
        params.append(status_faturamento)

    checklist = (checklist or "").strip().lower()
    if checklist == "nao_iniciado":
        where.append(
            "not exists (select 1 from checklist_avarias c where c.servico_id = servicos.id)"
        )
    elif checklist == "iniciado":
        where.append(
            "exists (select 1 from checklist_avarias c where c.servico_id = servicos.id)"
        )
        where.append(
            """not exists (
                 select 1 from checklist_assinaturas a
                  where a.servico_id = servicos.id
                    and coalesce(a.assinatura_origem_base64, '') <> ''
               )"""
        )
    elif checklist == "concluido":
        where.append(
            """exists (
                 select 1 from checklist_assinaturas a
                  where a.servico_id = servicos.id
                    and coalesce(a.assinatura_origem_base64, '') <> ''
               )"""
        )

    sql = "select * from servicos"
    if where:
        sql += " where " + " and ".join(where)
    sql += f" order by created_at desc limit {limit}"

    rows = q(sql, tuple(params), True)
    return [enriquecer_servico_central(normalizar_servico(r), raw_row=r) for r in rows]


def status_operacional_motorista_dashboard(motorista):
    if not motorista.get("online"):
        return "offline", "Offline", "#dc2626"
    ocupado = motorista_ocupado(motorista.get("id"))
    if not ocupado:
        return "livre", "Livre", "#16a34a"
    st = (ocupado.get("status") or "").lower()
    mapa = {
        "a caminho": ("a_caminho", "A caminho", "#eab308"),
        "na origem": ("na_origem", "Na origem", "#f97316"),
        "em transporte": ("em_transporte", "Em transporte", "#2563eb"),
        "enviado": ("enviado", "Enviado", "#eab308"),
        "aceito": ("aceito", "Aceito", "#16a34a"),
    }
    return mapa.get(st, ("livre", "Livre", "#16a34a"))


def distancia_origem_dashboard(motorista, servico):
    if not servico:
        return None
    mlat = motorista.get("lat")
    mlng = motorista.get("lng")
    lat = servico.get("origem_lat")
    lng = servico.get("origem_lng")
    if not mlat or not mlng or not lat or not lng:
        return None
    dk = distancia_km(mlat, mlng, lat, lng)
    if dk is None:
        return None
    return f"{dk} km aprox."


def motoristas_mapa_dashboard_enriquecido():
    rows = q(
        """
        select *
          from motoristas
         where coalesce(ativo, true) = true
        """,
        fetch=True,
    )
    lista = []
    for r in rows:
        m = normalizar_motorista(r)
        if not m:
            continue
        codigo, label, cor = status_operacional_motorista_dashboard(m)
        servico_ativo = None
        distancia_origem = "-"
        protocolo = None
        ocupado = motorista_ocupado(m.get("id"))
        if ocupado:
            servico_ativo = servico_by_id(ocupado.get("id"))
            if servico_ativo:
                protocolo = servico_ativo.get("protocolo")
                dist = distancia_origem_dashboard(m, servico_ativo)
                if dist:
                    distancia_origem = dist
        try:
            lat = float(m.get("lat")) if m.get("lat") is not None else None
            lng = float(m.get("lng")) if m.get("lng") is not None else None
        except Exception:
            lat, lng = None, None
        lista.append(
            {
                "id": str(m.get("id")),
                "nome": m.get("nome") or "Motorista",
                "placa": m.get("placa_atual") or m.get("placa") or "-",
                "online": bool(m.get("online")),
                "lat": lat,
                "lng": lng,
                "status_operacional": codigo,
                "status_label": label,
                "cor_marcador": cor,
                "gps_atualizacao": dt_str(m.get("ultima_atualizacao")) or "-",
                "servico_protocolo": protocolo,
                "distancia_origem": distancia_origem,
            }
        )
    return lista


ORDEM_STATUS_DASHBOARD = [
    ("enviado", "Enviado"),
    ("a caminho", "A caminho"),
    ("na origem", "Na origem"),
    ("em transporte", "Em transporte"),
    ("finalizado", "Finalizado"),
    ("cancelado", "Cancelado"),
]


def resumo_status_operacional_dashboard():
    rows = q(
        f"""
        select lower(coalesce(status, 'novo')) as status, count(*)::int as total
          from servicos
         where {_dash_where_servico_hoje()}
         group by 1
        """,
        fetch=True,
    )
    mapa = {r.get("status"): int(r.get("total") or 0) for r in rows}
    return [
        {"chave": chave, "status": rotulo, "total": mapa.get(chave, 0)}
        for chave, rotulo in ORDEM_STATUS_DASHBOARD
    ]


def alertas_operacionais_dashboard():
    alertas = []
    ts_atualizado = _dash_sql_ts("atualizado_em")

    sem_placa = q(
        f"""
        select id, protocolo
          from servicos
         where {_dash_where_servico_hoje()}
           and coalesce(status, '') not in ('finalizado', 'recusado', 'cancelado')
           and coalesce(placa_veiculo_removido, placa_removida, '') = ''
         order by {_dash_sql_ts("created_at")} desc nulls last
         limit 20
        """,
        fetch=True,
    )
    for s in sem_placa:
        alertas.append(
            {
                "tipo": "sem_placa",
                "icone": "⚠",
                "texto": f"Serviço sem placa — {s.get('protocolo') or s.get('id')}",
                "link": "/central",
            }
        )

    sem_checklist = q(
        f"""
        select s.id, s.protocolo
          from servicos s
          left join checklist_avarias c on c.servico_id = s.id
         where {_dash_where_servico_hoje("s.")}
           and coalesce(s.status, '') not in ('finalizado', 'recusado', 'cancelado')
         group by s.id, s.protocolo
        having count(c.id) = 0
         order by max({_dash_sql_ts("s.created_at")}) desc nulls last
         limit 20
        """,
        fetch=True,
    )
    for s in sem_checklist:
        alertas.append(
            {
                "tipo": "sem_checklist",
                "icone": "⚠",
                "texto": f"Checklist não iniciado — {s.get('protocolo') or s.get('id')}",
                "link": f"/servicos/{s.get('id')}/checklist",
            }
        )

    sem_gps = q(
        """
        select nome
          from motoristas
         where coalesce(ativo, true) = true
           and coalesce(online, false) = true
           and (lat is null or lng is null)
         order by nome
         limit 20
        """,
        fetch=True,
    )
    for m in sem_gps:
        alertas.append(
            {
                "tipo": "sem_gps",
                "icone": "⚠",
                "texto": f"Motorista sem GPS — {m.get('nome')}",
                "link": "/motoristas?online=1",
            }
        )

    parados = q(
        f"""
        select protocolo
          from servicos
         where {_dash_where_servico_hoje()}
           and coalesce(status, '') not in ('finalizado', 'recusado', 'cancelado')
           and {ts_atualizado} is not null
           and {ts_atualizado} < now() - interval '30 minutes'
         order by {ts_atualizado} asc nulls last
         limit 20
        """,
        fetch=True,
    )
    for s in parados:
        alertas.append(
            {
                "tipo": "parado",
                "icone": "⚠",
                "texto": f"Serviço parado há mais de 30 min — {s.get('protocolo')}",
                "link": "/central",
            }
        )

    na_origem_longo = q(
        f"""
        select protocolo
          from servicos
         where {_dash_where_servico_hoje()}
           and lower(coalesce(status, '')) = 'na origem'
           and {ts_atualizado} is not null
           and {ts_atualizado} < now() - interval '60 minutes'
         order by {ts_atualizado} asc nulls last
         limit 20
        """,
        fetch=True,
    )
    for s in na_origem_longo:
        alertas.append(
            {
                "tipo": "origem_longo",
                "icone": "⚠",
                "texto": f"Na origem há mais de 60 min — {s.get('protocolo')}",
                "link": "/central?statuses=na origem",
            }
        )

    sem_motorista = q(
        f"""
        select protocolo
          from servicos
         where {_dash_where_servico_hoje()}
           and lower(coalesce(status, '')) = 'enviado'
           and (motorista_id is null or coalesce(motorista_nome, '') = '')
         order by {_dash_sql_ts("created_at")} desc nulls last
         limit 20
        """,
        fetch=True,
    )
    for s in sem_motorista:
        alertas.append(
            {
                "tipo": "sem_motorista",
                "icone": "⚠",
                "texto": f"Enviado sem motorista — {s.get('protocolo')}",
                "link": "/central?statuses=enviado",
            }
        )

    return alertas[:40]


def ranking_motoristas_hoje_dashboard():
    rows = q(
        f"""
        select coalesce(motorista_nome, 'Sem nome') as nome,
               motorista_id,
               count(*)::int as qtd
          from servicos
         where {_dash_where_servico_hoje()}
           and motorista_id is not null
         group by motorista_nome, motorista_id
         order by qtd desc, nome asc
         limit 8
        """,
        fetch=True,
    )
    return [
        {
            "nome": r.get("nome"),
            "qtd": int(r.get("qtd") or 0),
            "tempo_medio": "0 min",
        }
        for r in rows
    ]


def payload_dashboard_live():
    return {
        "ok": True,
        "data_ref": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "kpis": dados_dashboard(),
        "resumo_status": resumo_status_operacional_dashboard(),
        "resumo_financeiro": resumo_financeiro_dashboard(),
        "motoristas_mapa": motoristas_mapa_dashboard_enriquecido(),
        "alertas": alertas_operacionais_dashboard(),
        "ranking": ranking_motoristas_hoje_dashboard(),
    }


def resumo_operacional_dashboard():
    rows = q(
        """
        select coalesce(status, 'novo') as status, count(*)::int as total
          from servicos
         where created_at::date = current_date
         group by 1
         order by total desc
        """,
        fetch=True,
    )
    return [dict(r) for r in rows]


def resumo_financeiro_dashboard():
    ts_created = _dash_sql_ts("created_at")
    rows = q(
        f"""
        select coalesce(status_faturamento, 'para_conferir') as status,
               count(*)::int as total,
               coalesce(sum(valor_total), 0) as valor
          from servicos
         where {ts_created} is not null
           and date_trunc('month', {ts_created}) = date_trunc('month', now())
         group by 1
         order by total desc
        """,
        fetch=True,
    )
    out = []
    for r in rows:
        item = dict(r)
        item["valor_fmt"] = fmt_moeda(float(item.get("valor") or 0))
        out.append(item)
    return out


# ============================================================
# CONTROLE — frota, danos, abastecimentos e compliance
# ============================================================

CONTROLE_MODULOS = {
    "danos": {
        "slug": "danos",
        "titulo": "Danos",
        "icone": "💥",
        "descricao": "Registro e acompanhamento de danos em viaturas, objetos e operações.",
        "tabela": "controle_danos",
        "rota": "/controle/danos",
        "api": "/api/controle/danos",
        "colunas": [
            ("data_dano", "Data"),
            ("profissional", "Profissional"),
            ("viatura", "Viatura"),
            ("origem", "Origem"),
            ("tipo_dano", "Tipo"),
            ("objeto", "Objeto"),
            ("valor", "Valor"),
            ("situacao", "Situação"),
            ("status", "Status"),
        ],
    },
    "abastecimentos": {
        "slug": "abastecimentos",
        "titulo": "Abastecimentos",
        "icone": "⛽",
        "descricao": "Controle de combustível, postos, litros e custos por viatura.",
        "tabela": "controle_abastecimentos",
        "rota": "/controle/abastecimentos",
        "api": "/api/controle/abastecimentos",
        "colunas": [
            ("data_abastecimento", "Data"),
            ("viatura", "Viatura"),
            ("posto", "Posto"),
            ("combustivel", "Combustível"),
            ("litros", "Litros"),
            ("valor_total", "Total"),
            ("hodometro", "Hodômetro"),
            ("profissional", "Profissional"),
            ("status", "Status"),
        ],
    },
    "checklist-viatura": {
        "slug": "checklist-viatura",
        "titulo": "Checklist Viatura",
        "icone": "✅",
        "descricao": "Checklists de saída, retorno e inspeção das viaturas da frota.",
        "tabela": "controle_checklists_viatura",
        "rota": "/controle/checklist-viatura",
        "api": "/api/controle/checklist-viatura",
        "colunas": [
            ("data_checklist", "Data"),
            ("viatura", "Viatura"),
            ("motorista", "Motorista"),
            ("tipo_checklist", "Tipo"),
            ("km", "KM"),
            ("status_checklist", "Checklist"),
            ("status", "Status"),
        ],
    },
    "manutencoes": {
        "slug": "manutencoes",
        "titulo": "Manutenções",
        "icone": "🔧",
        "descricao": "Ordens de serviço, fornecedores e custos de manutenção preventiva e corretiva.",
        "tabela": "controle_manutencoes",
        "rota": "/controle/manutencoes",
        "api": "/api/controle/manutencoes",
        "colunas": [
            ("data_manutencao", "Data"),
            ("documento", "Documento"),
            ("fornecedor", "Fornecedor"),
            ("viatura", "Viatura"),
            ("hodometro", "Hodômetro"),
            ("total", "Total"),
            ("status", "Status"),
        ],
    },
    "multas": {
        "slug": "multas",
        "titulo": "Multas",
        "icone": "🚨",
        "descricao": "Infrações, condutores, vencimentos e situação de pagamento.",
        "tabela": "controle_multas",
        "rota": "/controle/multas",
        "api": "/api/controle/multas",
        "colunas": [
            ("data_infracao", "Data"),
            ("viatura", "Viatura"),
            ("auto_infracao", "Auto"),
            ("municipio", "Município"),
            ("condutor", "Condutor"),
            ("valor", "Valor"),
            ("vencimento", "Vencimento"),
            ("situacao", "Situação"),
            ("status", "Status"),
        ],
    },
    "seguros": {
        "slug": "seguros",
        "titulo": "Seguros",
        "icone": "🛡️",
        "descricao": "Apólices, vigências, parcelas e situação dos seguros da frota.",
        "tabela": "controle_seguros",
        "rota": "/controle/seguros",
        "api": "/api/controle/seguros",
        "colunas": [
            ("viatura", "Viatura"),
            ("seguradora", "Seguradora"),
            ("apolice", "Apólice"),
            ("vigencia_inicial", "Início"),
            ("vigencia_final", "Fim"),
            ("valor", "Valor"),
            ("parcelas", "Parcelas"),
            ("situacao", "Situação"),
            ("status", "Status"),
        ],
    },
}


def _formatar_valor_controle(campo, valor):
    if valor is None or valor == "":
        return "-"
    if campo in ("valor", "valor_total", "valor_unitario", "total", "valor_pago", "total_produtos", "total_servicos"):
        try:
            return fmt_moeda(float(valor))
        except Exception:
            return str(valor)
    if campo in ("litros", "hodometro", "km", "parcelas"):
        try:
            n = float(valor)
            return str(int(n)) if n == int(n) else str(n)
        except Exception:
            return str(valor)
    if hasattr(valor, "strftime"):
        if isinstance(valor, datetime):
            return valor.strftime("%d/%m/%Y %H:%M")
        return valor.strftime("%d/%m/%Y")
    texto = str(valor).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}[T\s]", texto):
        try:
            return datetime.fromisoformat(texto.replace("Z", "+00:00")[:19]).strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass
    if re.match(r"^\d{4}-\d{2}-\d{2}", texto):
        try:
            return datetime.strptime(texto[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            pass
    return texto


def _data_iso_controle(item):
    for k in (
        "data_hora_dano",
        "data_hora",
        "data_hora_infracao",
        "data_abastecimento",
        "data_checklist",
        "data_manutencao",
        "data_infracao",
        "data_dano",
        "vigencia_inicial",
        "created_at",
    ):
        v = item.get(k)
        if v and hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
    return ""


def _txt_controle(v, max_len=2000):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    if max_len and len(s) > max_len:
        return s[:max_len]
    return s


def _num_controle(v, default=0.0):
    if v is None or v == "":
        return default
    s = str(v).strip().replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return default


def _int_controle(v, default=0):
    try:
        return int(float(_num_controle(v, default)))
    except Exception:
        return default


def _bool_controle(v):
    if isinstance(v, bool):
        return v
    return str(v or "").strip().lower() in ("1", "true", "sim", "on", "yes")


def _parse_data_controle(v):
    if not v:
        return None
    if hasattr(v, "strftime") and not isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s:
        return None
    if "T" in s:
        try:
            return datetime.fromisoformat(s[:19]).date()
        except Exception:
            pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10] if fmt == "%Y-%m-%d" else s[:10], fmt).date()
        except Exception:
            continue
    dt = parse_datetime_br(s)
    return dt.date() if dt else None


def _parse_datetime_controle(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=None) if v.tzinfo else v
    s = str(v).strip()
    if not s:
        return None
    if "T" in s:
        try:
            return datetime.fromisoformat(s.replace("Z", "")[:19])
        except Exception:
            pass
    dt = parse_datetime_br(s)
    if dt:
        return dt
    d = _parse_data_controle(s)
    return datetime.combine(d, datetime.min.time()) if d else None


def inserir_controle_danos(p):
    data_hora = _parse_datetime_controle(p.get("data_hora_dano"))
    data_dano = data_hora.date() if data_hora else _parse_data_controle(p.get("data_dano"))
    row = one(
        """
        insert into controle_danos (
          data_dano, data_hora_dano, tipo_dano, origem, profissional, aviso, situacao,
          assistencia, viatura, objeto, identificacao_objeto, valor, parcelas,
          data_desconto, descricao_dano, analise_interna, observacoes, status, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ativo',now())
        returning *
        """,
        (
            data_dano,
            data_hora,
            _txt_controle(p.get("tipo_dano"), 120),
            _txt_controle(p.get("origem"), 120),
            _txt_controle(p.get("profissional"), 200),
            _txt_controle(p.get("aviso"), 200),
            _txt_controle(p.get("situacao"), 120),
            _txt_controle(p.get("assistencia"), 200),
            _txt_controle(p.get("viatura"), 80),
            _txt_controle(p.get("objeto"), 200),
            _txt_controle(p.get("identificacao_objeto"), 200),
            _num_controle(p.get("valor")),
            _int_controle(p.get("parcelas"), 1) or 1,
            _parse_data_controle(p.get("data_desconto")),
            _txt_controle(p.get("descricao_dano")),
            _txt_controle(p.get("analise_interna")),
            _txt_controle(p.get("observacoes")),
        ),
    )
    return normalizar_registro_controle(row)


def inserir_controle_abastecimentos(p):
    data_hora = _parse_datetime_controle(p.get("data_hora"))
    data_abast = data_hora.date() if data_hora else _parse_data_controle(p.get("data_abastecimento"))
    row = one(
        """
        insert into controle_abastecimentos (
          data_abastecimento, data_hora, viatura, posto, profissional, combustivel,
          litros, valor_unitario, valor_total, hodometro, observacoes,
          desconsiderar_ultimo_km, gerar_contas_pagar, status, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ativo',now())
        returning *
        """,
        (
            data_abast,
            data_hora,
            _txt_controle(p.get("viatura"), 80),
            _txt_controle(p.get("posto"), 200),
            _txt_controle(p.get("profissional"), 200),
            _txt_controle(p.get("combustivel"), 80),
            _num_controle(p.get("litros")),
            _num_controle(p.get("valor_unitario")),
            _num_controle(p.get("valor_total")),
            _int_controle(p.get("hodometro")) or None,
            _txt_controle(p.get("observacao") or p.get("observacoes")),
            _bool_controle(p.get("desconsiderar_ultimo_km")),
            _bool_controle(p.get("gerar_contas_pagar")),
        ),
    )
    return normalizar_registro_controle(row)


def inserir_controle_manutencoes(p):
    itens = p.get("itens") or []
    if isinstance(itens, str):
        try:
            itens = json.loads(itens)
        except Exception:
            itens = []
    itens_norm = []
    for it in itens:
        if not isinstance(it, dict):
            continue
        qtd = _num_controle(it.get("quantidade"), 1) or 1
        vu = _num_controle(it.get("valor_unitario"))
        vt = _num_controle(it.get("valor_total"), qtd * vu)
        desc = _txt_controle(it.get("descricao"), 300)
        if desc:
            itens_norm.append(
                {
                    "descricao": desc,
                    "valor_unitario": vu,
                    "quantidade": qtd,
                    "valor_total": vt,
                    "tipo": _txt_controle(it.get("tipo"), 20) or "produto",
                }
            )
    total_calc = sum(i["valor_total"] for i in itens_norm)
    total_prod = _num_controle(p.get("total_produtos"))
    total_serv = _num_controle(p.get("total_servicos"))
    total_geral = _num_controle(p.get("total"), total_calc or (total_prod + total_serv))
    data_hora = _parse_datetime_controle(p.get("data_hora"))
    data_man = _parse_data_controle(p.get("data")) or (data_hora.date() if data_hora else None)
    row = one(
        """
        insert into controle_manutencoes (
          data_manutencao, viatura, hodometro, documento, fornecedor, observacoes,
          itens, total_produtos, total_servicos, total, status, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ativo',now())
        returning *
        """,
        (
            data_man,
            _txt_controle(p.get("viatura"), 80),
            _int_controle(p.get("hodometro")) or None,
            _txt_controle(p.get("documento"), 120),
            _txt_controle(p.get("fornecedor"), 200),
            _txt_controle(p.get("observacao") or p.get("observacoes")),
            Json(itens_norm),
            total_prod,
            total_serv,
            total_geral,
        ),
    )
    return normalizar_registro_controle(row)


def inserir_controle_multas(p):
    data_hora = _parse_datetime_controle(p.get("data_hora_infracao"))
    data_infracao = data_hora.date() if data_hora else _parse_data_controle(p.get("data_infracao"))
    row = one(
        """
        insert into controle_multas (
          viatura, auto_infracao, data_infracao, data_hora_infracao, data_limite_indicacao,
          municipio, endereco, descricao_multa, natureza, condutor, condutor_responsavel,
          parcelas, vencimento, valor, valor_pago, observacoes, situacao,
          contas_pagar, extrato_profissional, status, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ativo',now())
        returning *
        """,
        (
            _txt_controle(p.get("viatura"), 80),
            _txt_controle(p.get("auto_infracao"), 80),
            data_infracao,
            data_hora,
            _parse_data_controle(p.get("data_limite_indicacao")),
            _txt_controle(p.get("municipio"), 120),
            _txt_controle(p.get("endereco")),
            _txt_controle(p.get("descricao_multa")),
            _txt_controle(p.get("natureza"), 120),
            _txt_controle(p.get("condutor"), 200),
            _txt_controle(p.get("condutor_responsavel"), 200),
            _int_controle(p.get("parcelas"), 1) or 1,
            _parse_data_controle(p.get("vencimento") or p.get("data_vencimento")),
            _num_controle(p.get("valor")),
            _num_controle(p.get("valor_pago")),
            _txt_controle(p.get("observacao") or p.get("observacoes")),
            _txt_controle(p.get("situacao"), 80) or "pendente",
            _bool_controle(p.get("contas_pagar")),
            _bool_controle(p.get("extrato_profissional")),
        ),
    )
    return normalizar_registro_controle(row)


def inserir_controle_seguros(p):
    row = one(
        """
        insert into controle_seguros (
          viatura, seguradora, apolice, corretor, telefone,
          vigencia_inicial, vigencia_final, valor, parcelas, vencimento,
          observacoes, situacao, status, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ativo',now())
        returning *
        """,
        (
            _txt_controle(p.get("viatura"), 80),
            _txt_controle(p.get("seguradora"), 200),
            _txt_controle(p.get("apolice") or p.get("numero_apolice"), 120),
            _txt_controle(p.get("corretor"), 200),
            _txt_controle(p.get("telefone"), 40),
            _parse_data_controle(p.get("vigencia_inicial")),
            _parse_data_controle(p.get("vigencia_final")),
            _num_controle(p.get("valor") or p.get("valor_total")),
            _int_controle(p.get("parcelas"), 1) or 1,
            _parse_data_controle(p.get("vencimento")),
            _txt_controle(p.get("observacoes") or p.get("observacao")),
            _txt_controle(p.get("situacao"), 80) or "ativo",
        ),
    )
    return normalizar_registro_controle(row)


def inserir_controle_checklist_viatura(p):
    data_hora = _parse_datetime_controle(p.get("data_hora"))
    data_chk = data_hora.date() if data_hora else _parse_data_controle(p.get("data_checklist"))
    row = one(
        """
        insert into controle_checklists_viatura (
          viatura, motorista, tipo_checklist, data_checklist, data_hora, km,
          status_checklist, observacoes, itens_conferidos, itens_problema, status, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ativo',now())
        returning *
        """,
        (
            _txt_controle(p.get("viatura"), 80),
            _txt_controle(p.get("motorista"), 200),
            _txt_controle(p.get("tipo_checklist"), 80),
            data_chk,
            data_hora,
            _int_controle(p.get("km")) or None,
            _txt_controle(p.get("status_checklist") or p.get("status"), 80),
            _txt_controle(p.get("observacoes") or p.get("observacao")),
            _txt_controle(p.get("itens_conferidos")),
            _txt_controle(p.get("itens_problema")),
        ),
    )
    return normalizar_registro_controle(row)


CONTROLE_INSERIR = {
    "danos": inserir_controle_danos,
    "abastecimentos": inserir_controle_abastecimentos,
    "checklist-viatura": inserir_controle_checklist_viatura,
    "manutencoes": inserir_controle_manutencoes,
    "multas": inserir_controle_multas,
    "seguros": inserir_controle_seguros,
}


async def api_controle_salvar(slug: str, request: Request):
    cfg = CONTROLE_MODULOS.get(slug)
    if not cfg:
        return JSONResponse({"ok": False, "erro": "Módulo inválido"}, status_code=404)
    inserir = CONTROLE_INSERIR.get(slug)
    if not inserir:
        return JSONResponse({"ok": False, "erro": "Cadastro indisponível"}, status_code=400)
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    try:
        registro = inserir(payload or {})
        if not registro:
            return JSONResponse({"ok": False, "erro": "Falha ao gravar registro"}, status_code=500)
        return {
            "ok": True,
            "registro": registro,
            "kpis": resumo_modulo_controle(cfg["tabela"]),
        }
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


def normalizar_registro_controle(row):
    if not row:
        return None
    item = dict(row)
    item["id"] = str(item.get("id"))
    item["_data_iso"] = _data_iso_controle(item)
    item["created_at"] = dt_str(item.get("created_at"))
    item["updated_at"] = dt_str(item.get("updated_at"))
    exibicao = {}
    for chave in item.keys():
        if chave.startswith("_"):
            continue
        exibicao[chave] = _formatar_valor_controle(chave, item.get(chave))
    item["_fmt"] = exibicao
    return item


def listar_registros_controle(tabela):
    if tabela not in {m["tabela"] for m in CONTROLE_MODULOS.values()}:
        return []
    rows = q(
        f"select * from {tabela} order by created_at desc nulls last limit 500",
        fetch=True,
    )
    return [normalizar_registro_controle(r) for r in rows]


def resumo_modulo_controle(tabela):
    if tabela not in {m["tabela"] for m in CONTROLE_MODULOS.values()}:
        return {"total": 0, "ativos": 0, "mes": 0}
    row = one(
        f"""
        select
          count(*)::int as total,
          count(*) filter (where coalesce(status, 'ativo') = 'ativo')::int as ativos,
          count(*) filter (
            where created_at is not null
              and date_trunc('month', created_at::timestamp) = date_trunc('month', now())
          )::int as mes
        from {tabela}
        """
    ) or {}
    return {
        "total": int(row.get("total") or 0),
        "ativos": int(row.get("ativos") or 0),
        "mes": int(row.get("mes") or 0),
    }


def pagina_modulo_controle(slug: str, request: Request):
    cfg = CONTROLE_MODULOS.get(slug)
    if not cfg:
        return RedirectResponse("/controle", status_code=303)
    kpis = resumo_modulo_controle(cfg["tabela"])
    registros = listar_registros_controle(cfg["tabela"])
    colunas = [{"chave": c[0], "rotulo": c[1]} for c in cfg["colunas"]]
    return templates.TemplateResponse(
        "controle/modulo.html",
        {
            "request": request,
            "nav_ativo": "controle",
            "nav_som": False,
            "modulo": cfg,
            "kpis": kpis,
            "registros": registros,
            "colunas": colunas,
        },
    )


# ============================================================
# CADASTROS — viaturas e profissionais (premium)
# ============================================================

def _txt_cad(v, max_len=500):
    return _txt_controle(v, max_len)


def _status_cad(v):
    s = (_txt_cad(v, 20) or "ativo").lower()
    return "inativo" if s in ("inativo", "0", "false") else "ativo"


def _bool_cad(v):
    return _bool_controle(v)


def normalizar_placa_viatura(placa):
    import re
    if not placa:
        return ""
    return re.sub(r"[\s\-.]", "", str(placa).upper().strip())


def mapear_origem_viatura(classificacao):
    raw = (classificacao or "").strip()
    norm = raw.upper().replace("Ç", "C").replace("Ã", "A").replace("Õ", "O").replace("É", "E")
    if "TERCEIRO" in norm:
        return {"origem": "terceiro", "terceiro": True, "classificacao_original": raw or "TERCEIRO"}
    return {"origem": "propria", "terceiro": False, "classificacao_original": raw or "LOG SOLUÇÕES"}


def normalizar_viatura_cadastro(row, com_arquivos=False):
    if not row:
        return None
    item = dict(row)
    item["id"] = str(item.get("id"))
    item["status"] = _status_cad(item.get("status"))
    item["terceiro"] = bool(item.get("terceiro"))
    origem = (_txt_cad(item.get("origem")) or "").lower()
    if not origem:
        origem = "terceiro" if item["terceiro"] else "propria"
    item["origem"] = origem
    item["origem_label"] = "Terceiro" if origem == "terceiro" else "Própria"
    item["classificacao_original"] = (_txt_cad(item.get("classificacao_original")) or "").strip()
    if not item["classificacao_original"]:
        item["classificacao_original"] = "TERCEIRO" if origem == "terceiro" else "LOG SOLUÇÕES"
    item["exibicao"] = (item.get("exibicao") or item.get("placa") or item.get("modelo") or "-").strip()
    item["nome_exibicao"] = item["exibicao"]
    item["personalizacao"] = (_txt_cad(item.get("personalizacao")) or "").strip()
    item["personalizada"] = bool(item["personalizacao"])
    item["placa"] = normalizar_placa_viatura(item.get("placa")) or (_txt_cad(item.get("placa")) or "").upper()
    item["created_at"] = dt_str(item.get("created_at"))
    item["updated_at"] = dt_str(item.get("updated_at"))
    if com_arquivos:
        item["arquivos"] = listar_arquivos_viatura_cadastro(item["id"])
    item["qtd_arquivos"] = int(item.get("qtd_arquivos") or 0)
    return item


def _fmt_data_cad(v):
    if not v:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y")
    return str(v)[:10]


def _nome_exibicao_profissional(item):
    return (
        (item.get("nome_trabalho") or "").strip()
        or (item.get("nome_completo") or "").strip()
        or (item.get("nome") or "").strip()
        or "-"
    )


def normalizar_cpf_profissional(cpf):
    if not cpf:
        return ""
    digits = re.sub(r"\D", "", str(cpf))
    return digits


def formatar_cpf_profissional(cpf):
    n = normalizar_cpf_profissional(cpf)
    if len(n) == 11:
        return f"{n[:3]}.{n[3:6]}.{n[6:9]}-{n[9:]}"
    if len(n) == 14:
        return f"{n[:2]}.{n[2:5]}.{n[5:8]}/{n[8:12]}-{n[12:]}"
    return (cpf or "").strip()


def normalizar_telefone_profissional(tel):
    if not tel:
        return ""
    return re.sub(r"\D", "", str(tel))


def formatar_telefone_profissional(tel):
    n = normalizar_telefone_profissional(tel)
    if len(n) == 11:
        return f"({n[:2]}) {n[2:7]}-{n[7:]}"
    if len(n) == 10:
        return f"({n[:2]}) {n[2:6]}-{n[6:]}"
    return (tel or "").strip()


def normalizar_nome_trabalho_profissional(nome):
    return " ".join(str(nome or "").split())


def normalizar_nome_completo_profissional(nome):
    return " ".join(str(nome or "").upper().split())


def mapear_classificacao_vinculo_profissional(classificacao):
    raw = (classificacao or "").strip()
    norm = raw.upper().replace("Ç", "C").replace("Ã", "A").replace("Õ", "O").replace("É", "E").replace("Ô", "O")
    if any(t in norm for t in ("TERCEIRO", "TERCEIROS", "PRESTADOR", "AUTONOMO", "AUTÔNOMO", "TERCEIRIZADO", "TERCEIRIZADA")):
        return {"classificacao_vinculo": "terceiro", "terceiro": True, "classificacao_original": raw or "TERCEIRO"}
    if any(t in norm for t in ("CONTRATADO", "FUNCIONARIO", "FUNCIONÁRIO", "COLABORADOR", "FIXO")):
        return {"classificacao_vinculo": "contratado", "terceiro": False, "classificacao_original": raw or "CONTRATADO"}
    return {"classificacao_vinculo": "contratado", "terceiro": False, "classificacao_original": raw or "CONTRATADO"}


_MOTORISTA_KEYWORDS = (
    "MOTORISTA", "LEVE", "PESADO", "PRANCHAO", "PRANCHÃO", "APOIO", "REBOQUEIRO", "CARRETINHA", "CARRETINHO",
)


def inferir_tipo_e_funcao_profissional(funcao=None, categoria=None, nome_trabalho=None):
    partes = [funcao, categoria, nome_trabalho]
    texto = " ".join(str(p or "") for p in partes).upper()
    texto = texto.replace("Ã", "A").replace("Ç", "C").replace("Õ", "O").replace("Ô", "O")
    if any(kw in texto for kw in _MOTORISTA_KEYWORDS):
        func = _txt_cad(funcao) or _txt_cad(categoria) or ""
        if not func:
            nt = (nome_trabalho or "").upper()
            for k in ("CARRETINHA", "CARRETINHO", "PESADO", "PRANCHAO", "PRANCHÃO", "LEVE", "REBOQUEIRO", "MOTORISTA", "APOIO"):
                if k in nt:
                    func = k.title().replace("Pranchao", "Pranchão")
                    break
        return {"tipo_profissional": "motorista", "funcao": func or "Motorista"}
    func = _txt_cad(funcao) or _txt_cad(categoria) or ""
    if not func:
        nt = (nome_trabalho or "").upper()
        for k in ("DIARISTA", "PLANTAO", "PLANTÃO", "ADMINISTRATIVO", "OPERACIONAL"):
            if k in nt:
                func = k.title().replace("Plantao", "Plantão")
                break
    return {"tipo_profissional": "operacional", "funcao": func or "Operacional"}


def _vinculo_profissional_de_item(item):
    vinculo = (_txt_cad(item.get("classificacao_vinculo")) or "").lower()
    if vinculo in ("contratado", "terceiro"):
        return vinculo
    return "terceiro" if bool(item.get("terceiro")) else "contratado"


def normalizar_profissional_cadastro(row, com_arquivos=False):
    if not row:
        return None
    item = dict(row)
    item["id"] = str(item.get("id"))
    if item.get("motorista_id"):
        item["motorista_id"] = str(item["motorista_id"])
    item["nome_completo"] = normalizar_nome_completo_profissional(
        item.get("nome_completo") or item.get("nome") or ""
    )
    item["nome_trabalho"] = normalizar_nome_trabalho_profissional(
        item.get("nome_trabalho") or item.get("nome_completo") or item.get("nome") or ""
    )
    item["nome_exibicao"] = _nome_exibicao_profissional(item)
    item["nome"] = item["nome_exibicao"]
    tel_raw = item.get("telefone_movel") or item.get("telefone_fixo") or item.get("telefone") or ""
    item["telefone"] = formatar_telefone_profissional(tel_raw)
    item["telefone_norm"] = normalizar_telefone_profissional(tel_raw)
    item["cpf_norm"] = normalizar_cpf_profissional(item.get("cpf"))
    item["cpf"] = formatar_cpf_profissional(item.get("cpf")) if item["cpf_norm"] else (item.get("cpf") or "").strip()
    item["classificacao_vinculo"] = _vinculo_profissional_de_item(item)
    item["terceiro"] = item["classificacao_vinculo"] == "terceiro"
    item["vinculo_label"] = "Terceiro" if item["terceiro"] else "Contratado"
    tp = (_txt_cad(item.get("tipo_profissional")) or "operacional").lower()
    item["tipo_profissional"] = "motorista" if tp == "motorista" else "operacional"
    item["tipo_profissional_label"] = "Motorista" if item["tipo_profissional"] == "motorista" else "Operacional"
    item["cnh_numero"] = (item.get("cnh_numero") or item.get("cnh") or "").strip()
    item["cnh_vencimento"] = _fmt_data_cad(item.get("cnh_vencimento") or item.get("validade_cnh"))
    item["cnh_categoria"] = (item.get("cnh_categoria") or item.get("categoria_cnh") or "").strip()
    item["data_nascimento_fmt"] = _fmt_data_cad(item.get("data_nascimento"))
    item["status"] = _status_cad(item.get("status"))
    item["pode_receber_servicos"] = bool(item.get("pode_receber_servicos"))
    item["pode_aparecer_controle"] = bool(item.get("pode_aparecer_controle"))
    item["created_at"] = dt_str(item.get("created_at"))
    item["updated_at"] = dt_str(item.get("updated_at"))
    item["detalhes_registro"] = {
        "nome_completo": item["nome_completo"] or "-",
        "cpf": (item.get("cpf") or "-").strip(),
        "estado_civil": (item.get("estado_civil") or "-").strip(),
        "data_nascimento": item["data_nascimento_fmt"] or "-",
    }
    if com_arquivos:
        item["arquivos"] = listar_arquivos_profissional_cadastro(item["id"])
    item["qtd_arquivos"] = int(item.get("qtd_arquivos") or 0)
    return item


def resumo_viaturas_cadastro():
    row = one(
        """
        select
          count(*)::int as total,
          count(*) filter (where coalesce(status,'ativo')='ativo')::int as ativas,
          count(*) filter (where coalesce(status,'ativo')='inativo')::int as inativas,
          count(*) filter (where exists (
            select 1 from cadastro_viatura_arquivos a where a.viatura_id = cadastro_viaturas.id
          ))::int as com_documentos,
          count(*) filter (where coalesce(trim(personalizacao),'') <> '')::int as personalizadas,
          count(*) filter (
            where coalesce(origem, case when coalesce(terceiro,false) then 'terceiro' else 'propria' end) = 'propria'
          )::int as proprias,
          count(*) filter (
            where coalesce(origem, case when coalesce(terceiro,false) then 'terceiro' else 'propria' end) = 'terceiro'
          )::int as terceiros
        from cadastro_viaturas
        """
    ) or {}
    return {
        "total": int(row.get("total") or 0),
        "ativas": int(row.get("ativas") or 0),
        "inativas": int(row.get("inativas") or 0),
        "com_documentos": int(row.get("com_documentos") or 0),
        "personalizadas": int(row.get("personalizadas") or 0),
        "proprias": int(row.get("proprias") or 0),
        "terceiros": int(row.get("terceiros") or 0),
    }


def resumo_profissionais_cadastro():
    row = one(
        """
        select
          count(*)::int as total,
          count(*) filter (where coalesce(status,'ativo')='ativo')::int as ativos,
          count(*) filter (where coalesce(status,'ativo')='inativo')::int as inativos,
          count(*) filter (
            where coalesce(tipo_profissional, '') = 'motorista'
               or coalesce(funcao, '') ilike '%%motorista%%'
          )::int as motoristas,
          count(*) filter (
            where coalesce(tipo_profissional, '') = 'operacional'
               and coalesce(funcao, '') not ilike '%%motorista%%'
          )::int as operacionais,
          count(*) filter (
            where coalesce(classificacao_vinculo, case when coalesce(terceiro,false) then 'terceiro' else 'contratado' end) = 'contratado'
          )::int as contratados,
          count(*) filter (
            where coalesce(classificacao_vinculo, case when coalesce(terceiro,false) then 'terceiro' else 'contratado' end) = 'terceiro'
          )::int as terceiros
        from cadastro_profissionais
        """
    ) or {}
    return {
        "total": int(row.get("total") or 0),
        "ativos": int(row.get("ativos") or 0),
        "inativos": int(row.get("inativos") or 0),
        "motoristas": int(row.get("motoristas") or 0),
        "operacionais": int(row.get("operacionais") or 0),
        "contratados": int(row.get("contratados") or 0),
        "terceiros": int(row.get("terceiros") or 0),
    }


def _nome_exibicao_contato(item):
    if not item:
        return "-"
    if (item.get("tipo_pessoa") or "").lower() == "fisica":
        return (item.get("nome") or "").strip() or "-"
    return (
        (item.get("nome_fantasia") or "").strip()
        or (item.get("razao_social") or "").strip()
        or "-"
    )


def _documento_exibicao_contato(item):
    if not item:
        return ""
    if (item.get("tipo_pessoa") or "").lower() == "fisica":
        return (item.get("cpf") or "").strip()
    return (item.get("cnpj") or "").strip()


def _where_contatos_cadastro(cliente=None, fornecedor=None):
    parts = []
    params = []
    if cliente:
        parts.append("coalesce(cliente, false) = true")
    if fornecedor:
        parts.append("coalesce(fornecedor, false) = true")
    sql = (" where " + " and ".join(parts)) if parts else ""
    return sql, params


def normalizar_contato_cadastro(row, com_arquivos=False):
    if not row:
        return None
    item = dict(row)
    item["id"] = str(item.get("id"))
    item["tipo_pessoa"] = (item.get("tipo_pessoa") or "juridica").lower()
    item["status"] = _status_cad(item.get("status"))
    item["cliente"] = bool(item.get("cliente"))
    item["fornecedor"] = bool(item.get("fornecedor"))
    item["nome_exibicao"] = _nome_exibicao_contato(item)
    item["documento"] = _documento_exibicao_contato(item)
    item["limite_credito"] = float(item.get("limite_credito") or 0)
    item["data_nascimento_fmt"] = _fmt_data_cad(item.get("data_nascimento"))
    item["created_at_fmt"] = dt_str(item.get("created_at"))
    item["updated_at_fmt"] = dt_str(item.get("updated_at"))
    if com_arquivos:
        item["arquivos"] = listar_arquivos_contato_cadastro(item["id"])
    return item


def resumo_contatos_cadastro(cliente=False, fornecedor=False):
    where, _ = _where_contatos_cadastro(
        cliente=True if cliente else None,
        fornecedor=True if fornecedor else None,
    )
    row = one(
        f"""
        select
          count(*)::int as total,
          count(*) filter (where coalesce(status,'ativo')='ativo')::int as ativos,
          count(*) filter (
            where created_at >= date_trunc('month', current_date)
          )::int as novos_mes
        from cadastro_contatos
        {where}
        """
    ) or {}
    return {
        "total": int(row.get("total") or 0),
        "ativos": int(row.get("ativos") or 0),
        "novos_mes": int(row.get("novos_mes") or 0),
    }


def listar_contatos_cadastro(cliente=None, fornecedor=None, status=None):
    where, params = _where_contatos_cadastro(cliente=cliente, fornecedor=fornecedor)
    if status:
        where += (" and " if where else " where ") + "coalesce(status,'ativo')=%s"
        params.append(status)
    rows = q(
        f"select * from cadastro_contatos{where} order by coalesce(nome_fantasia, razao_social, nome) asc nulls last",
        tuple(params) if params else None,
        fetch=True,
    )
    return [normalizar_contato_cadastro(r) for r in rows]


def contato_cadastro_por_id(cid):
    row = one("select * from cadastro_contatos where id=%s", (str(cid),))
    return normalizar_contato_cadastro(row, com_arquivos=True) if row else None


def _params_contato_cadastro(p):
    p = p or {}
    tipo_pessoa = (p.get("tipo_pessoa") or "juridica").lower()
    if tipo_pessoa not in ("fisica", "juridica"):
        tipo_pessoa = "juridica"
    cliente = _bool_cad(p.get("cliente"))
    fornecedor = _bool_cad(p.get("fornecedor"))
    if not cliente and not fornecedor:
        raise ValueError("Marque Cliente e/ou Fornecedor")
    if tipo_pessoa == "fisica":
        nome = _txt_cad(p.get("nome"))
        if not nome:
            raise ValueError("Nome é obrigatório para pessoa física")
    else:
        razao = _txt_cad(p.get("razao_social"))
        if not razao:
            raise ValueError("Razão social é obrigatória para pessoa jurídica")
    return (
        tipo_pessoa,
        _status_cad(p.get("status")),
        _txt_cad(p.get("razao_social")),
        _txt_cad(p.get("nome_fantasia")),
        _txt_cad(p.get("cnpj")),
        _txt_cad(p.get("inscricao_estadual")),
        _txt_cad(p.get("contribuinte_icms")),
        _txt_cad(p.get("nome")),
        _txt_cad(p.get("cpf")),
        _txt_cad(p.get("rg")),
        _parse_data_controle(p.get("data_nascimento")),
        _txt_cad(p.get("cep")),
        _txt_cad(p.get("logradouro")),
        _txt_cad(p.get("numero")),
        _txt_cad(p.get("complemento")),
        _txt_cad(p.get("bairro")),
        _txt_cad(p.get("cidade")),
        _txt_cad(p.get("uf")),
        _txt_cad(p.get("email")),
        _txt_cad(p.get("email_financeiro")),
        _txt_cad(p.get("telefone")),
        _txt_cad(p.get("celular")),
        cliente,
        fornecedor,
        _num_controle(p.get("limite_credito")),
        _txt_cad(p.get("prazo_recebimento")),
        _txt_cad(p.get("observacoes_comercial_cliente")),
        _txt_cad(p.get("prazo_pagamento")),
        _txt_cad(p.get("categoria_fornecedor")),
        _txt_cad(p.get("observacoes_comercial_fornecedor")),
        _txt_cad(p.get("observacoes")),
    )


def salvar_contato_cadastro(payload):
    p = payload or {}
    cid = (p.get("id") or "").strip()
    params = _params_contato_cadastro(p)
    cols = """
      tipo_pessoa=%s, status=%s, razao_social=%s, nome_fantasia=%s, cnpj=%s,
      inscricao_estadual=%s, contribuinte_icms=%s, nome=%s, cpf=%s, rg=%s,
      data_nascimento=%s, cep=%s, logradouro=%s, numero=%s, complemento=%s,
      bairro=%s, cidade=%s, uf=%s, email=%s, email_financeiro=%s,
      telefone=%s, celular=%s, cliente=%s, fornecedor=%s,
      limite_credito=%s, prazo_recebimento=%s, observacoes_comercial_cliente=%s,
      prazo_pagamento=%s, categoria_fornecedor=%s, observacoes_comercial_fornecedor=%s,
      observacoes=%s, updated_at=now()
    """
    if cid:
        q(f"update cadastro_contatos set {cols} where id=%s", params + (cid,))
        return contato_cadastro_por_id(cid)
    row = one(
        f"""
        insert into cadastro_contatos (
          tipo_pessoa, status, razao_social, nome_fantasia, cnpj,
          inscricao_estadual, contribuinte_icms, nome, cpf, rg,
          data_nascimento, cep, logradouro, numero, complemento,
          bairro, cidade, uf, email, email_financeiro,
          telefone, celular, cliente, fornecedor,
          limite_credito, prazo_recebimento, observacoes_comercial_cliente,
          prazo_pagamento, categoria_fornecedor, observacoes_comercial_fornecedor,
          observacoes, updated_at
        ) values ({",".join(["%s"] * len(params))}, now())
        returning id
        """,
        params,
    )
    return contato_cadastro_por_id(row["id"])


def excluir_contato_cadastro(cid):
    q("delete from cadastro_contato_arquivos where contato_id=%s", (str(cid),))
    q("delete from cadastro_contatos where id=%s", (str(cid),))


def listar_arquivos_contato_cadastro(cid):
    rows = q(
        "select * from cadastro_contato_arquivos where contato_id=%s order by created_at desc",
        (str(cid),),
        fetch=True,
    )
    out = []
    for r in rows:
        item = dict(r)
        item["id"] = str(item["id"])
        item["contato_id"] = str(item["contato_id"])
        item["nome_arquivo"] = item.get("nome_arquivo") or item.get("nome") or item.get("filename") or "-"
        item["tipo"] = item.get("tipo") or item.get("tipo_documento") or "-"
        item["caminho_arquivo"] = item.get("caminho_arquivo") or item.get("url") or ""
        item["created_at_fmt"] = dt_str(item.get("created_at"))
        out.append(item)
    return out


def excluir_arquivo_contato_cadastro(aid):
    row = one("select caminho_arquivo, url from cadastro_contato_arquivos where id=%s", (str(aid),))
    if row:
        path_url = row.get("caminho_arquivo") or row.get("url") or ""
        if path_url.startswith("/static/"):
            fpath = path_url.lstrip("/").replace("/", os.sep)
            if os.path.isfile(fpath):
                try:
                    os.remove(fpath)
                except OSError:
                    pass
    q("delete from cadastro_contato_arquivos where id=%s", (str(aid),))


async def salvar_arquivo_contato_cadastro(cid, form, arquivo):
    if not arquivo or not getattr(arquivo, "filename", None):
        raise ValueError("Arquivo obrigatório")
    pasta = os.path.join(CADASTROS_UPLOAD_DIR, "contatos", str(cid))
    os.makedirs(pasta, exist_ok=True)
    ext = os.path.splitext(arquivo.filename)[1].lower()
    fname = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(pasta, fname)
    content = await arquivo.read()
    with open(path, "wb") as f:
        f.write(content)
    caminho = f"/static/uploads/cadastros/contatos/{cid}/{fname}"
    nome_arq = _txt_cad(form.get("nome")) or arquivo.filename
    tipo_doc = _txt_cad(form.get("tipo") or form.get("tipo_documento")) or "Outros"
    row = one(
        """
        insert into cadastro_contato_arquivos (
          contato_id, nome_arquivo, extensao, tipo, caminho_arquivo,
          nome, tipo_documento, data_documento, filename, url
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) returning *
        """,
        (
            str(cid),
            nome_arq,
            ext.lstrip("."),
            tipo_doc,
            caminho,
            nome_arq,
            tipo_doc,
            _parse_data_controle(form.get("data_documento")),
            fname,
            caminho,
        ),
    )
    return dict(row)


def opcoes_contatos_cadastro(papel=None):
    cliente = (papel or "").lower() == "cliente"
    fornecedor = (papel or "").lower() == "fornecedor"
    itens = listar_contatos_cadastro(
        cliente=True if cliente else None,
        fornecedor=True if fornecedor else None,
        status="ativo",
    )
    return [
        {
            "id": i["id"],
            "nome": i["nome_exibicao"],
            "documento": i.get("documento") or "",
            "cidade": i.get("cidade") or "",
        }
        for i in itens
    ]


def listar_viaturas_cadastro():
    rows = q(
        """
        select v.*,
               (select count(*)::int from cadastro_viatura_arquivos a where a.viatura_id = v.id) as qtd_arquivos
          from cadastro_viaturas v
         order by coalesce(v.exibicao, v.placa, v.modelo) asc nulls last
        """,
        fetch=True,
    )
    return [normalizar_viatura_cadastro(r) for r in rows]


def viatura_cadastro_por_id(vid):
    row = one(
        """
        select v.*,
               (select count(*)::int from cadastro_viatura_arquivos a where a.viatura_id = v.id) as qtd_arquivos
          from cadastro_viaturas v
         where v.id = %s
        """,
        (str(vid),),
    )
    return normalizar_viatura_cadastro(row, com_arquivos=True) if row else None


def viatura_cadastro_por_placa(placa):
    placa_n = normalizar_placa_viatura(placa)
    if not placa_n:
        return None
    row = one(
        """
        select v.*,
               (select count(*)::int from cadastro_viatura_arquivos a where a.viatura_id = v.id) as qtd_arquivos
          from cadastro_viaturas v
         where upper(replace(replace(replace(coalesce(v.placa,''), ' ', ''), '-', ''), '.', '')) = %s
         limit 1
        """,
        (placa_n,),
    )
    return normalizar_viatura_cadastro(row) if row else None


def salvar_viatura_cadastro(payload):
    p = payload or {}
    vid = (p.get("id") or "").strip()
    placa = normalizar_placa_viatura(p.get("placa")) or (_txt_cad(p.get("placa"), 12) or "").upper()
    if not placa:
        raise ValueError("Placa é obrigatória")
    exibicao = _txt_cad(p.get("exibicao") or p.get("nome_exibicao")) or placa
    origem_in = (_txt_cad(p.get("origem")) or "").lower()
    if origem_in in ("propria", "terceiro"):
        origem = origem_in
        terceiro = origem == "terceiro"
    else:
        mapped = mapear_origem_viatura(p.get("classificacao_original"))
        origem = mapped["origem"]
        terceiro = mapped["terceiro"]
    classificacao_original = _txt_cad(p.get("classificacao_original")) or mapear_origem_viatura(origem)["classificacao_original"]
    params = (
        placa,
        _txt_cad(p.get("marca")),
        _txt_cad(p.get("modelo")),
        _txt_cad(p.get("renavam")),
        _txt_cad(p.get("chassi")),
        _txt_cad(p.get("ano_fabricacao"), 8),
        _txt_cad(p.get("ano_modelo"), 8),
        _txt_cad(p.get("combustivel")),
        _num_controle(p.get("capacidade_litros")) or None,
        _txt_cad(p.get("cor")),
        _txt_cad(p.get("estado_placa")),
        _txt_cad(p.get("tipo_viatura")),
        _txt_cad(p.get("observacoes")),
        _status_cad(p.get("status")),
        _txt_cad(p.get("cpf_cnpj_crlv")),
        exibicao,
        _txt_cad(p.get("telefone")),
        _txt_cad(p.get("personalizacao")),
        _num_controle(p.get("consumo_km_l")) or None,
        _int_controle(p.get("hodometro")) or None,
        terceiro,
        classificacao_original,
        origem,
    )
    if vid:
        q(
            """
            update cadastro_viaturas set
              placa=%s, marca=%s, modelo=%s, renavam=%s, chassi=%s,
              ano_fabricacao=%s, ano_modelo=%s, combustivel=%s, capacidade_litros=%s,
              cor=%s, estado_placa=%s, tipo_viatura=%s, observacoes=%s, status=%s,
              cpf_cnpj_crlv=%s, exibicao=%s, telefone=%s, personalizacao=%s,
              consumo_km_l=%s, hodometro=%s, terceiro=%s, classificacao_original=%s, origem=%s,
              updated_at=now()
            where id=%s
            """,
            params + (vid,),
        )
        return viatura_cadastro_por_id(vid)
    existente = viatura_cadastro_por_placa(placa)
    if existente:
        p2 = dict(p)
        p2["id"] = existente["id"]
        return salvar_viatura_cadastro(p2)
    row = one(
        """
        insert into cadastro_viaturas (
          placa, marca, modelo, renavam, chassi, ano_fabricacao, ano_modelo,
          combustivel, capacidade_litros, cor, estado_placa, tipo_viatura, observacoes,
          status, cpf_cnpj_crlv, exibicao, telefone, personalizacao,
          consumo_km_l, hodometro, terceiro, classificacao_original, origem, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,now())
        returning id
        """,
        params,
    )
    return viatura_cadastro_por_id(row["id"])


def excluir_viatura_cadastro(vid):
    q("delete from cadastro_viatura_arquivos where viatura_id=%s", (str(vid),))
    q("delete from cadastro_viaturas where id=%s", (str(vid),))


def listar_arquivos_viatura_cadastro(vid):
    rows = q(
        "select * from cadastro_viatura_arquivos where viatura_id=%s order by created_at desc",
        (str(vid),),
        fetch=True,
    )
    out = []
    for r in rows:
        item = dict(r)
        item["id"] = str(item["id"])
        item["viatura_id"] = str(item["viatura_id"])
        if item.get("data_documento") and hasattr(item["data_documento"], "strftime"):
            item["data_documento"] = item["data_documento"].strftime("%Y-%m-%d")
        out.append(item)
    return out


def listar_profissionais_cadastro():
    rows = q("select * from cadastro_profissionais order by nome asc nulls last", fetch=True)
    return [normalizar_profissional_cadastro(r) for r in rows]


def profissional_cadastro_por_id(pid):
    row = one("select * from cadastro_profissionais where id=%s", (str(pid),))
    return normalizar_profissional_cadastro(row, com_arquivos=True) if row else None


def profissional_cadastro_por_cpf(cpf):
    cpf_n = normalizar_cpf_profissional(cpf)
    if not cpf_n:
        return None
    row = one(
        """
        select * from cadastro_profissionais
         where regexp_replace(coalesce(cpf,''), '[^0-9]', '', 'g') = %s
         limit 1
        """,
        (cpf_n,),
    )
    return normalizar_profissional_cadastro(row) if row else None


def profissional_cadastro_por_nome_telefone(nome_completo, telefone):
    nome_n = normalizar_nome_completo_profissional(nome_completo)
    tel_n = normalizar_telefone_profissional(telefone)
    if not nome_n or not tel_n:
        return None
    row = one(
        """
        select * from cadastro_profissionais
         where upper(trim(coalesce(nome_completo, nome))) = %s
           and regexp_replace(coalesce(telefone_movel, telefone, ''), '[^0-9]', '', 'g') = %s
         limit 1
        """,
        (nome_n, tel_n),
    )
    return normalizar_profissional_cadastro(row) if row else None


def profissional_cadastro_por_motorista_id(mid):
    if not mid:
        return None
    row = one(
        "select * from cadastro_profissionais where motorista_id=%s limit 1",
        (str(mid),),
    )
    return normalizar_profissional_cadastro(row) if row else None


def _sync_motorista_profissional(prof_id, payload):
    """Cria/atualiza registro em motoristas para profissionais do tipo motorista."""
    prof = profissional_cadastro_por_id(prof_id)
    if not prof:
        return
    if (prof.get("tipo_profissional") or "").lower() != "motorista":
        return
    nome_exib = prof.get("nome_exibicao") or prof.get("nome_trabalho") or prof.get("nome_completo")
    tel = prof.get("telefone_norm") or normalizar_telefone_profissional(prof.get("telefone"))
    cpf = prof.get("cpf_norm") or normalizar_cpf_profissional(prof.get("cpf"))
    cnh = (prof.get("cnh_numero") or "").strip()
    mid = prof.get("motorista_id")
    if mid:
        q(
            """
            update motoristas set
              nome=%s, telefone=%s, cpf=%s, cnh=%s, tipo=%s, ativo=%s, ultima_atualizacao=now()
             where id=%s
            """,
            (
                nome_exib,
                formatar_telefone_profissional(tel) if tel else None,
                cpf or None,
                cnh or None,
                prof.get("funcao") or "Motorista",
                prof.get("status") == "ativo",
                str(mid),
            ),
        )
        return
    row = one(
        """
        insert into motoristas (nome, telefone, cpf, cnh, tipo, ativo, online, ultima_atualizacao)
        values (%s, %s, %s, %s, %s, %s, false, now())
        returning id
        """,
        (
            nome_exib,
            formatar_telefone_profissional(tel) if tel else None,
            cpf or None,
            cnh or None,
            prof.get("funcao") or "Motorista",
            prof.get("status") == "ativo",
        ),
    )
    if row:
        q("update cadastro_profissionais set motorista_id=%s, updated_at=now() where id=%s", (str(row["id"]), str(prof_id)))


def _params_profissional_cadastro(p):
    p = p or {}
    nome_completo = normalizar_nome_completo_profissional(p.get("nome_completo") or p.get("nome"))
    if not nome_completo:
        raise ValueError("Nome completo é obrigatório")
    nome_trabalho = normalizar_nome_trabalho_profissional(
        p.get("nome_trabalho") or p.get("nome_exibicao") or nome_completo
    )
    nome_legacy = nome_trabalho
    vinculo_in = (_txt_cad(p.get("classificacao_vinculo")) or "").lower()
    if vinculo_in in ("contratado", "terceiro"):
        vinculo = vinculo_in
        terceiro = vinculo == "terceiro"
    else:
        mapped = mapear_classificacao_vinculo_profissional(p.get("classificacao_original") or p.get("classificacao"))
        vinculo = mapped["classificacao_vinculo"]
        terceiro = mapped["terceiro"]
    tipo_in = (_txt_cad(p.get("tipo_profissional")) or "").lower()
    if tipo_in in ("motorista", "operacional"):
        tipo_prof = tipo_in
    else:
        infer = inferir_tipo_e_funcao_profissional(p.get("funcao"), p.get("categoria"), nome_trabalho)
        tipo_prof = infer["tipo_profissional"]
    funcao = _txt_cad(p.get("funcao")) or inferir_tipo_e_funcao_profissional(
        p.get("funcao"), p.get("categoria"), nome_trabalho
    ).get("funcao")
    cnh_numero = _txt_cad(p.get("cnh_numero") or p.get("cnh"))
    cnh_venc = _parse_data_controle(p.get("cnh_vencimento") or p.get("validade_cnh"))
    cnh_cat = _txt_cad(p.get("cnh_categoria") or p.get("categoria_cnh"))
    tel_movel_raw = normalizar_telefone_profissional(p.get("telefone_movel") or p.get("telefone"))
    tel_movel = tel_movel_raw or None
    tel_fixo_raw = normalizar_telefone_profissional(p.get("telefone_fixo"))
    tel_fixo = tel_fixo_raw or None
    cpf = normalizar_cpf_profissional(p.get("cpf")) or None
    remuneracao = _txt_cad(p.get("remuneracao"))
    if remuneracao is None and p.get("comissao_padrao") not in (None, ""):
        remuneracao = str(p.get("comissao_padrao"))
    return (
        _txt_cad(p.get("filial_cnpj")),
        nome_completo,
        nome_trabalho,
        _parse_data_controle(p.get("data_nascimento")),
        cpf,
        _txt_cad(p.get("rg")),
        tel_movel,
        tel_fixo,
        _txt_cad(p.get("estado_civil")),
        _txt_cad(p.get("email")),
        _txt_cad(p.get("cep")),
        _txt_cad(p.get("logradouro")),
        _txt_cad(p.get("bairro")),
        _txt_cad(p.get("cidade")),
        _txt_cad(p.get("uf")),
        _txt_cad(p.get("observacoes")),
        terceiro,
        _txt_cad(p.get("cnpj")),
        funcao,
        remuneracao,
        _txt_cad(p.get("forma_pagamento")),
        _parse_data_controle(p.get("data_admissao")),
        _parse_data_controle(p.get("data_demissao")),
        _txt_cad(p.get("hora_inicio")),
        _txt_cad(p.get("hora_termino")),
        _txt_cad(p.get("carga_horaria")),
        _txt_cad(p.get("intervalo")),
        _txt_cad(p.get("escala")),
        _txt_cad(p.get("registro_ctps")),
        cnh_numero,
        cnh_venc,
        cnh_cat,
        _status_cad(p.get("status")),
        nome_legacy,
        cnh_numero,
        cnh_cat,
        cnh_venc,
        tel_movel,
        _txt_cad(p.get("endereco")),
        tipo_prof,
        _num_controle(p.get("comissao_padrao")),
        _bool_cad(p.get("pode_receber_servicos")),
        _bool_cad(p.get("pode_aparecer_controle")),
        (p.get("motorista_id") or "").strip() or None,
        vinculo,
    )


def salvar_profissional_cadastro(payload):
    p = payload or {}
    pid = (p.get("id") or "").strip()
    if not pid:
        cpf_busca = normalizar_cpf_profissional(p.get("cpf"))
        if cpf_busca:
            existente = profissional_cadastro_por_cpf(cpf_busca)
            if existente:
                pid = existente["id"]
        if not pid:
            existente = profissional_cadastro_por_nome_telefone(
                p.get("nome_completo") or p.get("nome"),
                p.get("telefone_movel") or p.get("telefone"),
            )
            if existente:
                pid = existente["id"]
    if pid:
        p = dict(p)
        p["id"] = pid
    params = _params_profissional_cadastro(p)
    cols_update = """
      filial_cnpj=%s, nome_completo=%s, nome_trabalho=%s, data_nascimento=%s,
      cpf=%s, rg=%s, telefone_movel=%s, telefone_fixo=%s, estado_civil=%s, email=%s,
      cep=%s, logradouro=%s, bairro=%s, cidade=%s, uf=%s, observacoes=%s,
      terceiro=%s, cnpj=%s, funcao=%s, remuneracao=%s, forma_pagamento=%s,
      data_admissao=%s, data_demissao=%s, hora_inicio=%s, hora_termino=%s,
      carga_horaria=%s, intervalo=%s, escala=%s, registro_ctps=%s,
      cnh_numero=%s, cnh_vencimento=%s, cnh_categoria=%s, status=%s,
      nome=%s, cnh=%s, categoria_cnh=%s, validade_cnh=%s, telefone=%s, endereco=%s,
      tipo_profissional=%s, comissao_padrao=%s, pode_receber_servicos=%s,
      pode_aparecer_controle=%s, motorista_id=%s, classificacao_vinculo=%s, updated_at=now()
    """
    if pid:
        q(f"update cadastro_profissionais set {cols_update} where id=%s", params + (pid,))
        _sync_motorista_profissional(pid, p)
        return profissional_cadastro_por_id(pid)
    row = one(
        f"""
        insert into cadastro_profissionais (
          filial_cnpj, nome_completo, nome_trabalho, data_nascimento,
          cpf, rg, telefone_movel, telefone_fixo, estado_civil, email,
          cep, logradouro, bairro, cidade, uf, observacoes,
          terceiro, cnpj, funcao, remuneracao, forma_pagamento,
          data_admissao, data_demissao, hora_inicio, hora_termino,
          carga_horaria, intervalo, escala, registro_ctps,
          cnh_numero, cnh_vencimento, cnh_categoria, status,
          nome, cnh, categoria_cnh, validade_cnh, telefone, endereco,
          tipo_profissional, comissao_padrao, pode_receber_servicos,
          pode_aparecer_controle, motorista_id, classificacao_vinculo, updated_at
        ) values ({",".join(["%s"] * len(params))}, now())
        returning id
        """,
        params,
    )
    new_id = str(row["id"])
    _sync_motorista_profissional(new_id, p)
    return profissional_cadastro_por_id(new_id)


def excluir_profissional_cadastro(pid):
    q("delete from cadastro_profissional_arquivos where profissional_id=%s", (str(pid),))
    q("delete from cadastro_profissionais where id=%s", (str(pid),))


def listar_arquivos_profissional_cadastro(pid):
    rows = q(
        "select * from cadastro_profissional_arquivos where profissional_id=%s order by created_at desc",
        (str(pid),),
        fetch=True,
    )
    out = []
    for r in rows:
        item = dict(r)
        item["id"] = str(item["id"])
        item["profissional_id"] = str(item["profissional_id"])
        item["nome_arquivo"] = item.get("nome_arquivo") or item.get("nome") or item.get("filename") or "-"
        item["tipo"] = item.get("tipo") or item.get("tipo_documento") or "-"
        item["caminho_arquivo"] = item.get("caminho_arquivo") or item.get("url") or ""
        item["extensao"] = (item.get("extensao") or "").strip().lower()
        if not item["extensao"] and item.get("filename"):
            item["extensao"] = os.path.splitext(item["filename"])[1].lstrip(".").lower()
        item["created_at_fmt"] = dt_str(item.get("created_at"))
        out.append(item)
    return out


def excluir_arquivo_profissional_cadastro(aid):
    row = one("select caminho_arquivo, url, filename from cadastro_profissional_arquivos where id=%s", (str(aid),))
    if row:
        path_url = row.get("caminho_arquivo") or row.get("url") or ""
        if path_url.startswith("/static/"):
            fpath = path_url.lstrip("/").replace("/", os.sep)
            if os.path.isfile(fpath):
                try:
                    os.remove(fpath)
                except OSError:
                    pass
    q("delete from cadastro_profissional_arquivos where id=%s", (str(aid),))


async def salvar_arquivo_viatura_cadastro(vid, form, arquivo: UploadFile):
    if not arquivo or not arquivo.filename:
        raise ValueError("Arquivo obrigatório")
    pasta = os.path.join(CADASTROS_UPLOAD_DIR, "viaturas", str(vid))
    os.makedirs(pasta, exist_ok=True)
    ext = os.path.splitext(arquivo.filename)[1].lower()
    fname = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(pasta, fname)
    content = await arquivo.read()
    with open(path, "wb") as f:
        f.write(content)
    url = f"/static/uploads/cadastros/viaturas/{vid}/{fname}"
    row = one(
        """
        insert into cadastro_viatura_arquivos (viatura_id, nome, tipo_documento, data_documento, filename, url)
        values (%s,%s,%s,%s,%s,%s) returning *
        """,
        (
            str(vid),
            _txt_cad(form.get("nome")) or arquivo.filename,
            _txt_cad(form.get("tipo_documento")),
            _parse_data_controle(form.get("data_documento")),
            fname,
            url,
        ),
    )
    return dict(row)


async def salvar_arquivo_profissional_cadastro(pid, form, arquivo):
    if not arquivo or not getattr(arquivo, "filename", None):
        raise ValueError("Arquivo obrigatório")
    pasta = os.path.join(CADASTROS_UPLOAD_DIR, "profissionais", str(pid))
    os.makedirs(pasta, exist_ok=True)
    ext = os.path.splitext(arquivo.filename)[1].lower()
    fname = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(pasta, fname)
    content = await arquivo.read()
    with open(path, "wb") as f:
        f.write(content)
    caminho = f"/static/uploads/cadastros/profissionais/{pid}/{fname}"
    nome_arq = _txt_cad(form.get("nome")) or arquivo.filename
    tipo_doc = _txt_cad(form.get("tipo") or form.get("tipo_documento")) or "Outros"
    row = one(
        """
        insert into cadastro_profissional_arquivos (
          profissional_id, nome_arquivo, extensao, tipo, caminho_arquivo,
          nome, tipo_documento, data_documento, filename, url
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) returning *
        """,
        (
            str(pid),
            nome_arq,
            ext.lstrip("."),
            tipo_doc,
            caminho,
            nome_arq,
            tipo_doc,
            _parse_data_controle(form.get("data_documento")),
            fname,
            caminho,
        ),
    )
    return dict(row)


@app.get("/api/cadastros/viaturas")
def api_cadastros_lista_viaturas():
    return {"ok": True, "itens": listar_viaturas_cadastro(), "kpis": resumo_viaturas_cadastro()}


@app.get("/api/cadastros/viaturas/{vid}")
def api_cadastros_get_viatura(vid: str):
    item = viatura_cadastro_por_id(vid)
    if not item:
        return JSONResponse({"ok": False, "erro": "Viatura não encontrada"}, status_code=404)
    return {"ok": True, "item": item}


@app.post("/api/cadastros/viaturas")
async def api_cadastros_post_viatura(request: Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    try:
        item = salvar_viatura_cadastro(payload)
        return {"ok": True, "item": item, "kpis": resumo_viaturas_cadastro()}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.delete("/api/cadastros/viaturas/{vid}")
def api_cadastros_delete_viatura(vid: str):
    try:
        excluir_viatura_cadastro(vid)
        return {"ok": True, "kpis": resumo_viaturas_cadastro()}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.post("/api/cadastros/viaturas/{vid}/arquivos")
async def api_cadastros_upload_viatura(vid: str, request: Request):
    form = await request.form()
    arquivo = form.get("arquivo")
    try:
        row = await salvar_arquivo_viatura_cadastro(vid, form, arquivo)
        item = viatura_cadastro_por_id(vid)
        return {"ok": True, "arquivo": row, "item": item}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.get("/api/cadastros/profissionais")
def api_cadastros_lista_profissionais():
    return {"ok": True, "itens": listar_profissionais_cadastro(), "kpis": resumo_profissionais_cadastro()}


@app.get("/api/cadastros/profissionais/{pid}")
def api_cadastros_get_profissional(pid: str):
    item = profissional_cadastro_por_id(pid)
    if not item:
        return JSONResponse({"ok": False, "erro": "Profissional não encontrado"}, status_code=404)
    return {"ok": True, "item": item}


@app.post("/api/cadastros/profissionais")
async def api_cadastros_post_profissional(request: Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    try:
        item = salvar_profissional_cadastro(payload)
        return {"ok": True, "item": item, "kpis": resumo_profissionais_cadastro()}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.put("/api/cadastros/profissionais/{pid}")
async def api_cadastros_put_profissional(pid: str, request: Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    payload["id"] = pid
    try:
        item = salvar_profissional_cadastro(payload)
        return {"ok": True, "item": item, "kpis": resumo_profissionais_cadastro()}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.delete("/api/cadastros/profissionais/{pid}")
def api_cadastros_delete_profissional(pid: str):
    try:
        excluir_profissional_cadastro(pid)
        return {"ok": True, "kpis": resumo_profissionais_cadastro()}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.get("/api/cadastros/profissionais/{pid}/arquivos")
def api_cadastros_lista_arquivos_profissional(pid: str):
    return {"ok": True, "itens": listar_arquivos_profissional_cadastro(pid)}


@app.post("/api/cadastros/profissionais/{pid}/arquivos")
async def api_cadastros_upload_profissional(pid: str, request: Request):
    form = await request.form()
    arquivo = form.get("arquivo")
    try:
        row = await salvar_arquivo_profissional_cadastro(pid, form, arquivo)
        return {"ok": True, "arquivo": row, "itens": listar_arquivos_profissional_cadastro(pid)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.delete("/api/cadastros/profissionais/{pid}/arquivos/{aid}")
def api_cadastros_delete_arquivo_profissional(pid: str, aid: str):
    try:
        excluir_arquivo_profissional_cadastro(aid)
        return {"ok": True, "itens": listar_arquivos_profissional_cadastro(pid)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.get("/api/cadastros/contatos")
def api_cadastros_lista_contatos(
    cliente: str = "",
    fornecedor: str = "",
    status: str = "",
):
    filtro_c = cliente.lower() in ("1", "true", "sim", "s")
    filtro_f = fornecedor.lower() in ("1", "true", "sim", "s")
    itens = listar_contatos_cadastro(
        cliente=True if filtro_c else None,
        fornecedor=True if filtro_f else None,
        status=status or None,
    )
    kpis = resumo_contatos_cadastro(
        cliente=filtro_c or None,
        fornecedor=filtro_f or None,
    )
    return {"ok": True, "itens": itens, "kpis": kpis}


@app.get("/api/cadastros/contatos/opcoes")
def api_cadastros_opcoes_contatos(papel: str = ""):
    return {"ok": True, "itens": opcoes_contatos_cadastro(papel)}


@app.get("/api/cadastros/contatos/{cid}")
def api_cadastros_get_contato(cid: str):
    item = contato_cadastro_por_id(cid)
    if not item:
        return JSONResponse({"ok": False, "erro": "Contato não encontrado"}, status_code=404)
    return {"ok": True, "item": item}


@app.post("/api/cadastros/contatos")
async def api_cadastros_post_contato(request: Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    try:
        filtro_c = bool(payload.get("filtro_cliente"))
        filtro_f = bool(payload.get("filtro_fornecedor"))
        item = salvar_contato_cadastro(payload)
        kpis = resumo_contatos_cadastro(cliente=filtro_c or None, fornecedor=filtro_f or None)
        return {"ok": True, "item": item, "kpis": kpis}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.put("/api/cadastros/contatos/{cid}")
async def api_cadastros_put_contato(cid: str, request: Request):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    payload["id"] = cid
    try:
        filtro_c = bool(payload.get("filtro_cliente"))
        filtro_f = bool(payload.get("filtro_fornecedor"))
        item = salvar_contato_cadastro(payload)
        kpis = resumo_contatos_cadastro(cliente=filtro_c or None, fornecedor=filtro_f or None)
        return {"ok": True, "item": item, "kpis": kpis}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.delete("/api/cadastros/contatos/{cid}")
def api_cadastros_delete_contato(cid: str, cliente: str = "", fornecedor: str = ""):
    try:
        excluir_contato_cadastro(cid)
        filtro_c = cliente.lower() in ("1", "true", "sim", "s")
        filtro_f = fornecedor.lower() in ("1", "true", "sim", "s")
        kpis = resumo_contatos_cadastro(cliente=filtro_c or None, fornecedor=filtro_f or None)
        return {"ok": True, "kpis": kpis}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.get("/api/cadastros/contatos/{cid}/arquivos")
def api_cadastros_lista_arquivos_contato(cid: str):
    return {"ok": True, "itens": listar_arquivos_contato_cadastro(cid)}


@app.post("/api/cadastros/contatos/{cid}/arquivos")
async def api_cadastros_upload_contato(cid: str, request: Request):
    form = await request.form()
    arquivo = form.get("arquivo")
    try:
        row = await salvar_arquivo_contato_cadastro(cid, form, arquivo)
        return {"ok": True, "arquivo": row, "itens": listar_arquivos_contato_cadastro(cid)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.delete("/api/cadastros/contatos/{cid}/arquivos/{aid}")
def api_cadastros_delete_arquivo_contato(cid: str, aid: str):
    try:
        excluir_arquivo_contato_cadastro(aid)
        return {"ok": True, "itens": listar_arquivos_contato_cadastro(cid)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


def opcoes_profissionais_controle():
    rows = q(
        """
        select id,
               coalesce(nullif(trim(nome_trabalho),''), nullif(trim(nome_completo),''), nome) as nome
          from cadastro_profissionais
         where coalesce(status, 'ativo') = 'ativo'
           and coalesce(pode_aparecer_controle, true) = true
           and coalesce(trim(coalesce(nome_completo, nome)), '') <> ''
         order by nome asc
        """,
        fetch=True,
    )
    return [
        {"id": str(r["id"]), "nome": (r.get("nome") or "").strip(), "status": "ativo"}
        for r in rows
    ]


def opcoes_motoristas_legado():
    rows = q(
        """
        select id, nome
          from motoristas
         where coalesce(ativo, true) = true
           and coalesce(trim(nome), '') <> ''
         order by nome asc
        """,
        fetch=True,
    )
    return [
        {"id": str(r["id"]), "nome": (r.get("nome") or "").strip(), "status": "ativo"}
        for r in rows
    ]


def opcoes_motoristas_controle():
    """Profissionais ativos aptos a receber serviços ou função de motorista."""
    rows = q(
        """
        select id,
               coalesce(nullif(trim(nome_trabalho),''), nullif(trim(nome_completo),''), nome) as nome,
               nome_completo, nome_trabalho, funcao, tipo_profissional, classificacao_vinculo, terceiro
          from cadastro_profissionais
         where coalesce(status, 'ativo') = 'ativo'
           and (
             coalesce(pode_receber_servicos, true) = true
             or coalesce(funcao, '') ilike '%%motorista%%'
             or coalesce(tipo_profissional, '') ilike '%%motorista%%'
           )
           and coalesce(trim(coalesce(nome_completo, nome)), '') <> ''
         order by nome asc
        """,
        fetch=True,
    )
    out = []
    for r in rows:
        item = normalizar_profissional_cadastro(r)
        out.append({
            "id": item["id"],
            "nome": item.get("nome_exibicao") or item.get("nome"),
            "nome_exibicao": item.get("nome_exibicao"),
            "nome_completo": item.get("nome_completo"),
            "funcao": item.get("funcao"),
            "tipo_profissional": item.get("tipo_profissional"),
            "tipo_profissional_label": item.get("tipo_profissional_label"),
            "classificacao_vinculo": item.get("classificacao_vinculo"),
            "vinculo_label": item.get("vinculo_label"),
            "status": "ativo",
        })
    if out:
        return out
    return opcoes_motoristas_legado()


def opcoes_viaturas_controle():
    rows = q(
        """
        select id, placa, marca, modelo, exibicao, tipo_viatura, personalizacao, origem, terceiro, classificacao_original
          from cadastro_viaturas
         where coalesce(status, 'ativo') = 'ativo'
         order by coalesce(exibicao, placa) asc nulls last
        """,
        fetch=True,
    )
    out = []
    for r in rows:
        item = normalizar_viatura_cadastro(r)
        placa = item.get("placa") or ""
        nome = item.get("nome_exibicao") or item.get("modelo") or placa
        label = f"{nome} — {placa}" if placa and nome and nome != placa else (nome or placa or "Sem identificação")
        if item.get("personalizacao"):
            label += f" [{item['personalizacao']}]"
        out.append({
            "id": item["id"],
            "placa": placa,
            "nome": nome,
            "nome_exibicao": nome,
            "status": "ativo",
            "label": label,
            "origem": item.get("origem"),
            "origem_label": item.get("origem_label"),
            "personalizacao": item.get("personalizacao"),
            "personalizada": item.get("personalizada"),
            "classificacao_original": item.get("classificacao_original"),
        })
    if out:
        return out
    rows = q(
        """
        select id, placa, modelo, tipo
          from veiculos
         where coalesce(ativo, true) = true
         order by placa asc nulls last
        """,
        fetch=True,
    )
    for r in rows:
        placa = (r.get("placa") or "").strip().upper()
        modelo = (r.get("modelo") or "").strip()
        tipo = (r.get("tipo") or "").strip()
        nome = modelo or tipo or placa
        label = f"{placa} — {nome}" if placa and nome and nome != placa else (placa or nome)
        out.append({"id": str(r["id"]), "placa": placa, "nome": nome, "status": "ativo", "label": label})
    return out


@app.get("/api/opcoes/motoristas")
def api_opcoes_motoristas():
    return opcoes_motoristas_controle()


@app.get("/api/opcoes/profissionais")
def api_opcoes_profissionais():
    return opcoes_profissionais_controle()


@app.get("/api/opcoes/viaturas")
def api_opcoes_viaturas():
    return opcoes_viaturas_controle()


# ============================================================
# FROTA — Abastecimentos (premium)
# ============================================================

FROTA_POSTOS_ABASTECIMENTO = ("INTERNO", "BOMBA EXTERNA", "BOMBA INTERNA")
FROTA_COMBUSTIVEIS = ("DIESEL S10", "DIESEL S500", "GASOLINA", "ETANOL", "ARLA 32")


def _label_viatura_abastecimento(v):
    if not v:
        return ""
    exib = (v.get("nome_exibicao") or v.get("exibicao") or v.get("modelo") or "").strip()
    placa = (v.get("placa") or "").strip()
    if exib and placa:
        return f"{exib} — {placa}"
    return exib or placa or ""


def _fornecedor_pagar_abastecimento(posto):
    p = (posto or "").strip().upper()
    if p == "INTERNO":
        return "Posto Interno"
    return (posto or "").strip() or "Posto"


def _criar_conta_pagar_abastecimento(abast):
    row_cat = one(
        """
        select id, descricao from financeiro_categorias
         where natureza='Despesa'
           and (descricao ilike '%%Combust%%' or descricao ilike '%%Abastec%%')
         order by case when descricao ilike '%%Combust%%' then 0 else 1 end, ordem asc
         limit 1
        """
    )
    categoria_id = str(row_cat["id"]) if row_cat else None
    categoria = (row_cat.get("descricao") if row_cat else None) or "Combustível"
    viatura_label = (abast.get("viatura") or abast.get("placa") or "").strip()
    combustivel = (abast.get("combustivel") or "").strip()
    litros = float(abast.get("litros") or 0)
    litros_txt = f"{litros:.3f}".rstrip("0").rstrip(".")
    descricao = f"Abastecimento - {viatura_label} - {combustivel} - {litros_txt}L"
    venc = abast.get("data_abastecimento")
    if hasattr(venc, "isoformat"):
        venc = venc
    elif isinstance(venc, str) and venc:
        venc = _parse_data_controle(venc)
    else:
        venc = None
    valor = float(abast.get("valor_total") or 0)
    if valor <= 0:
        raise ValueError("Valor total inválido para contas a pagar")
    row = one(
        """
        insert into financeiro_contas_pagar (
          fornecedor, descricao, categoria, categoria_id, competencia, emissao, vencimento,
          valor, parcelas, status, observacoes, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,1,'em_aberto',%s,now())
        returning id
        """,
        (
            _fornecedor_pagar_abastecimento(abast.get("posto")),
            descricao,
            categoria,
            categoria_id,
            venc,
            venc,
            venc,
            valor,
            f"Gerado automaticamente do abastecimento {abast.get('id') or ''}".strip(),
        ),
    )
    return str(row["id"])


def ultimo_abastecimento_viatura_frota(viatura_id, excluir_id=None):
    if not viatura_id:
        return None
    params = [str(viatura_id)]
    sql = """
        select id, hodometro, data_hora, data_abastecimento, viatura, placa
          from controle_abastecimentos
         where coalesce(status, 'ativo') = 'ativo'
           and viatura_id = %s
           and hodometro is not null
    """
    if excluir_id:
        sql += " and id <> %s"
        params.append(str(excluir_id))
    sql += """
         order by coalesce(data_hora, data_abastecimento::timestamp) desc nulls last,
                  hodometro desc nulls last
         limit 1
    """
    return one(sql, tuple(params))


def _calcular_km_media_abastecimento(hodometro_atual, litros, ultimo, desconsiderar=False):
    hod = _int_controle(hodometro_atual)
    lit = _num_controle(litros)
    km_rodado = None
    media = None
    alerta = ""
    ult_hod = None
    if ultimo and ultimo.get("hodometro") is not None:
        ult_hod = _int_controle(ultimo.get("hodometro"))
        if hod is not None and ult_hod is not None:
            km_rodado = hod - ult_hod
            if km_rodado < 0 and not desconsiderar:
                alerta = (
                    f"Hodômetro informado ({hod}) é menor que o último registrado ({ult_hod}) "
                    f"para esta viatura."
                )
            elif km_rodado >= 0 and lit > 0:
                media = round(km_rodado / lit, 3)
            elif km_rodado < 0 and desconsiderar:
                km_rodado = None
                media = None
    return {
        "km_rodado": km_rodado,
        "media_km_litro": media,
        "alerta_hodometro": alerta,
        "ultimo_hodometro": ult_hod,
    }


def normalizar_abastecimento_frota(row):
    if not row:
        return None
    item = dict(row)
    item["id"] = str(item.get("id"))
    if item.get("viatura_id"):
        item["viatura_id"] = str(item["viatura_id"])
    if item.get("profissional_id"):
        item["profissional_id"] = str(item["profissional_id"])
    if item.get("contas_pagar_id"):
        item["contas_pagar_id"] = str(item["contas_pagar_id"])
    item["litros"] = float(item.get("litros") or 0)
    item["valor_unitario"] = float(item.get("valor_unitario") or 0)
    item["valor_total"] = float(item.get("valor_total") or 0)
    item["hodometro"] = _int_controle(item.get("hodometro")) if item.get("hodometro") is not None else None
    item["km_rodado"] = _int_controle(item.get("km_rodado")) if item.get("km_rodado") is not None else None
    item["media_km_litro"] = float(item["media_km_litro"]) if item.get("media_km_litro") is not None else None
    item["desconsiderar_ultimo_km"] = bool(item.get("desconsiderar_ultimo_km"))
    item["gerar_contas_pagar"] = bool(item.get("gerar_contas_pagar"))
    item["status"] = (item.get("status") or "ativo").lower()
    dh = item.get("data_hora")
    if dh and hasattr(dh, "strftime"):
        item["data_hora_fmt"] = dh.strftime("%d/%m/%Y %H:%M")
        item["data_hora_iso"] = dh.strftime("%Y-%m-%dT%H:%M")
    else:
        item["data_hora_fmt"] = _formatar_valor_controle("data_hora", dh)
        item["data_hora_iso"] = str(dh)[:16] if dh else ""
    item["litros_fmt"] = f"{item['litros']:.3f}".rstrip("0").rstrip(".")
    item["valor_unitario_fmt"] = fmt_moeda(item["valor_unitario"])
    item["valor_total_fmt"] = fmt_moeda(item["valor_total"])
    item["media_km_litro_fmt"] = f"{item['media_km_litro']:.2f}" if item.get("media_km_litro") is not None else "-"
    item["km_rodado_fmt"] = str(item["km_rodado"]) if item.get("km_rodado") is not None else "-"
    item["hodometro_fmt"] = str(item["hodometro"]) if item.get("hodometro") is not None else "-"
    item["posto_interno"] = (item.get("posto") or "").strip().upper() == "INTERNO"
    item["created_at"] = dt_str(item.get("created_at"))
    item["updated_at"] = dt_str(item.get("updated_at"))
    return item


def _where_abastecimentos_frota(filtros=None):
    filtros = filtros or {}
    parts = ["coalesce(a.status, 'ativo') <> 'cancelado'"]
    params = []
    if filtros.get("data_ini"):
        parts.append("coalesce(a.data_hora, a.data_abastecimento::timestamp)::date >= %s")
        params.append(filtros["data_ini"])
    if filtros.get("data_fim"):
        parts.append("coalesce(a.data_hora, a.data_abastecimento::timestamp)::date <= %s")
        params.append(filtros["data_fim"])
    if filtros.get("viatura_id"):
        parts.append("a.viatura_id = %s")
        params.append(str(filtros["viatura_id"]))
    if filtros.get("profissional_id"):
        parts.append("a.profissional_id = %s")
        params.append(str(filtros["profissional_id"]))
    if filtros.get("posto"):
        parts.append("upper(trim(coalesce(a.posto,''))) = upper(trim(%s))")
        params.append(filtros["posto"])
    if filtros.get("combustivel"):
        parts.append("upper(trim(coalesce(a.combustivel,''))) = upper(trim(%s))")
        params.append(filtros["combustivel"])
    if filtros.get("busca"):
        parts.append(
            """
            (
              coalesce(a.viatura,'') ilike %s
              or coalesce(a.placa,'') ilike %s
              or coalesce(a.profissional,'') ilike %s
              or coalesce(a.posto,'') ilike %s
              or coalesce(a.combustivel,'') ilike %s
              or coalesce(a.observacoes,'') ilike %s
            )
            """
        )
        b = f"%{filtros['busca']}%"
        params.extend([b, b, b, b, b, b])
    where = " where " + " and ".join(parts) if parts else ""
    return where, params


def resumo_abastecimentos_frota(filtros=None):
    where, params = _where_abastecimentos_frota(filtros)
    row = one(
        f"""
        select
          count(*)::int as total,
          coalesce(sum(litros), 0)::float as litros,
          coalesce(sum(valor_total), 0)::float as valor_total,
          coalesce(sum(km_rodado), 0)::float as km_total,
          count(*) filter (where upper(trim(coalesce(posto,''))) = 'INTERNO')::int as internos,
          count(*) filter (where upper(trim(coalesce(posto,''))) <> 'INTERNO')::int as externos
        from controle_abastecimentos a
        {where}
        """,
        tuple(params) if params else None,
    ) or {}
    litros = float(row.get("litros") or 0)
    km_total = float(row.get("km_total") or 0)
    media_geral = round(km_total / litros, 2) if litros > 0 and km_total > 0 else 0
    return {
        "total": int(row.get("total") or 0),
        "litros": round(litros, 3),
        "valor_total": round(float(row.get("valor_total") or 0), 2),
        "media_km_litro": media_geral,
        "internos": int(row.get("internos") or 0),
        "externos": int(row.get("externos") or 0),
    }


def listar_abastecimentos_frota(filtros=None):
    where, params = _where_abastecimentos_frota(filtros)
    rows = q(
        f"""
        select a.*
          from controle_abastecimentos a
        {where}
         order by coalesce(a.data_hora, a.data_abastecimento::timestamp) desc nulls last,
                  a.created_at desc nulls last
         limit 1000
        """,
        tuple(params) if params else None,
        fetch=True,
    )
    return [normalizar_abastecimento_frota(r) for r in rows]


def abastecimento_frota_por_id(aid):
    row = one("select * from controle_abastecimentos where id=%s", (str(aid),))
    return normalizar_abastecimento_frota(row) if row else None


def _validar_payload_abastecimento_frota(p, aid=None):
    p = p or {}
    viatura_id = (p.get("viatura_id") or "").strip()
    if not viatura_id:
        raise ValueError("Viatura é obrigatória")
    viatura = viatura_cadastro_por_id(viatura_id)
    if not viatura:
        raise ValueError("Viatura não encontrada")
    data_hora = _parse_datetime_controle(p.get("data_hora"))
    if not data_hora:
        raise ValueError("Data e hora são obrigatórias")
    posto = (_txt_controle(p.get("posto"), 80) or "").upper()
    if posto not in FROTA_POSTOS_ABASTECIMENTO:
        raise ValueError("Posto inválido")
    profissional_id = (p.get("profissional_id") or "").strip()
    if not profissional_id:
        raise ValueError("Profissional é obrigatório")
    prof = profissional_cadastro_por_id(profissional_id)
    if not prof:
        raise ValueError("Profissional não encontrado")
    combustivel = (_txt_controle(p.get("combustivel"), 80) or "").upper()
    if combustivel not in FROTA_COMBUSTIVEIS:
        raise ValueError("Combustível inválido")
    litros = _num_controle(p.get("litros"))
    valor_unit = _num_controle(p.get("valor_unitario"))
    valor_total = _num_controle(p.get("valor_total"))
    if litros <= 0:
        raise ValueError("Litros abastecidos deve ser maior que zero")
    if valor_unit <= 0:
        raise ValueError("Valor unitário deve ser maior que zero")
    if valor_total <= 0:
        raise ValueError("Valor total deve ser maior que zero")
    desconsiderar = _bool_controle(p.get("desconsiderar_ultimo_km"))
    hod_raw = p.get("hodometro")
    hod_str = str(hod_raw or "").strip()
    if not hod_str:
        if not desconsiderar:
            raise ValueError("Hodômetro (KM) é obrigatório")
        hodometro = None
    else:
        hodometro = _int_controle(hod_raw)
    ultimo = ultimo_abastecimento_viatura_frota(viatura_id, excluir_id=aid)
    km_info = _calcular_km_media_abastecimento(hodometro, litros, ultimo, desconsiderar)
    if km_info.get("alerta_hodometro") and not desconsiderar:
        raise ValueError(km_info["alerta_hodometro"])
    viatura_label = _label_viatura_abastecimento(viatura)
    profissional_label = prof.get("nome_exibicao") or prof.get("nome_trabalho") or prof.get("nome_completo")
    return {
        "viatura_id": viatura_id,
        "viatura": viatura_label,
        "placa": viatura.get("placa") or "",
        "profissional_id": profissional_id,
        "profissional": profissional_label,
        "data_hora": data_hora,
        "data_abastecimento": data_hora.date(),
        "posto": posto,
        "combustivel": combustivel,
        "litros": litros,
        "valor_unitario": valor_unit,
        "valor_total": valor_total,
        "hodometro": hodometro,
        "observacoes": _txt_controle(p.get("observacao") or p.get("observacoes")),
        "desconsiderar_ultimo_km": desconsiderar,
        "gerar_contas_pagar": _bool_controle(p.get("gerar_contas_pagar")),
        "km_rodado": km_info.get("km_rodado"),
        "media_km_litro": km_info.get("media_km_litro"),
        "alerta_hodometro": km_info.get("alerta_hodometro") or None,
        "status": _status_cad(p.get("status")) if p.get("status") else "ativo",
    }


def salvar_abastecimento_frota(payload):
    p = payload or {}
    aid = (p.get("id") or "").strip()
    dados = _validar_payload_abastecimento_frota(p, aid=aid or None)
    params = (
        dados["data_abastecimento"],
        dados["data_hora"],
        dados["viatura"],
        dados["placa"],
        dados["viatura_id"],
        dados["posto"],
        dados["profissional"],
        dados["profissional_id"],
        dados["combustivel"],
        dados["litros"],
        dados["valor_unitario"],
        dados["valor_total"],
        dados["hodometro"],
        dados["km_rodado"],
        dados["media_km_litro"],
        dados["observacoes"],
        dados["desconsiderar_ultimo_km"],
        dados["gerar_contas_pagar"],
        dados["alerta_hodometro"],
        dados["status"],
    )
    if aid:
        q(
            """
            update controle_abastecimentos set
              data_abastecimento=%s, data_hora=%s, viatura=%s, placa=%s, viatura_id=%s,
              posto=%s, profissional=%s, profissional_id=%s, combustivel=%s,
              litros=%s, valor_unitario=%s, valor_total=%s, hodometro=%s,
              km_rodado=%s, media_km_litro=%s, observacoes=%s,
              desconsiderar_ultimo_km=%s, gerar_contas_pagar=%s, alerta_hodometro=%s,
              status=%s, updated_at=now()
            where id=%s
            """,
            params + (aid,),
        )
        item = abastecimento_frota_por_id(aid)
    else:
        row = one(
            """
            insert into controle_abastecimentos (
              data_abastecimento, data_hora, viatura, placa, viatura_id,
              posto, profissional, profissional_id, combustivel,
              litros, valor_unitario, valor_total, hodometro,
              km_rodado, media_km_litro, observacoes,
              desconsiderar_ultimo_km, gerar_contas_pagar, alerta_hodometro,
              status, updated_at
            ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,now())
            returning id
            """,
            params,
        )
        item = abastecimento_frota_por_id(row["id"])
    if dados["gerar_contas_pagar"] and not item.get("contas_pagar_id"):
        cp_id = _criar_conta_pagar_abastecimento(item)
        q(
            "update controle_abastecimentos set contas_pagar_id=%s, updated_at=now() where id=%s",
            (cp_id, item["id"]),
        )
        item = abastecimento_frota_por_id(item["id"])
    return item


def excluir_abastecimento_frota(aid):
    q(
        "update controle_abastecimentos set status='cancelado', updated_at=now() where id=%s",
        (str(aid),),
    )


def opcoes_viaturas_frota_abastecimento():
    rows = q(
        """
        select id, placa, exibicao, marca, modelo, status
          from cadastro_viaturas
         where coalesce(status, 'ativo') = 'ativo'
         order by coalesce(exibicao, placa) asc nulls last
        """,
        fetch=True,
    )
    out = []
    for r in rows:
        v = normalizar_viatura_cadastro(r)
        label = _label_viatura_abastecimento(v)
        out.append({
            "id": v["id"],
            "placa": v.get("placa") or "",
            "nome_exibicao": v.get("nome_exibicao") or v.get("exibicao") or "",
            "label": label,
        })
    return out


def _labels_profissional_operacional(p):
    nome = (
        p.get("nome_exibicao")
        or p.get("nome_trabalho")
        or p.get("nome_completo")
        or p.get("nome")
        or ""
    ).strip()
    funcao = (p.get("funcao") or "Profissional").strip()
    vinc = (p.get("vinculo_label") or p.get("classificacao_vinculo") or "").strip()
    label_operacional = f"{nome} — {funcao}" if nome else funcao
    label_admin = f"{label_operacional} ({vinc})" if vinc else label_operacional
    return {
        "nome_exibicao": nome,
        "funcao": funcao,
        "label_operacional": label_operacional,
        "label_admin": label_admin,
    }


def _label_profissional_frota(p):
    return _labels_profissional_operacional(p)["label_operacional"]


def opcoes_profissionais_frota_abastecimento():
    rows = q(
        """
        select id, nome_trabalho, nome_completo, funcao, classificacao_vinculo, terceiro, status
          from cadastro_profissionais
         where coalesce(status, 'ativo') = 'ativo'
         order by coalesce(nullif(trim(nome_trabalho),''), nome_completo) asc nulls last
        """,
        fetch=True,
    )
    out = []
    for r in rows:
        p = normalizar_profissional_cadastro(r)
        labels = _labels_profissional_operacional(p)
        out.append({
            "id": p["id"],
            "nome_exibicao": labels["nome_exibicao"],
            "nome_completo": p.get("nome_completo") or "",
            "funcao": labels["funcao"],
            "vinculo_label": p.get("vinculo_label") or "",
            "label": labels["label_operacional"],
            "label_operacional": labels["label_operacional"],
            "label_admin": labels["label_admin"],
        })
    return out


FROTA_MANUTENCAO_STATUS = ("pendente", "finalizada", "cancelada")
FROTA_MANUTENCAO_TIPOS_ITEM = ("ESTOQUE", "PRODUTO", "SERVICO", "SERVIÇO")


def _status_manutencao_frota(v):
    s = (_txt_cad(v) or "finalizada").lower()
    if s == "ativo":
        return "finalizada"
    if s in FROTA_MANUTENCAO_STATUS:
        return s
    return "finalizada"


def _label_status_manutencao(st):
    m = {"pendente": "Pendente", "finalizada": "Finalizada", "cancelada": "Cancelada"}
    return m.get(st, st.capitalize() if st else "Finalizada")


def ultimo_hodometro_viatura_frota(viatura_id, excluir_manutencao_id=None):
    if not viatura_id:
        return None
    params = [str(viatura_id), str(viatura_id)]
    excl_sql = ""
    if excluir_manutencao_id:
        excl_sql = " and not (origem = 'manutencao' and registro_id = %s)"
        params.append(str(excluir_manutencao_id))
    row = one(
        f"""
        select hodometro, origem, registro_id from (
          select hodometro, 'abastecimento' as origem, id as registro_id,
                 coalesce(data_hora, data_abastecimento::timestamp) as dt
            from controle_abastecimentos
           where viatura_id = %s
             and hodometro is not null
             and coalesce(status, 'ativo') <> 'cancelado'
          union all
          select hodometro, 'manutencao', id,
                 coalesce(data_manutencao::timestamp, created_at)
            from controle_manutencoes
           where viatura_id = %s
             and hodometro is not null
             and coalesce(status, 'finalizada') not in ('cancelada', 'cancelado')
        ) u
        where hodometro is not null {excl_sql}
        order by hodometro desc, dt desc nulls last
        limit 1
        """,
        tuple(params),
    )
    if not row:
        return None
    return {
        "hodometro": _int_controle(row.get("hodometro")),
        "origem": row.get("origem"),
        "registro_id": str(row["registro_id"]) if row.get("registro_id") else None,
    }


def _normalizar_itens_manutencao(raw_itens):
    itens = raw_itens or []
    if isinstance(itens, str):
        try:
            itens = json.loads(itens)
        except Exception:
            itens = []
    out = []
    for it in itens:
        if not isinstance(it, dict):
            continue
        tipo = (_txt_cad(it.get("tipo")) or "PRODUTO").upper().replace("Ç", "C")
        if tipo not in ("ESTOQUE", "PRODUTO", "SERVICO"):
            tipo = "PRODUTO"
        nome_item = _txt_cad(it.get("item") or it.get("descricao"), 300)
        if not nome_item:
            continue
        qtd = _num_controle(it.get("quantidade"), 1) or 1
        if qtd <= 0:
            continue
        vu = _num_controle(it.get("valor_unitario"))
        if vu < 0:
            vu = 0
        vt = _num_controle(it.get("valor_total"), qtd * vu)
        if vt < 0:
            vt = 0
        out.append({
            "tipo": tipo,
            "item": nome_item,
            "descricao": nome_item,
            "valor_unitario": vu,
            "quantidade": qtd,
            "valor_total": vt,
            "observacao": _txt_cad(it.get("observacao") or it.get("obs"), 500) or "",
        })
    return out


def _totais_manutencao_itens(itens):
    total_prod = 0.0
    total_serv = 0.0
    for it in itens or []:
        t = (it.get("tipo") or "").upper().replace("Ç", "C")
        v = float(it.get("valor_total") or 0)
        if t == "SERVICO":
            total_serv += v
        else:
            total_prod += v
    return round(total_prod, 2), round(total_serv, 2), round(total_prod + total_serv, 2)


def _criar_conta_pagar_manutencao(manut):
    row_cat = one(
        """
        select id, descricao from financeiro_categorias
         where natureza='Despesa'
           and descricao ilike '%%Manuten%%'
         order by ordem asc limit 1
        """
    )
    categoria_id = str(row_cat["id"]) if row_cat else None
    categoria = (row_cat.get("descricao") if row_cat else None) or "Manutenção"
    viatura_label = (manut.get("viatura") or manut.get("placa") or "").strip()
    documento = (manut.get("documento") or "").strip()
    doc_txt = f" - Documento {documento}" if documento else ""
    descricao = f"Manutenção - {viatura_label}{doc_txt}"
    venc = manut.get("data_manutencao")
    if isinstance(venc, str):
        venc = _parse_data_controle(venc)
    valor = float(manut.get("total") or 0)
    if valor <= 0:
        raise ValueError("Valor total inválido para contas a pagar")
    row = one(
        """
        insert into financeiro_contas_pagar (
          fornecedor, descricao, categoria, categoria_id, competencia, emissao, vencimento,
          valor, parcelas, status, observacoes, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,1,'em_aberto',%s,now())
        returning id
        """,
        (
            (manut.get("fornecedor") or "Fornecedor").strip(),
            descricao,
            categoria,
            categoria_id,
            venc,
            venc,
            venc,
            valor,
            f"Gerado automaticamente da manutenção {manut.get('id') or ''}".strip(),
        ),
    )
    return str(row["id"])


def normalizar_manutencao_frota(row):
    if not row:
        return None
    item = dict(row)
    item["id"] = str(item.get("id"))
    if item.get("viatura_id"):
        item["viatura_id"] = str(item["viatura_id"])
    if item.get("fornecedor_id"):
        item["fornecedor_id"] = str(item["fornecedor_id"])
    if item.get("contas_pagar_id"):
        item["contas_pagar_id"] = str(item["contas_pagar_id"])
    item["status"] = _status_manutencao_frota(item.get("status"))
    item["status_label"] = _label_status_manutencao(item["status"])
    raw_itens = item.get("itens")
    if isinstance(raw_itens, str):
        try:
            raw_itens = json.loads(raw_itens)
        except Exception:
            raw_itens = []
    item["itens"] = _normalizar_itens_manutencao(raw_itens)
    tp, ts, tg = _totais_manutencao_itens(item["itens"])
    item["total_produtos"] = float(item.get("total_produtos") or tp)
    item["total_servicos"] = float(item.get("total_servicos") or ts)
    item["total"] = float(item.get("total") or tg)
    item["hodometro"] = _int_controle(item.get("hodometro")) if item.get("hodometro") is not None else None
    item["gerar_contas_pagar"] = bool(item.get("gerar_contas_pagar"))
    item["hodometro_menor_ultimo"] = bool(item.get("hodometro_menor_ultimo"))
    dm = item.get("data_manutencao")
    if dm and hasattr(dm, "strftime"):
        item["data_fmt"] = dm.strftime("%d/%m/%Y")
        item["data_iso"] = dm.strftime("%Y-%m-%d")
    else:
        item["data_fmt"] = _formatar_valor_controle("data_manutencao", dm)
        item["data_iso"] = str(dm)[:10] if dm else ""
    item["total_produtos_fmt"] = fmt_moeda(item["total_produtos"])
    item["total_servicos_fmt"] = fmt_moeda(item["total_servicos"])
    item["total_fmt"] = fmt_moeda(item["total"])
    item["hodometro_fmt"] = str(item["hodometro"]) if item.get("hodometro") is not None else "-"
    item["created_at"] = dt_str(item.get("created_at"))
    item["updated_at"] = dt_str(item.get("updated_at"))
    return item


def _where_manutencoes_frota(filtros=None):
    filtros = filtros or {}
    parts = ["coalesce(m.status, 'finalizada') not in ('cancelada', 'cancelado')"]
    params = []
    if filtros.get("data_ini"):
        parts.append("m.data_manutencao >= %s")
        params.append(filtros["data_ini"])
    if filtros.get("data_fim"):
        parts.append("m.data_manutencao <= %s")
        params.append(filtros["data_fim"])
    if filtros.get("viatura_id"):
        parts.append("m.viatura_id = %s")
        params.append(str(filtros["viatura_id"]))
    if filtros.get("fornecedor"):
        parts.append("coalesce(m.fornecedor,'') ilike %s")
        params.append(f"%{filtros['fornecedor']}%")
    if filtros.get("documento"):
        parts.append("coalesce(m.documento,'') ilike %s")
        params.append(f"%{filtros['documento']}%")
    if filtros.get("status"):
        parts.append("coalesce(m.status, 'finalizada') = %s")
        params.append(filtros["status"])
    if filtros.get("busca"):
        parts.append(
            """
            (
              coalesce(m.viatura,'') ilike %s
              or coalesce(m.placa,'') ilike %s
              or coalesce(m.fornecedor,'') ilike %s
              or coalesce(m.documento,'') ilike %s
              or coalesce(m.observacoes,'') ilike %s
            )
            """
        )
        b = f"%{filtros['busca']}%"
        params.extend([b, b, b, b, b])
    where = " where " + " and ".join(parts) if parts else ""
    return where, params


def resumo_manutencoes_frota(filtros=None):
    where, params = _where_manutencoes_frota(filtros)
    row = one(
        f"""
        select
          count(*)::int as total,
          coalesce(sum(total), 0)::float as valor_total,
          coalesce(sum(total_produtos), 0)::float as total_produtos,
          coalesce(sum(total_servicos), 0)::float as total_servicos,
          count(*) filter (
            where date_trunc('month', coalesce(data_manutencao, created_at::date))
              = date_trunc('month', current_date)
          )::int as mes
        from controle_manutencoes m
        {where}
        """,
        tuple(params) if params else None,
    ) or {}
    total = int(row.get("total") or 0)
    valor = float(row.get("valor_total") or 0)
    return {
        "total": total,
        "valor_total": round(valor, 2),
        "total_produtos": round(float(row.get("total_produtos") or 0), 2),
        "total_servicos": round(float(row.get("total_servicos") or 0), 2),
        "media_manutencao": round(valor / total, 2) if total > 0 else 0,
        "mes": int(row.get("mes") or 0),
    }


def listar_manutencoes_frota(filtros=None):
    where, params = _where_manutencoes_frota(filtros)
    rows = q(
        f"""
        select m.* from controle_manutencoes m
        {where}
         order by coalesce(m.data_manutencao, m.created_at::date) desc nulls last,
                  m.created_at desc nulls last
         limit 1000
        """,
        tuple(params) if params else None,
        fetch=True,
    )
    return [normalizar_manutencao_frota(r) for r in rows]


def manutencao_frota_por_id(mid):
    row = one("select * from controle_manutencoes where id=%s", (str(mid),))
    return normalizar_manutencao_frota(row) if row else None


def opcoes_fornecedores_frota():
    rows = q(
        """
        select id,
               coalesce(nullif(trim(nome_fantasia),''), nullif(trim(razao_social),''), nullif(trim(nome),'')) as nome
          from cadastro_contatos
         where coalesce(fornecedor, false) = true
           and coalesce(status, 'ativo') = 'ativo'
         order by nome asc nulls last
        """,
        fetch=True,
    )
    out = []
    for r in rows:
        nome = (r.get("nome") or "").strip()
        if not nome:
            continue
        out.append({"id": str(r["id"]), "nome": nome, "label": nome})
    return out


def _validar_payload_manutencao_frota(p, mid=None):
    p = p or {}
    viatura_id = (p.get("viatura_id") or "").strip()
    if not viatura_id:
        raise ValueError("Viatura é obrigatória")
    viatura = viatura_cadastro_por_id(viatura_id)
    if not viatura:
        raise ValueError("Viatura não encontrada")
    data_man = _parse_data_controle(p.get("data") or p.get("data_manutencao"))
    if not data_man:
        raise ValueError("Data é obrigatória")
    hod_str = str(p.get("hodometro") or "").strip()
    if not hod_str:
        raise ValueError("Hodômetro (KM) é obrigatório")
    hodometro = _int_controle(hod_str)
    fornecedor = _txt_controle(p.get("fornecedor"), 200)
    if not fornecedor:
        raise ValueError("Fornecedor é obrigatório")
    itens = _normalizar_itens_manutencao(p.get("itens"))
    if not itens:
        raise ValueError("Adicione ao menos um item à manutenção")
    total_prod, total_serv, total_geral = _totais_manutencao_itens(itens)
    ultimo = ultimo_hodometro_viatura_frota(viatura_id, excluir_manutencao_id=mid)
    alerta = ""
    hodometro_menor = False
    if ultimo and ultimo.get("hodometro") is not None and hodometro is not None:
        if hodometro < ultimo["hodometro"]:
            hodometro_menor = True
            alerta = (
                f"Hodômetro informado ({hodometro}) é menor que o último registrado "
                f"({ultimo['hodometro']} km — origem: {ultimo.get('origem', '-')})."
            )
    status = _status_manutencao_frota(p.get("status") or "finalizada")
    obs = _txt_controle(p.get("observacao") or p.get("observacoes"))
    if hodometro_menor and alerta:
        extra = f"[Alerta hodômetro] {alerta}"
        obs = f"{obs}\n{extra}".strip() if obs else extra
    return {
        "viatura_id": viatura_id,
        "viatura": _label_viatura_abastecimento(viatura),
        "placa": viatura.get("placa") or "",
        "data_manutencao": data_man,
        "hodometro": hodometro,
        "documento": _txt_controle(p.get("documento"), 120),
        "fornecedor": fornecedor,
        "fornecedor_id": (p.get("fornecedor_id") or "").strip() or None,
        "observacoes": obs,
        "itens": itens,
        "total_produtos": total_prod,
        "total_servicos": total_serv,
        "total": total_geral,
        "gerar_contas_pagar": _bool_controle(p.get("gerar_contas_pagar")),
        "alerta_hodometro": alerta or None,
        "hodometro_menor_ultimo": hodometro_menor,
        "status": status,
    }


def salvar_manutencao_frota(payload):
    p = payload or {}
    mid = (p.get("id") or "").strip()
    dados = _validar_payload_manutencao_frota(p, mid=mid or None)
    params = (
        dados["data_manutencao"],
        dados["viatura"],
        dados["placa"],
        dados["viatura_id"],
        dados["hodometro"],
        dados["documento"],
        dados["fornecedor"],
        dados["fornecedor_id"],
        dados["observacoes"],
        Json(dados["itens"]),
        dados["total_produtos"],
        dados["total_servicos"],
        dados["total"],
        dados["gerar_contas_pagar"],
        dados["alerta_hodometro"],
        dados["hodometro_menor_ultimo"],
        dados["status"],
    )
    if mid:
        q(
            """
            update controle_manutencoes set
              data_manutencao=%s, viatura=%s, placa=%s, viatura_id=%s, hodometro=%s,
              documento=%s, fornecedor=%s, fornecedor_id=%s, observacoes=%s,
              itens=%s, total_produtos=%s, total_servicos=%s, total=%s,
              gerar_contas_pagar=%s, alerta_hodometro=%s, hodometro_menor_ultimo=%s,
              status=%s, updated_at=now()
            where id=%s
            """,
            params + (mid,),
        )
        item = manutencao_frota_por_id(mid)
    else:
        row = one(
            """
            insert into controle_manutencoes (
              data_manutencao, viatura, placa, viatura_id, hodometro,
              documento, fornecedor, fornecedor_id, observacoes,
              itens, total_produtos, total_servicos, total,
              gerar_contas_pagar, alerta_hodometro, hodometro_menor_ultimo,
              status, updated_at
            ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,now())
            returning id
            """,
            params,
        )
        item = manutencao_frota_por_id(row["id"])
    if dados["gerar_contas_pagar"] and not item.get("contas_pagar_id"):
        cp_id = _criar_conta_pagar_manutencao(item)
        q(
            "update controle_manutencoes set contas_pagar_id=%s, updated_at=now() where id=%s",
            (cp_id, item["id"]),
        )
        item = manutencao_frota_por_id(item["id"])
    return item


def excluir_manutencao_frota(mid):
    q(
        "update controle_manutencoes set status='cancelada', updated_at=now() where id=%s",
        (str(mid),),
    )


def _parse_filtros_manutencoes_request(request):
    qparams = request.query_params if request else {}
    return {
        "data_ini": _parse_data_controle(qparams.get("data_ini")),
        "data_fim": _parse_data_controle(qparams.get("data_fim")),
        "viatura_id": (qparams.get("viatura_id") or "").strip() or None,
        "fornecedor": (qparams.get("fornecedor") or "").strip() or None,
        "documento": (qparams.get("documento") or "").strip() or None,
        "status": (qparams.get("status") or "").strip().lower() or None,
        "busca": (qparams.get("busca") or "").strip() or None,
    }


def _parse_filtros_abastecimentos_request(request):
    qparams = request.query_params if request else {}
    return {
        "data_ini": _parse_data_controle(qparams.get("data_ini")),
        "data_fim": _parse_data_controle(qparams.get("data_fim")),
        "viatura_id": (qparams.get("viatura_id") or "").strip() or None,
        "profissional_id": (qparams.get("profissional_id") or "").strip() or None,
        "posto": (qparams.get("posto") or "").strip() or None,
        "combustivel": (qparams.get("combustivel") or "").strip() or None,
        "busca": (qparams.get("busca") or "").strip() or None,
    }


@app.get("/frota/abastecimentos", response_class=HTMLResponse)
def frota_pagina_abastecimentos(request: Request):
    return templates.TemplateResponse(
        "frota/abastecimentos.html",
        {
            "request": request,
            "nav_ativo": "frota",
            "nav_som": False,
            "kpis": resumo_abastecimentos_frota(),
            "postos": FROTA_POSTOS_ABASTECIMENTO,
            "combustiveis": FROTA_COMBUSTIVEIS,
        },
    )


@app.get("/frota/manutencoes", response_class=HTMLResponse)
def frota_pagina_manutencoes(request: Request):
    return templates.TemplateResponse(
        "frota/manutencoes.html",
        {
            "request": request,
            "nav_ativo": "frota",
            "nav_som": False,
            "kpis": resumo_manutencoes_frota(),
            "status_opcoes": FROTA_MANUTENCAO_STATUS,
            "tipos_item": ("ESTOQUE", "PRODUTO", "SERVIÇO"),
        },
    )


@app.get("/frota/controle-patio", response_class=HTMLResponse)
def frota_pagina_controle_patio(request: Request):
    return templates.TemplateResponse(
        "frota/controle_patio.html",
        {
            "request": request,
            "nav_ativo": "frota",
            "nav_som": False,
            "kpis": resumo_patio_frota(),
            "status_opcoes": FROTA_PATIO_STATUS,
            "responsavel_padrao": "Operação",
        },
    )


@app.get("/api/frota/abastecimentos")
def api_frota_lista_abastecimentos(request: Request):
    filtros = _parse_filtros_abastecimentos_request(request)
    return {
        "ok": True,
        "itens": listar_abastecimentos_frota(filtros),
        "kpis": resumo_abastecimentos_frota(filtros),
    }


@app.get("/api/frota/abastecimentos/opcoes/viaturas")
def api_frota_opcoes_viaturas_abastecimento():
    return {"ok": True, "itens": opcoes_viaturas_frota_abastecimento()}


@app.get("/api/frota/abastecimentos/opcoes/profissionais")
def api_frota_opcoes_profissionais_abastecimento():
    return {"ok": True, "itens": opcoes_profissionais_frota_abastecimento()}


@app.get("/api/frota/abastecimentos/ultimo-km")
def api_frota_ultimo_km_abastecimento(viatura_id: str, hodometro: str = "", excluir_id: str = ""):
    ultimo = ultimo_abastecimento_viatura_frota(viatura_id, excluir_id=excluir_id or None)
    desconsiderar = False
    km_info = _calcular_km_media_abastecimento(hodometro, 1, ultimo, desconsiderar)
    return {
        "ok": True,
        "ultimo_hodometro": km_info.get("ultimo_hodometro"),
        "ultimo_id": str(ultimo["id"]) if ultimo else None,
        "alerta": km_info.get("alerta_hodometro") or "",
    }


@app.get("/api/frota/abastecimentos/{aid}")
def api_frota_get_abastecimento(aid: str):
    item = abastecimento_frota_por_id(aid)
    if not item:
        return JSONResponse({"ok": False, "erro": "Abastecimento não encontrado"}, status_code=404)
    return {"ok": True, "item": item}


@app.post("/api/frota/abastecimentos")
async def api_frota_post_abastecimento(request: Request):
    try:
        payload = await request.json()
        item = salvar_abastecimento_frota(payload)
        filtros = _parse_filtros_abastecimentos_request(request)
        return {"ok": True, "item": item, "kpis": resumo_abastecimentos_frota(filtros)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.put("/api/frota/abastecimentos/{aid}")
async def api_frota_put_abastecimento(aid: str, request: Request):
    try:
        payload = await request.json()
        payload["id"] = aid
        item = salvar_abastecimento_frota(payload)
        filtros = _parse_filtros_abastecimentos_request(request)
        return {"ok": True, "item": item, "kpis": resumo_abastecimentos_frota(filtros)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.delete("/api/frota/abastecimentos/{aid}")
def api_frota_delete_abastecimento(aid: str):
    try:
        excluir_abastecimento_frota(aid)
        return {"ok": True, "kpis": resumo_abastecimentos_frota()}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.get("/api/frota/viaturas-opcoes")
def api_frota_viaturas_opcoes():
    return {"ok": True, "itens": opcoes_viaturas_frota_abastecimento()}


@app.get("/api/frota/fornecedores-opcoes")
def api_frota_fornecedores_opcoes():
    return {"ok": True, "itens": opcoes_fornecedores_frota()}


@app.get("/api/frota/ultimo-hodometro/{viatura_id}")
def api_frota_ultimo_hodometro(viatura_id: str, hodometro: str = "", excluir_id: str = ""):
    ultimo = ultimo_hodometro_viatura_frota(viatura_id, excluir_manutencao_id=excluir_id or None)
    alerta = ""
    hod = _int_controle(hodometro) if str(hodometro or "").strip() else None
    if ultimo and hod is not None and ultimo.get("hodometro") is not None and hod < ultimo["hodometro"]:
        alerta = (
            f"Hodômetro informado ({hod}) é menor que o último registrado "
            f"({ultimo['hodometro']} km)."
        )
    return {
        "ok": True,
        "ultimo_hodometro": ultimo.get("hodometro") if ultimo else None,
        "origem": ultimo.get("origem") if ultimo else None,
        "alerta": alerta,
    }


@app.get("/api/frota/manutencoes")
def api_frota_lista_manutencoes(request: Request):
    filtros = _parse_filtros_manutencoes_request(request)
    return {
        "ok": True,
        "itens": listar_manutencoes_frota(filtros),
        "kpis": resumo_manutencoes_frota(filtros),
    }


@app.get("/api/frota/manutencoes/{mid}")
def api_frota_get_manutencao(mid: str):
    item = manutencao_frota_por_id(mid)
    if not item:
        return JSONResponse({"ok": False, "erro": "Manutenção não encontrada"}, status_code=404)
    return {"ok": True, "item": item}


@app.post("/api/frota/manutencoes")
async def api_frota_post_manutencao(request: Request):
    try:
        payload = await request.json()
        item = salvar_manutencao_frota(payload)
        filtros = _parse_filtros_manutencoes_request(request)
        return {"ok": True, "item": item, "kpis": resumo_manutencoes_frota(filtros)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.put("/api/frota/manutencoes/{mid}")
async def api_frota_put_manutencao(mid: str, request: Request):
    try:
        payload = await request.json()
        payload["id"] = mid
        item = salvar_manutencao_frota(payload)
        filtros = _parse_filtros_manutencoes_request(request)
        return {"ok": True, "item": item, "kpis": resumo_manutencoes_frota(filtros)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.delete("/api/frota/manutencoes/{mid}")
def api_frota_delete_manutencao(mid: str):
    try:
        excluir_manutencao_frota(mid)
        return {"ok": True, "kpis": resumo_manutencoes_frota()}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


FROTA_MULTAS_NATUREZAS = ("Leve", "Média", "Grave", "Gravíssima")
FROTA_MULTAS_STATUS = ("pendente", "pago", "vencido", "cancelado")
EXTRATOS_PROFISSIONAIS_JSON = Path("data/extratos_profissionais.json")


def _add_meses_data(d, meses):
    import calendar
    from datetime import date as date_cls

    if not d:
        return None
    if isinstance(d, str):
        d = _parse_data_controle(d)
    if not d:
        return None
    month = d.month - 1 + int(meses)
    year = d.year + month // 12
    month = month % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return date_cls(year, month, day)


def _calcular_status_multa_frota(valor, valor_pago, vencimento, status_manual=None):
    st = (_txt_cad(status_manual) or "").lower()
    if st in ("cancelado", "cancelada"):
        return "cancelado"
    valor = float(valor or 0)
    valor_pago = float(valor_pago or 0)
    if valor > 0 and valor_pago >= valor:
        return "pago"
    from datetime import date as date_cls

    venc = vencimento
    if isinstance(venc, str):
        venc = _parse_data_controle(venc)
    if venc and venc < date_cls.today():
        return "vencido"
    return "pendente"


def normalizar_multa_frota(row, recalcular_status=True):
    if not row:
        return None
    item = dict(row)
    item["id"] = str(item.get("id"))
    for fk in ("viatura_id", "condutor_id", "condutor_responsavel_id", "contas_pagar_id"):
        if item.get(fk):
            item[fk] = str(item[fk])
    item["valor"] = float(item.get("valor") or 0)
    item["valor_pago"] = float(item.get("valor_pago") or 0)
    item["valor_aberto"] = round(max(item["valor"] - item["valor_pago"], 0), 2)
    item["parcelas"] = int(item.get("parcelas") or 1)
    item["gerar_contas_pagar"] = bool(item.get("gerar_contas_pagar") or item.get("contas_pagar"))
    item["gerar_extrato_profissional"] = bool(
        item.get("gerar_extrato_profissional") or item.get("extrato_profissional")
    )
    cp_ids = item.get("contas_pagar_ids")
    if isinstance(cp_ids, str):
        try:
            cp_ids = json.loads(cp_ids)
        except Exception:
            cp_ids = []
    if not isinstance(cp_ids, list):
        cp_ids = []
    item["contas_pagar_ids"] = [str(x) for x in cp_ids if x]
    st = (item.get("status") or item.get("situacao") or "pendente").lower()
    if st == "ativo":
        st = "pendente"
    if recalcular_status and st != "cancelado":
        st = _calcular_status_multa_frota(
            item["valor"], item["valor_pago"], item.get("vencimento"), st
        )
    item["status"] = st
    item["situacao"] = st
    dh = item.get("data_hora_infracao")
    if dh and hasattr(dh, "strftime"):
        item["data_hora_infracao_fmt"] = dh.strftime("%d/%m/%Y %H:%M")
        item["data_hora_infracao_iso"] = dh.strftime("%Y-%m-%dT%H:%M")
    else:
        item["data_hora_infracao_fmt"] = _formatar_valor_controle("data_hora", dh)
        item["data_hora_infracao_iso"] = str(dh)[:16] if dh else ""
    di = item.get("data_infracao")
    if di and hasattr(di, "strftime"):
        item["data_infracao_fmt"] = di.strftime("%d/%m/%Y")
        item["data_infracao_iso"] = di.strftime("%Y-%m-%d")
    else:
        item["data_infracao_fmt"] = _formatar_valor_controle("data", di)
        item["data_infracao_iso"] = str(di)[:10] if di else ""
    for fld, sfx in (
        ("data_limite_indicacao", "data_limite_indicacao"),
        ("vencimento", "data_vencimento"),
    ):
        val = item.get(fld)
        if val and hasattr(val, "strftime"):
            item[f"{sfx}_fmt"] = val.strftime("%d/%m/%Y")
            item[f"{sfx}_iso"] = val.strftime("%Y-%m-%d")
        else:
            item[f"{sfx}_fmt"] = _formatar_valor_controle("data", val)
            item[f"{sfx}_iso"] = str(val)[:10] if val else ""
    item["valor_fmt"] = fmt_moeda(item["valor"])
    item["valor_pago_fmt"] = fmt_moeda(item["valor_pago"])
    item["valor_aberto_fmt"] = fmt_moeda(item["valor_aberto"])
    item["viatura_nome_exibicao"] = item.get("viatura") or ""
    item["condutor_nome_exibicao"] = item.get("condutor") or ""
    item["condutor_responsavel_nome_exibicao"] = item.get("condutor_responsavel") or ""
    item["created_at"] = dt_str(item.get("created_at"))
    item["updated_at"] = dt_str(item.get("updated_at"))
    return item


def _where_multas_frota(filtros=None):
    filtros = filtros or {}
    st_filtro = (filtros.get("status") or "").lower()
    if st_filtro == "cancelado":
        parts = ["coalesce(m.status, m.situacao, 'ativo') in ('cancelado', 'cancelada')"]
    else:
        parts = ["coalesce(m.status, m.situacao, 'ativo') not in ('cancelado', 'cancelada')"]
    params = []
    if filtros.get("data_ini"):
        parts.append(
            "coalesce(m.data_hora_infracao, m.data_infracao::timestamp)::date >= %s"
        )
        params.append(filtros["data_ini"])
    if filtros.get("data_fim"):
        parts.append(
            "coalesce(m.data_hora_infracao, m.data_infracao::timestamp)::date <= %s"
        )
        params.append(filtros["data_fim"])
    if filtros.get("viatura_id"):
        parts.append("m.viatura_id = %s")
        params.append(str(filtros["viatura_id"]))
    if filtros.get("condutor_id"):
        parts.append("(m.condutor_id = %s or m.condutor_responsavel_id = %s)")
        params.extend([str(filtros["condutor_id"]), str(filtros["condutor_id"])])
    if filtros.get("municipio"):
        parts.append("coalesce(m.municipio,'') ilike %s")
        params.append(f"%{filtros['municipio']}%")
    if filtros.get("status"):
        st = filtros["status"].lower()
        if st == "pago":
            parts.append("coalesce(m.valor_pago,0) >= coalesce(m.valor,0) and coalesce(m.valor,0) > 0")
        elif st == "vencido":
            parts.append(
                "coalesce(m.valor_pago,0) < coalesce(m.valor,0) and m.vencimento is not null and m.vencimento < current_date"
            )
        elif st == "pendente":
            parts.append(
                "coalesce(m.valor_pago,0) < coalesce(m.valor,0) and (m.vencimento is null or m.vencimento >= current_date)"
            )
    if filtros.get("busca"):
        parts.append(
            """
            (
              coalesce(m.viatura,'') ilike %s
              or coalesce(m.placa,'') ilike %s
              or coalesce(m.auto_infracao,'') ilike %s
              or coalesce(m.municipio,'') ilike %s
              or coalesce(m.condutor,'') ilike %s
              or coalesce(m.condutor_responsavel,'') ilike %s
              or coalesce(m.descricao_multa,'') ilike %s
              or coalesce(m.natureza,'') ilike %s
            )
            """
        )
        b = f"%{filtros['busca']}%"
        params.extend([b] * 8)
    where = " where " + " and ".join(parts) if parts else ""
    return where, params


def resumo_multas_frota(filtros=None):
    from datetime import date as date_cls

    where, params = _where_multas_frota(filtros)
    row = one(
        f"""
        select
          count(*)::int as total,
          coalesce(sum(valor), 0)::float as valor_total,
          coalesce(sum(valor_pago), 0)::float as valor_pago,
          coalesce(sum(greatest(valor - coalesce(valor_pago,0), 0)), 0)::float as valor_aberto,
          count(*) filter (
            where coalesce(valor_pago,0) < coalesce(valor,0)
              and vencimento is not null
              and vencimento < current_date
          )::int as vencidas,
          count(*) filter (
            where date_trunc('month', coalesce(data_hora_infracao, data_infracao::timestamp))
              = date_trunc('month', current_timestamp)
          )::int as mes
        from controle_multas m
        {where}
        """,
        tuple(params) if params else None,
    ) or {}
    return {
        "total": int(row.get("total") or 0),
        "valor_total": round(float(row.get("valor_total") or 0), 2),
        "valor_pago": round(float(row.get("valor_pago") or 0), 2),
        "valor_aberto": round(float(row.get("valor_aberto") or 0), 2),
        "vencidas": int(row.get("vencidas") or 0),
        "mes": int(row.get("mes") or 0),
    }


def listar_multas_frota(filtros=None):
    where, params = _where_multas_frota(filtros)
    rows = q(
        f"""
        select m.*
          from controle_multas m
        {where}
         order by coalesce(m.data_hora_infracao, m.data_infracao::timestamp) desc nulls last,
                  m.created_at desc nulls last
         limit 1000
        """,
        tuple(params) if params else None,
        fetch=True,
    )
    return [normalizar_multa_frota(r) for r in rows]


def multa_frota_por_id(mid):
    row = one("select * from controle_multas where id=%s", (str(mid),))
    return normalizar_multa_frota(row) if row else None


def opcoes_municipios_multas_frota():
    rows = q(
        """
        select distinct trim(municipio) as municipio
          from controle_multas
         where coalesce(trim(municipio),'') <> ''
         order by municipio asc
         limit 500
        """,
        fetch=True,
    )
    return [r["municipio"] for r in rows if r.get("municipio")]


def _validar_payload_multa_frota(p, mid=None):
    p = p or {}
    viatura_id = (p.get("viatura_id") or "").strip()
    if not viatura_id:
        raise ValueError("Viatura é obrigatória")
    viatura = viatura_cadastro_por_id(viatura_id)
    if not viatura:
        raise ValueError("Viatura não encontrada")
    auto = (_txt_controle(p.get("auto_infracao"), 80) or "").strip()
    if not auto:
        raise ValueError("Auto de Infração é obrigatório")
    data_hora = _parse_datetime_controle(p.get("data_hora_infracao"))
    if not data_hora:
        raise ValueError("Data e Hora da Infração são obrigatórias")
    municipio = (_txt_controle(p.get("municipio"), 120) or "").strip()
    if not municipio:
        raise ValueError("Município é obrigatório")
    endereco = (_txt_controle(p.get("endereco")) or "").strip()
    if not endereco:
        raise ValueError("Endereço é obrigatório")
    descricao = (_txt_controle(p.get("descricao_multa")) or "").strip()
    if not descricao:
        raise ValueError("Descrição da Multa é obrigatória")
    natureza = (_txt_controle(p.get("natureza"), 80) or "").strip()
    if not natureza:
        raise ValueError("Natureza é obrigatória")
    condutor_id = (p.get("condutor_id") or "").strip()
    if not condutor_id:
        raise ValueError("Condutor é obrigatório")
    cond = profissional_cadastro_por_id(condutor_id)
    if not cond:
        raise ValueError("Condutor não encontrado")
    resp_id = (p.get("condutor_responsavel_id") or "").strip()
    if not resp_id:
        raise ValueError("Condutor Responsável é obrigatório")
    resp = profissional_cadastro_por_id(resp_id)
    if not resp:
        raise ValueError("Condutor Responsável não encontrado")
    parcelas = _int_controle(p.get("parcelas"), 1) or 1
    if parcelas < 1:
        raise ValueError("Parcelas deve ser >= 1")
    vencimento = _parse_data_controle(p.get("data_vencimento") or p.get("vencimento"))
    if not vencimento:
        raise ValueError("Data de Vencimento é obrigatória")
    valor = _num_controle(p.get("valor"))
    if valor <= 0:
        raise ValueError("Valor deve ser maior que zero")
    valor_pago = _num_controle(p.get("valor_pago"))
    if valor_pago < 0:
        raise ValueError("Valor pago não pode ser negativo")
    valor_aberto = round(max(valor - valor_pago, 0), 2)
    status_manual = (_txt_controle(p.get("status"), 40) or "").lower()
    if mid:
        atual = multa_frota_por_id(mid)
        if atual and atual.get("status") == "cancelado":
            status_manual = "cancelado"
    status = _calcular_status_multa_frota(valor, valor_pago, vencimento, status_manual)
    viatura_label = _label_viatura_abastecimento(viatura)
    return {
        "viatura_id": viatura_id,
        "viatura": viatura_label,
        "placa": viatura.get("placa") or "",
        "auto_infracao": auto,
        "data_hora_infracao": data_hora,
        "data_infracao": data_hora.date(),
        "data_limite_indicacao": _parse_data_controle(p.get("data_limite_indicacao")),
        "municipio": municipio,
        "endereco": endereco,
        "descricao_multa": descricao,
        "natureza": natureza,
        "condutor_id": condutor_id,
        "condutor": _label_profissional_frota(cond),
        "condutor_responsavel_id": resp_id,
        "condutor_responsavel": _label_profissional_frota(resp),
        "parcelas": parcelas,
        "vencimento": vencimento,
        "valor": valor,
        "valor_pago": valor_pago,
        "valor_aberto": valor_aberto,
        "observacoes": _txt_controle(p.get("observacao") or p.get("observacoes")),
        "gerar_contas_pagar": _bool_controle(p.get("gerar_contas_pagar")),
        "gerar_extrato_profissional": _bool_controle(p.get("gerar_extrato_profissional")),
        "status": status,
    }


def _criar_contas_pagar_multa_frota(multa):
    row_cat = one(
        """
        select id, descricao from financeiro_categorias
         where natureza='Despesa'
           and (descricao ilike '%%Multa%%' or descricao ilike '%%Infra%%')
         order by case when descricao ilike '%%Multa%%' then 0 else 1 end, ordem asc
         limit 1
        """
    )
    categoria_id = str(row_cat["id"]) if row_cat else None
    categoria = (row_cat.get("descricao") if row_cat else None) or "Multas"
    viatura_label = (multa.get("viatura") or multa.get("placa") or "").strip()
    auto = (multa.get("auto_infracao") or "").strip()
    condutor = (multa.get("condutor") or "").strip()
    hist_base = f"Multa - {viatura_label} - Auto {auto} - Condutor {condutor}"
    valor_base = float(multa.get("valor_aberto") or multa.get("valor") or 0)
    if valor_base <= 0:
        raise ValueError("Valor em aberto inválido para contas a pagar")
    parcelas = int(multa.get("parcelas") or 1)
    venc0 = multa.get("vencimento")
    if isinstance(venc0, str):
        venc0 = _parse_data_controle(venc0)
    valor_parcela = round(valor_base / parcelas, 2)
    ids = []
    for i in range(parcelas):
        venc = _add_meses_data(venc0, i) if venc0 else None
        valor_linha = valor_parcela
        if i == parcelas - 1:
            valor_linha = round(valor_base - valor_parcela * (parcelas - 1), 2)
        desc = hist_base if parcelas == 1 else f"{hist_base} — Parcela {i + 1}/{parcelas}"
        row = one(
            """
            insert into financeiro_contas_pagar (
              fornecedor, descricao, categoria, categoria_id, competencia, emissao, vencimento,
              valor, parcelas, status, observacoes, updated_at
            ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,'em_aberto',%s,now())
            returning id
            """,
            (
                "Multa de Trânsito",
                desc,
                categoria,
                categoria_id,
                venc,
                venc,
                venc,
                valor_linha,
                parcelas,
                f"Gerado automaticamente da multa {multa.get('id') or ''}".strip(),
            ),
        )
        ids.append(str(row["id"]))
    return ids


def _registrar_extrato_profissional_multa(multa):
    EXTRATOS_PROFISSIONAIS_JSON.parent.mkdir(parents=True, exist_ok=True)
    itens = []
    if EXTRATOS_PROFISSIONAIS_JSON.exists():
        try:
            itens = json.loads(EXTRATOS_PROFISSIONAIS_JSON.read_text(encoding="utf-8"))
        except Exception:
            itens = []
    if not isinstance(itens, list):
        itens = []
    ext_id = str(uuid.uuid4())
    venc = multa.get("vencimento")
    data_ref = multa.get("data_infracao") or multa.get("data_hora_infracao")
    if hasattr(data_ref, "date"):
        data_ref = data_ref.date()
    if hasattr(venc, "isoformat"):
        data_ext = venc.isoformat()
    elif data_ref and hasattr(data_ref, "isoformat"):
        data_ext = data_ref.isoformat()
    else:
        data_ext = str(venc or data_ref or "")[:10]
    registro = {
        "id": ext_id,
        "profissional_id": multa.get("condutor_responsavel_id"),
        "profissional_nome": multa.get("condutor_responsavel"),
        "tipo": "desconto",
        "categoria": "Multa",
        "valor": float(multa.get("valor_aberto") or multa.get("valor") or 0),
        "descricao": (
            f"Multa - Auto {multa.get('auto_infracao')} - Viatura {multa.get('placa')} - "
            f"{multa.get('descricao_multa') or ''}"
        ).strip(),
        "data": data_ext,
        "multa_id": multa.get("id"),
        "criado_em": datetime.now().isoformat(),
    }
    itens.append(registro)
    EXTRATOS_PROFISSIONAIS_JSON.write_text(
        json.dumps(itens, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return ext_id


def salvar_multa_frota(payload):
    p = payload or {}
    mid = (p.get("id") or "").strip()
    dados = _validar_payload_multa_frota(p, mid=mid or None)
    existente = multa_frota_por_id(mid) if mid else None
    cp_ids_existentes = (existente or {}).get("contas_pagar_ids") or []
    ext_id_existente = (existente or {}).get("extrato_profissional_id")
    params = (
        dados["viatura"],
        dados["placa"],
        dados["viatura_id"],
        dados["auto_infracao"],
        dados["data_infracao"],
        dados["data_hora_infracao"],
        dados["data_limite_indicacao"],
        dados["municipio"],
        dados["endereco"],
        dados["descricao_multa"],
        dados["natureza"],
        dados["condutor"],
        dados["condutor_id"],
        dados["condutor_responsavel"],
        dados["condutor_responsavel_id"],
        dados["parcelas"],
        dados["vencimento"],
        dados["valor"],
        dados["valor_pago"],
        dados["valor_aberto"],
        dados["observacoes"],
        dados["status"],
        dados["status"],
        dados["gerar_contas_pagar"],
        dados["gerar_extrato_profissional"],
        dados["gerar_contas_pagar"],
        dados["gerar_extrato_profissional"],
    )
    if mid:
        q(
            """
            update controle_multas set
              viatura=%s, placa=%s, viatura_id=%s, auto_infracao=%s,
              data_infracao=%s, data_hora_infracao=%s, data_limite_indicacao=%s,
              municipio=%s, endereco=%s, descricao_multa=%s, natureza=%s,
              condutor=%s, condutor_id=%s, condutor_responsavel=%s, condutor_responsavel_id=%s,
              parcelas=%s, vencimento=%s, valor=%s, valor_pago=%s, valor_aberto=%s,
              observacoes=%s, status=%s, situacao=%s,
              gerar_contas_pagar=%s, gerar_extrato_profissional=%s,
              contas_pagar=%s, extrato_profissional=%s,
              updated_at=now()
            where id=%s
            """,
            params + (mid,),
        )
        item = multa_frota_por_id(mid)
    else:
        row = one(
            """
            insert into controle_multas (
              viatura, placa, viatura_id, auto_infracao,
              data_infracao, data_hora_infracao, data_limite_indicacao,
              municipio, endereco, descricao_multa, natureza,
              condutor, condutor_id, condutor_responsavel, condutor_responsavel_id,
              parcelas, vencimento, valor, valor_pago, valor_aberto,
              observacoes, status, situacao,
              gerar_contas_pagar, gerar_extrato_profissional,
              contas_pagar, extrato_profissional,
              updated_at
            ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,now())
            returning id
            """,
            params,
        )
        item = multa_frota_por_id(row["id"])
    if dados["gerar_contas_pagar"] and not cp_ids_existentes and not item.get("contas_pagar_ids"):
        cp_ids = _criar_contas_pagar_multa_frota(item)
        cp_principal = cp_ids[0] if cp_ids else None
        q(
            """
            update controle_multas set
              contas_pagar_id=%s, contas_pagar_ids=%s, updated_at=now()
            where id=%s
            """,
            (cp_principal, Json(cp_ids), item["id"]),
        )
        item = multa_frota_por_id(item["id"])
    if dados["gerar_extrato_profissional"] and not ext_id_existente and not item.get("extrato_profissional_id"):
        ext_id = _registrar_extrato_profissional_multa(item)
        q(
            "update controle_multas set extrato_profissional_id=%s, updated_at=now() where id=%s",
            (ext_id, item["id"]),
        )
        item = multa_frota_por_id(item["id"])
    return item


def excluir_multa_frota(mid):
    q(
        """
        update controle_multas
           set status='cancelado', situacao='cancelado', updated_at=now()
         where id=%s
        """,
        (str(mid),),
    )


def _parse_filtros_multas_request(request):
    qparams = request.query_params if request else {}
    return {
        "data_ini": _parse_data_controle(qparams.get("data_ini")),
        "data_fim": _parse_data_controle(qparams.get("data_fim")),
        "viatura_id": (qparams.get("viatura_id") or "").strip() or None,
        "condutor_id": (qparams.get("condutor_id") or "").strip() or None,
        "municipio": (qparams.get("municipio") or "").strip() or None,
        "status": (qparams.get("status") or "").strip().lower() or None,
        "busca": (qparams.get("busca") or "").strip() or None,
    }


@app.get("/frota/multas", response_class=HTMLResponse)
def frota_pagina_multas(request: Request):
    return templates.TemplateResponse(
        "frota/multas.html",
        {
            "request": request,
            "nav_ativo": "frota",
            "nav_som": False,
            "kpis": resumo_multas_frota(),
            "naturezas": FROTA_MULTAS_NATUREZAS,
            "status_opcoes": FROTA_MULTAS_STATUS,
        },
    )


@app.get("/api/frota/multas")
def api_frota_lista_multas(request: Request):
    filtros = _parse_filtros_multas_request(request)
    return {
        "ok": True,
        "itens": listar_multas_frota(filtros),
        "kpis": resumo_multas_frota(filtros),
    }


@app.get("/api/frota/multas/opcoes/municipios")
def api_frota_multas_opcoes_municipios():
    return {"ok": True, "itens": opcoes_municipios_multas_frota()}


@app.get("/api/frota/multas/{mid}")
def api_frota_get_multa(mid: str):
    item = multa_frota_por_id(mid)
    if not item:
        return JSONResponse({"ok": False, "erro": "Multa não encontrada"}, status_code=404)
    return {"ok": True, "item": item}


@app.post("/api/frota/multas")
async def api_frota_post_multa(request: Request):
    try:
        payload = await request.json()
        item = salvar_multa_frota(payload)
        filtros = _parse_filtros_multas_request(request)
        return {"ok": True, "item": item, "kpis": resumo_multas_frota(filtros)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.put("/api/frota/multas/{mid}")
async def api_frota_put_multa(mid: str, request: Request):
    try:
        payload = await request.json()
        payload["id"] = mid
        item = salvar_multa_frota(payload)
        filtros = _parse_filtros_multas_request(request)
        return {"ok": True, "item": item, "kpis": resumo_multas_frota(filtros)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.delete("/api/frota/multas/{mid}")
def api_frota_delete_multa(mid: str):
    try:
        excluir_multa_frota(mid)
        return {"ok": True, "kpis": resumo_multas_frota()}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.get("/api/frota/profissionais-opcoes")
def api_frota_profissionais_opcoes():
    return {"ok": True, "itens": opcoes_profissionais_frota_abastecimento()}


FROTA_PATIO_STATUS = ("no_patio", "saiu", "cancelado")
FROTA_PATIO_STATUS_LABEL = {
    "no_patio": "No pátio",
    "saiu": "Saiu",
    "cancelado": "Cancelado",
}


def _label_status_patio(st):
    s = (_txt_cad(st) or "no_patio").lower()
    return FROTA_PATIO_STATUS_LABEL.get(s, s.replace("_", " ").title())


def _combinar_data_hora_patio(data_val, hora_val=None, datetime_val=None):
    if datetime_val:
        dt = _parse_datetime_controle(datetime_val)
        if dt:
            return dt
    d = _parse_data_controle(data_val)
    if not d:
        return None
    if hora_val:
        if hasattr(hora_val, "hour"):
            t = hora_val
        else:
            txt = str(hora_val).strip()
            if len(txt) >= 5 and ":" in txt:
                parts = txt.split(":")
                try:
                    from datetime import time as time_cls

                    t = time_cls(int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 0)
                except Exception:
                    t = None
            else:
                t = None
        if t:
            return datetime.combine(d, t)
    return datetime.combine(d, datetime.min.time())


def _calcular_tempo_patio_minutos(chegada, saida=None, status="no_patio"):
    if not chegada:
        return 0
    if isinstance(chegada, str):
        chegada = _parse_datetime_controle(chegada)
    if not chegada:
        return 0
    st = (_txt_cad(status) or "no_patio").lower()
    if st == "no_patio" or not saida:
        fim = datetime.now()
    else:
        fim = saida if not isinstance(saida, str) else _parse_datetime_controle(saida)
    if not fim:
        fim = datetime.now()
    delta = fim - chegada
    return max(int(delta.total_seconds() // 60), 0)


def _formatar_tempo_patio(minutos):
    if minutos is None:
        return "-"
    mins = int(minutos or 0)
    if mins <= 0:
        return "0min"
    if mins < 60:
        return f"{mins}min"
    horas = mins // 60
    resto = mins % 60
    if horas < 24:
        return f"{horas}h {resto}min" if resto else f"{horas}h"
    dias = horas // 24
    horas_r = horas % 24
    partes = [f"{dias} dia{'s' if dias > 1 else ''}"]
    if horas_r:
        partes.append(f"{horas_r}h")
    elif resto:
        partes.append(f"{resto}min")
    return " ".join(partes)


def normalizar_patio_frota(row):
    if not row:
        return None
    item = dict(row)
    item["id"] = str(item.get("id"))
    item["placa"] = normalizar_placa_viatura(item.get("placa"))
    st = (_txt_cad(item.get("status")) or "no_patio").lower()
    if st == "ativo":
        st = "no_patio"
    item["status"] = st
    item["status_label"] = _label_status_patio(st)
    chegada = item.get("data_hora_chegada")
    if not chegada and item.get("data_entrada"):
        chegada = _combinar_data_hora_patio(item.get("data_entrada"), item.get("hora_chegada"))
    saida = item.get("data_hora_saida")
    if not saida and item.get("data_saida"):
        saida = _combinar_data_hora_patio(item.get("data_saida"), item.get("hora_saida"))
    item["data_hora_chegada"] = chegada
    item["data_hora_saida"] = saida
    if chegada and hasattr(chegada, "date"):
        item["data_entrada"] = chegada.date()
        item["data_entrada_fmt"] = chegada.strftime("%d/%m/%Y")
        item["data_entrada_iso"] = chegada.strftime("%Y-%m-%d")
        item["hora_chegada_fmt"] = chegada.strftime("%H:%M")
    else:
        item["data_entrada_fmt"] = _formatar_valor_controle("data", item.get("data_entrada"))
        item["data_entrada_iso"] = str(item.get("data_entrada") or "")[:10]
        item["hora_chegada_fmt"] = str(item.get("hora_chegada") or "")[:5]
    if saida and hasattr(saida, "strftime"):
        item["data_saida"] = saida.date()
        item["data_saida_fmt"] = saida.strftime("%d/%m/%Y")
        item["data_saida_iso"] = saida.strftime("%Y-%m-%d")
        item["hora_saida_fmt"] = saida.strftime("%H:%M")
    else:
        item["data_saida_fmt"] = _formatar_valor_controle("data", item.get("data_saida")) if item.get("data_saida") else "-"
        item["data_saida_iso"] = str(item.get("data_saida") or "")[:10]
        item["hora_saida_fmt"] = str(item.get("hora_saida") or "")[:5] if item.get("hora_saida") else "-"
    mins = _calcular_tempo_patio_minutos(chegada, saida, st)
    item["tempo_no_patio_minutos"] = mins
    item["tempo_no_patio_fmt"] = _formatar_tempo_patio(mins)
    obs = item.get("observacao") or item.get("observacoes") or ""
    item["observacao"] = obs
    item["observacao_resumo"] = (obs[:80] + "…") if len(obs) > 80 else obs
    item["created_at"] = dt_str(item.get("created_at"))
    item["updated_at"] = dt_str(item.get("updated_at"))
    return item


def _where_patio_frota(filtros=None):
    filtros = filtros or {}
    parts = []
    params = []
    st = (filtros.get("status") or "").lower()
    if st == "cancelado":
        parts.append("coalesce(p.status,'no_patio') = 'cancelado'")
    elif st:
        parts.append("coalesce(p.status,'no_patio') = %s")
        params.append(st)
    else:
        parts.append("coalesce(p.status,'no_patio') <> 'cancelado'")
    if filtros.get("data_ini"):
        parts.append("coalesce(p.data_entrada, p.data_hora_chegada::date) >= %s")
        params.append(filtros["data_ini"])
    if filtros.get("data_fim"):
        parts.append("coalesce(p.data_entrada, p.data_hora_chegada::date) <= %s")
        params.append(filtros["data_fim"])
    if filtros.get("seguradora"):
        parts.append("coalesce(p.seguradora,'') ilike %s")
        params.append(f"%{filtros['seguradora']}%")
    if filtros.get("placa"):
        placa_n = normalizar_placa_viatura(filtros["placa"])
        parts.append(
            "upper(replace(replace(replace(coalesce(p.placa,''), ' ', ''), '-', ''), '.', '')) = %s"
        )
        params.append(placa_n)
    if filtros.get("motorista"):
        parts.append("coalesce(p.motorista,'') ilike %s")
        params.append(f"%{filtros['motorista']}%")
    if filtros.get("responsavel"):
        parts.append(
            "(coalesce(p.responsavel_entrada,'') ilike %s or coalesce(p.responsavel_saida,'') ilike %s)"
        )
        b = f"%{filtros['responsavel']}%"
        params.extend([b, b])
    if filtros.get("busca"):
        parts.append(
            """
            (
              coalesce(p.placa,'') ilike %s
              or coalesce(p.motorista,'') ilike %s
              or coalesce(p.modelo,'') ilike %s
              or coalesce(p.cor,'') ilike %s
              or coalesce(p.seguradora,'') ilike %s
              or coalesce(p.responsavel_entrada,'') ilike %s
              or coalesce(p.observacao,'') ilike %s
            )
            """
        )
        b = f"%{filtros['busca']}%"
        params.extend([b, b, b, b, b, b, b])
    where = " where " + " and ".join(parts) if parts else ""
    return where, params


def resumo_patio_frota(filtros=None):
    where, params = _where_patio_frota(filtros)
    row = one(
        f"""
        select
          count(*) filter (where coalesce(p.status,'no_patio') = 'no_patio')::int as no_patio,
          count(*) filter (
            where coalesce(p.data_entrada, p.data_hora_chegada::date) = current_date
          )::int as entradas_hoje,
          count(*) filter (
            where coalesce(p.status,'no_patio') = 'saiu'
              and coalesce(p.data_saida, p.data_hora_saida::date) = current_date
          )::int as saidas_hoje,
          count(*)::int as total_periodo,
          coalesce(avg(
            case when coalesce(p.status,'no_patio') = 'saiu' and p.tempo_no_patio_minutos > 0
                 then p.tempo_no_patio_minutos end
          ), 0)::float as tempo_medio
        from controle_patio p
        {where}
        """,
        tuple(params) if params else None,
    ) or {}
    no_patio = int(row.get("no_patio") or 0)
    tempo_medio = int(round(float(row.get("tempo_medio") or 0)))
    return {
        "no_patio": no_patio,
        "entradas_hoje": int(row.get("entradas_hoje") or 0),
        "saidas_hoje": int(row.get("saidas_hoje") or 0),
        "pendentes_saida": no_patio,
        "tempo_medio_minutos": tempo_medio,
        "tempo_medio_fmt": _formatar_tempo_patio(tempo_medio),
        "total_periodo": int(row.get("total_periodo") or 0),
    }


def listar_patio_frota(filtros=None):
    where, params = _where_patio_frota(filtros)
    rows = q(
        f"""
        select p.*
          from controle_patio p
        {where}
         order by coalesce(p.data_hora_chegada, p.data_entrada::timestamp) desc nulls last,
                  p.created_at desc nulls last
         limit 1000
        """,
        tuple(params) if params else None,
        fetch=True,
    )
    return [normalizar_patio_frota(r) for r in rows]


def patio_frota_por_id(pid):
    row = one("select * from controle_patio where id=%s", (str(pid),))
    return normalizar_patio_frota(row) if row else None


# --- Integração experimental Central ↔ Pátio (Fase 1 — somente consulta) ---
CENTRAL_PATIO_BADGE_KEYS = ("no_patio", "pendente_saida", "saiu")


def _status_exibicao_central_patio(patio_status, servico_status):
    ps = (_txt_cad(patio_status) or "no_patio").lower()
    if ps == "saiu":
        return "SAIU DO PÁTIO", "saiu"
    if ps == "no_patio":
        ss = (_txt_cad(servico_status) or "").lower()
        if ss == "finalizado":
            return "PENDENTE DE SAÍDA", "pendente_saida"
        return "NO PÁTIO", "no_patio"
    return None, None


def patio_integracao_central_por_servico(sid):
    """Consulta read-only de registro de pátio vinculado a um serviço da Central."""
    s = servico_by_id(sid)
    if not s:
        return None
    placa = normalizar_placa_viatura(
        s.get("placa_veiculo_removido") or s.get("placa_removida") or ""
    )
    row = one(
        """
        select *
          from controle_patio
         where servico_id = %s
           and coalesce(status,'no_patio') <> 'cancelado'
         order by coalesce(data_hora_chegada, data_entrada::timestamp) desc nulls last
         limit 1
        """,
        (str(sid),),
    )
    vinculo_por = "servico_id" if row else None
    if not row and placa:
        row = one(
            """
            select *
              from controle_patio
             where coalesce(status,'no_patio') <> 'cancelado'
               and upper(replace(replace(replace(coalesce(placa,''), ' ', ''), '-', ''), '.', '')) = %s
             order by
               case when coalesce(status,'no_patio') = 'no_patio' then 0 else 1 end,
               coalesce(data_hora_chegada, data_entrada::timestamp) desc nulls last
             limit 1
            """,
            (placa,),
        )
        if row:
            vinculo_por = "placa"
    if not row:
        return None
    item = normalizar_patio_frota(row)
    st_db = (item.get("status") or "no_patio").lower()
    if st_db == "cancelado":
        return None
    label, badge_key = _status_exibicao_central_patio(st_db, s.get("status"))
    if not label:
        return None
    chegada = item.get("data_hora_chegada")
    saida = item.get("data_hora_saida")
    if chegada:
        entrada_fmt = formatar_data_hora_central(chegada)
    else:
        entrada_fmt = f"{item.get('data_entrada_fmt') or '-'} {item.get('hora_chegada_fmt') or ''}".strip()
    saida_fmt = formatar_data_hora_central(saida) if saida else None
    return {
        "id": item.get("id"),
        "status": badge_key,
        "status_label": label,
        "data_hora_entrada": entrada_fmt,
        "tempo_no_patio_fmt": item.get("tempo_no_patio_fmt")
        or _formatar_tempo_patio(item.get("tempo_no_patio_minutos") or 0),
        "responsavel_entrada": item.get("responsavel_entrada") or "-",
        "data_hora_saida": saida_fmt,
        "responsavel_saida": (item.get("responsavel_saida") or "").strip() or None,
        "url_registro": f"/frota/controle-patio?registro={item.get('id')}",
        "vinculo_por": vinculo_por,
        "placa": item.get("placa") or placa or "",
    }


def placa_no_patio_frota(placa, excluir_id=None):
    placa_n = normalizar_placa_viatura(placa)
    if not placa_n:
        return None
    params = [placa_n]
    sql = """
        select id, placa, motorista, data_hora_chegada, status
          from controle_patio
         where coalesce(status,'no_patio') = 'no_patio'
           and upper(replace(replace(replace(coalesce(placa,''), ' ', ''), '-', ''), '.', '')) = %s
    """
    if excluir_id:
        sql += " and id <> %s"
        params.append(str(excluir_id))
    sql += " order by data_hora_chegada desc nulls last limit 1"
    row = one(sql, tuple(params))
    return normalizar_patio_frota(row) if row else None


def opcoes_seguradoras_frota():
    nomes = set()
    rows_cli = q(
        """
        select coalesce(nullif(trim(nome_fantasia),''), nullif(trim(razao_social),''), nullif(trim(nome),'')) as nome
          from cadastro_contatos
         where coalesce(cliente, false) = true
           and coalesce(status,'ativo') = 'ativo'
        """,
        fetch=True,
    ) or []
    for r in rows_cli:
        n = (r.get("nome") or "").strip()
        if n:
            nomes.add(n)
    rows_srv = q(
        """
        select distinct trim(seguradora) as nome
          from servicos
         where coalesce(trim(seguradora),'') <> ''
         limit 500
        """,
        fetch=True,
    ) or []
    for r in rows_srv:
        n = (r.get("nome") or "").strip()
        if n:
            nomes.add(n)
    rows_pat = q(
        """
        select distinct trim(seguradora) as nome
          from controle_patio
         where coalesce(trim(seguradora),'') <> ''
         limit 500
        """,
        fetch=True,
    ) or []
    for r in rows_pat:
        n = (r.get("nome") or "").strip()
        if n:
            nomes.add(n)
    return sorted(nomes, key=lambda x: x.lower())


def _validar_payload_patio_frota(p, pid=None):
    p = p or {}
    placa = normalizar_placa_viatura(p.get("placa"))
    if not placa:
        raise ValueError("Placa é obrigatória")
    data_entrada = p.get("data_entrada")
    hora_chegada = p.get("hora_chegada")
    data_hora_chegada = _combinar_data_hora_patio(
        data_entrada, hora_chegada, p.get("data_hora_chegada")
    )
    if not data_hora_chegada:
        raise ValueError("Data e hora de chegada são obrigatórias")
    responsavel_entrada = (_txt_controle(p.get("responsavel_entrada"), 120) or "").strip()
    if not responsavel_entrada:
        raise ValueError("Responsável pela entrada é obrigatório")
    st = (_txt_controle(p.get("status"), 40) or "no_patio").lower()
    if st not in FROTA_PATIO_STATUS:
        st = "no_patio"
    if pid:
        atual = patio_frota_por_id(pid)
        if atual and atual.get("status") == "cancelado":
            st = "cancelado"
    data_hora_saida = None
    if p.get("data_hora_saida") or p.get("data_saida") or p.get("hora_saida"):
        data_hora_saida = _combinar_data_hora_patio(
            p.get("data_saida"), p.get("hora_saida"), p.get("data_hora_saida")
        )
    if st == "saiu" and not data_hora_saida:
        raise ValueError("Hora de saída é obrigatória para status Saiu")
    if data_hora_saida and data_hora_saida < data_hora_chegada:
        raise ValueError("Saída não pode ser anterior à chegada")
    tempo = _calcular_tempo_patio_minutos(data_hora_chegada, data_hora_saida, st)
    return {
        "data_entrada": data_hora_chegada.date(),
        "hora_chegada": data_hora_chegada.time(),
        "data_hora_chegada": data_hora_chegada,
        "motorista": _txt_controle(p.get("motorista"), 200),
        "placa": placa,
        "modelo": _txt_controle(p.get("modelo"), 120),
        "cor": _txt_controle(p.get("cor"), 80),
        "seguradora": _txt_controle(p.get("seguradora"), 200),
        "data_saida": data_hora_saida.date() if data_hora_saida else None,
        "hora_saida": data_hora_saida.time() if data_hora_saida else None,
        "data_hora_saida": data_hora_saida,
        "responsavel_entrada": responsavel_entrada,
        "responsavel_saida": _txt_controle(p.get("responsavel_saida"), 120),
        "observacao": _txt_controle(p.get("observacao") or p.get("observacoes")),
        "observacao_saida": _txt_controle(p.get("observacao_saida")),
        "status": st,
        "tempo_no_patio_minutos": tempo,
    }


def salvar_patio_frota(payload):
    p = payload or {}
    pid = (p.get("id") or "").strip()
    if not pid and not _bool_controle(p.get("confirmar_duplicata")):
        dup = placa_no_patio_frota(p.get("placa"))
        if dup:
            return {
                "requer_confirmacao": True,
                "mensagem": "Esta placa já está no pátio. Deseja continuar mesmo assim?",
                "placa_duplicada": dup,
            }
    dados = _validar_payload_patio_frota(p, pid=pid or None)
    params = (
        dados["data_entrada"],
        dados["hora_chegada"],
        dados["data_hora_chegada"],
        dados["motorista"],
        dados["placa"],
        dados["modelo"],
        dados["cor"],
        dados["seguradora"],
        dados["data_saida"],
        dados["hora_saida"],
        dados["data_hora_saida"],
        dados["responsavel_entrada"],
        dados["responsavel_saida"],
        dados["observacao"],
        dados["observacao_saida"],
        dados["status"],
        dados["tempo_no_patio_minutos"],
    )
    if pid:
        q(
            """
            update controle_patio set
              data_entrada=%s, hora_chegada=%s, data_hora_chegada=%s,
              motorista=%s, placa=%s, modelo=%s, cor=%s, seguradora=%s,
              data_saida=%s, hora_saida=%s, data_hora_saida=%s,
              responsavel_entrada=%s, responsavel_saida=%s,
              observacao=%s, observacao_saida=%s,
              status=%s, tempo_no_patio_minutos=%s, updated_at=now()
            where id=%s
            """,
            params + (pid,),
        )
        return patio_frota_por_id(pid)
    row = one(
        """
        insert into controle_patio (
          data_entrada, hora_chegada, data_hora_chegada,
          motorista, placa, modelo, cor, seguradora,
          data_saida, hora_saida, data_hora_saida,
          responsavel_entrada, responsavel_saida,
          observacao, observacao_saida,
          status, tempo_no_patio_minutos, updated_at
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,now())
        returning id
        """,
        params,
    )
    return patio_frota_por_id(row["id"])


def registrar_saida_patio_frota(pid, payload):
    item = patio_frota_por_id(pid)
    if not item:
        raise ValueError("Registro não encontrado")
    if item.get("status") == "cancelado":
        raise ValueError("Registro cancelado")
    if item.get("status") == "saiu":
        raise ValueError("Saída já registrada")
    p = payload or {}
    data_hora_saida = _combinar_data_hora_patio(
        p.get("data_saida") or datetime.now().date(),
        p.get("hora_saida"),
        p.get("data_hora_saida") or datetime.now().strftime("%Y-%m-%dT%H:%M"),
    )
    if not data_hora_saida:
        raise ValueError("Hora de saída é obrigatória")
    chegada = item.get("data_hora_chegada")
    if chegada and data_hora_saida < chegada:
        raise ValueError("Saída não pode ser anterior à chegada")
    resp = (_txt_controle(p.get("responsavel_saida"), 120) or "").strip()
    if not resp:
        raise ValueError("Responsável pela saída é obrigatório")
    tempo = _calcular_tempo_patio_minutos(chegada, data_hora_saida, "saiu")
    q(
        """
        update controle_patio set
          data_saida=%s, hora_saida=%s, data_hora_saida=%s,
          responsavel_saida=%s, observacao_saida=%s,
          status='saiu', tempo_no_patio_minutos=%s, updated_at=now()
        where id=%s
        """,
        (
            data_hora_saida.date(),
            data_hora_saida.time(),
            data_hora_saida,
            resp,
            _txt_controle(p.get("observacao_saida")),
            tempo,
            str(pid),
        ),
    )
    return patio_frota_por_id(pid)


def excluir_patio_frota(pid):
    q(
        "update controle_patio set status='cancelado', updated_at=now() where id=%s",
        (str(pid),),
    )


def _parse_filtros_patio_request(request):
    qparams = request.query_params if request else {}
    return {
        "data_ini": _parse_data_controle(qparams.get("data_ini")),
        "data_fim": _parse_data_controle(qparams.get("data_fim")),
        "status": (qparams.get("status") or "").strip().lower() or None,
        "seguradora": (qparams.get("seguradora") or "").strip() or None,
        "placa": (qparams.get("placa") or "").strip() or None,
        "motorista": (qparams.get("motorista") or "").strip() or None,
        "responsavel": (qparams.get("responsavel") or "").strip() or None,
        "busca": (qparams.get("busca") or "").strip() or None,
    }


@app.get("/api/frota/controle-patio")
def api_frota_lista_patio(request: Request):
    filtros = _parse_filtros_patio_request(request)
    return {
        "ok": True,
        "itens": listar_patio_frota(filtros),
        "kpis": resumo_patio_frota(filtros),
    }


@app.get("/api/frota/controle-patio/{pid}")
def api_frota_get_patio(pid: str):
    item = patio_frota_por_id(pid)
    if not item:
        return JSONResponse({"ok": False, "erro": "Registro não encontrado"}, status_code=404)
    return {"ok": True, "item": item}


@app.post("/api/frota/controle-patio")
async def api_frota_post_patio(request: Request):
    try:
        payload = await request.json()
        result = salvar_patio_frota(payload)
        if isinstance(result, dict) and result.get("requer_confirmacao"):
            return JSONResponse({"ok": False, **result}, status_code=409)
        filtros = _parse_filtros_patio_request(request)
        return {"ok": True, "item": result, "kpis": resumo_patio_frota(filtros)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.put("/api/frota/controle-patio/{pid}")
async def api_frota_put_patio(pid: str, request: Request):
    try:
        payload = await request.json()
        payload["id"] = pid
        result = salvar_patio_frota(payload)
        if isinstance(result, dict) and result.get("requer_confirmacao"):
            return JSONResponse({"ok": False, **result}, status_code=409)
        filtros = _parse_filtros_patio_request(request)
        return {"ok": True, "item": result, "kpis": resumo_patio_frota(filtros)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.post("/api/frota/controle-patio/{pid}/saida")
async def api_frota_saida_patio(pid: str, request: Request):
    try:
        payload = await request.json()
        item = registrar_saida_patio_frota(pid, payload)
        filtros = _parse_filtros_patio_request(request)
        return {"ok": True, "item": item, "kpis": resumo_patio_frota(filtros)}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.delete("/api/frota/controle-patio/{pid}")
def api_frota_delete_patio(pid: str):
    try:
        excluir_patio_frota(pid)
        return {"ok": True, "kpis": resumo_patio_frota()}
    except Exception as exc:
        return JSONResponse({"ok": False, "erro": str(exc)}, status_code=400)


@app.get("/api/frota/seguradoras-opcoes")
def api_frota_seguradoras_opcoes():
    return {"ok": True, "itens": opcoes_seguradoras_frota()}


@app.get("/api/integracao/central-patio/servico/{sid}")
def api_integracao_central_patio_servico(sid: str):
    """Fase 1 — consulta visual desacoplada; não altera fluxos operacionais."""
    if not servico_by_id(sid):
        return JSONResponse({"ok": False, "erro": "Serviço não encontrado"}, status_code=404)
    registro = patio_integracao_central_por_servico(sid)
    return {
        "ok": True,
        "experimental": True,
        "vinculo": registro is not None,
        "registro": registro,
    }


@app.get("/controle/danos", response_class=HTMLResponse)
def controle_pagina_danos(request: Request):
    return pagina_modulo_controle("danos", request)


@app.get("/controle/abastecimentos", response_class=HTMLResponse)
def controle_pagina_abastecimentos(request: Request):
    return pagina_modulo_controle("abastecimentos", request)


@app.get("/controle/checklist-viatura", response_class=HTMLResponse)
def controle_pagina_checklist_viatura(request: Request):
    return pagina_modulo_controle("checklist-viatura", request)


@app.get("/controle/manutencoes", response_class=HTMLResponse)
def controle_pagina_manutencoes(request: Request):
    return pagina_modulo_controle("manutencoes", request)


@app.get("/controle/multas", response_class=HTMLResponse)
def controle_pagina_multas(request: Request):
    return pagina_modulo_controle("multas", request)


@app.get("/controle/seguros", response_class=HTMLResponse)
def controle_pagina_seguros(request: Request):
    return pagina_modulo_controle("seguros", request)


@app.get("/api/controle/danos")
def api_controle_lista_danos():
    return {"ok": True, "itens": listar_registros_controle("controle_danos")}


@app.get("/api/controle/abastecimentos")
def api_controle_lista_abastecimentos():
    return {"ok": True, "itens": listar_registros_controle("controle_abastecimentos")}


@app.get("/api/controle/checklist-viatura")
def api_controle_lista_checklist_viatura():
    return {"ok": True, "itens": listar_registros_controle("controle_checklists_viatura")}


@app.get("/api/controle/manutencoes")
def api_controle_lista_manutencoes():
    return {"ok": True, "itens": listar_registros_controle("controle_manutencoes")}


@app.get("/api/controle/multas")
def api_controle_lista_multas():
    return {"ok": True, "itens": listar_registros_controle("controle_multas")}


@app.get("/api/controle/seguros")
def api_controle_lista_seguros():
    return {"ok": True, "itens": listar_registros_controle("controle_seguros")}


@app.post("/api/controle/danos")
async def api_controle_post_danos(request: Request):
    return await api_controle_salvar("danos", request)


@app.post("/api/controle/abastecimentos")
async def api_controle_post_abastecimentos(request: Request):
    return await api_controle_salvar("abastecimentos", request)


@app.post("/api/controle/checklist-viatura")
async def api_controle_post_checklist_viatura(request: Request):
    return await api_controle_salvar("checklist-viatura", request)


@app.post("/api/controle/manutencoes")
async def api_controle_post_manutencoes(request: Request):
    return await api_controle_salvar("manutencoes", request)


@app.post("/api/controle/multas")
async def api_controle_post_multas(request: Request):
    return await api_controle_salvar("multas", request)


@app.post("/api/controle/seguros")
async def api_controle_post_seguros(request: Request):
    return await api_controle_salvar("seguros", request)



@app.get('/central', response_class=HTMLResponse)
def central(
    request: Request,
    statuses: str = "",
    status: str = "",
    data_ini: str = "",
    data_fim: str = "",
    protocolo: str = "",
    placa: str = "",
    seguradora: str = "",
    motorista: str = "",
    tipo: str = "",
    origem: str = "",
    destino: str = "",
    checklist: str = "",
    status_faturamento: str = "",
):
    filtros = {
        "statuses": statuses,
        "status": status,
        "data_ini": data_ini,
        "data_fim": data_fim,
        "protocolo": protocolo,
        "placa": placa,
        "seguradora": seguradora,
        "motorista": motorista,
        "tipo": tipo,
        "origem": origem,
        "destino": destino,
        "checklist": checklist,
        "status_faturamento": status_faturamento,
    }
    titulo_filtro = _central_titulo_filtros(filtros)
    return templates.TemplateResponse(
        'index.html',
        {
            'request': request,
            'motoristas': lista_motoristas(),
            'servicos': lista_servicos_central_filtrada(**filtros),
            'lan_ip': get_lan_ip(),
            'tipos_servico': lista_tipos_servico(),
            'precos_por_tipo': {t: itens_padrao_tipo(t) for t in lista_tipos_servico()},
            'filtros': filtros,
            'filtro_statuses': statuses,
            'filtro_status': status,
            'titulo_filtro': titulo_filtro,
            'nav_ativo': 'central',
            'nav_som': True,
            'google_maps_api_key': google_api_key(),
            'status_operacionais': [
                'novo', 'enviado', 'aceito', 'a caminho', 'na origem',
                'em transporte', 'finalizado', 'recusado', 'cancelado',
            ],
        },
    )


@app.get('/servicos')
def servicos_alias():
    return RedirectResponse('/central', status_code=303)


@app.get('/servicos/novo', response_class=HTMLResponse)
def pagina_novo_servico(request: Request):
    return templates.TemplateResponse('novo_servico.html', {
        'request': request,
        'tipos_servico': lista_tipos_servico(),
        'precos_por_tipo': {t: itens_padrao_tipo(t) for t in lista_tipos_servico()}
    })

@app.get('/servicos/importar', response_class=HTMLResponse)
def pagina_importar_servicos(request: Request):
    return templates.TemplateResponse('importar_servicos.html', {'request': request})

@app.get('/cadastros/motoristas/novo', response_class=HTMLResponse)
def pagina_cadastro_motorista(request: Request):
    return templates.TemplateResponse('cadastro_motorista.html', {'request': request})

@app.get('/cadastros/veiculos/novo', response_class=HTMLResponse)
def pagina_cadastro_veiculo(request: Request):
    return templates.TemplateResponse('cadastro_veiculo.html', {'request': request})

@app.get('/cadastros/usuarios/novo', response_class=HTMLResponse)
def pagina_cadastro_usuario(request: Request):
    return templates.TemplateResponse('cadastro_usuario.html', {'request': request})

@app.post('/cadastros/veiculos')
async def salvar_veiculo(request: Request):
    form = await request.form()
    q("insert into veiculos (placa,modelo,tipo,ano,renavam,observacao,ativo) values (%s,%s,%s,%s,%s,%s,true)", (
        form.get('placa','').strip(), form.get('modelo','').strip(), form.get('tipo','').strip(),
        form.get('ano','').strip(), form.get('renavam','').strip(), form.get('observacao','').strip()
    ))
    return RedirectResponse('/central', 303)

@app.post('/cadastros/usuarios')
async def salvar_usuario(request: Request):
    form = await request.form()
    q("insert into usuarios_sistema (nome,cpf,telefone,email,senha,perfil,ativo) values (%s,%s,%s,%s,%s,%s,true)", (
        form.get('nome','').strip(), form.get('cpf','').strip(), form.get('telefone','').strip(),
        form.get('email','').strip(), form.get('senha','').strip(), form.get('perfil','').strip()
    ))
    return RedirectResponse('/central', 303)


def _slug_parte_checklist(parte):
    txt = re.sub(r'[^a-zA-Z0-9]+', '_', (parte or 'parte').strip().lower())
    return txt.strip('_') or 'parte'


CAMPOS_TIPO_SERVICO_CHECKLIST = (
    'tipo_servico', 'tipo', 'categoria', 'descricao_servico', 'descricao',
    'tipo_veiculo', 'observacao', 'obs',
)

_MAPA_ARQUIVOS_CHECKLIST = {
    'leve': {
        'Frente': 'carro-frente.png',
        'Traseira': 'carro-traseira.png',
        'Lateral esquerda': 'carro-lateral-esquerda.png',
        'Lateral direita': 'carro-lateral-direita.png',
        'Teto': 'carro-frente.png',
        'Rodas': 'carro-lateral-esquerda.png',
        'Vidros': 'carro-frente.png',
        'Parachoque': 'carro-traseira.png',
        'Faróis': 'carro-frente.png',
        'Lanternas': 'carro-traseira.png',
    },
    'utilitario': {
        'Frente': 'utilitario_frente.png',
        'Traseira': 'utilitario_traseira.png',
        'Lateral esquerda': 'utilitario_lateral_esquerda.png',
        'Lateral direita': 'utilitario_lateral_direita.png',
        'Teto': 'utilitario_frente.png',
        'Rodas': 'utilitario_lateral_esquerda.png',
        'Vidros': 'utilitario_frente.png',
        'Parachoque': 'utilitario_traseira.png',
        'Faróis': 'utilitario_frente.png',
        'Lanternas': 'utilitario_traseira.png',
    },
    'pesado': {
        'Frente': 'pesado_frente.png',
        'Traseira': 'pesado_traseira.png',
        'Lateral esquerda': 'pesado_lateral_esquerda.png',
        'Lateral direita': 'pesado_lateral_direita.png',
        'Teto': 'pesado_frente.png',
        'Rodas': 'pesado_lateral_esquerda.png',
        'Vidros': 'pesado_frente.png',
        'Parachoque': 'pesado_traseira.png',
        'Faróis': 'pesado_frente.png',
        'Lanternas': 'pesado_traseira.png',
    },
    'moto': {
        'Frente': 'moto_frente.png',
        'Traseira': 'moto_traseira.png',
        'Lateral esquerda': 'moto_lateral_esquerda.png',
        'Lateral direita': 'moto_lateral_direita.png',
        'Teto': 'moto_frente.png',
        'Rodas': 'moto_lateral_esquerda.png',
        'Vidros': 'moto_frente.png',
        'Parachoque': 'moto_traseira.png',
        'Faróis': 'moto_frente.png',
        'Lanternas': 'moto_traseira.png',
    },
}


def normalizar_texto_tipo(raw):
    import unicodedata
    txt = unicodedata.normalize('NFD', str(raw or ''))
    txt = ''.join(c for c in txt if unicodedata.category(c) != 'Mn')
    txt = txt.upper()
    txt = re.sub(r'[—–\-_/|\.]+', ' ', txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt


def coletar_texto_tipo_servico(*fontes):
    partes = []
    for fonte in fontes:
        if not fonte:
            continue
        if isinstance(fonte, str):
            txt = fonte.strip()
            if txt:
                partes.append(txt)
            continue
        if isinstance(fonte, dict):
            for campo in CAMPOS_TIPO_SERVICO_CHECKLIST:
                valor = fonte.get(campo)
                if valor is not None and str(valor).strip():
                    partes.append(str(valor))
    return normalizar_texto_tipo(' '.join(partes))


def inferir_tipo_veiculo(*fontes):
    """Ordem: pesado → moto → utilitário → leve (espelha app motorista)."""
    t = coletar_texto_tipo_servico(*fontes)
    if not t:
        return 'leve'

    if re.search(
        r'\bEXTRA\s+PESAD|\bPESAD(O|S)?\b|\bPESAD\b|CAMINHAO|PRANCHAO|BITREM|RODOTREM|CARRETA|'
        r'\bTRUCK\b|R\s*E\s*PES|R\s*L\s*PES|REBOQUE\s+PES|GUINCHO\s+PES|'
        r'\bR\s+E\s+PES|\bR\s+L\s+PES',
        t,
    ):
        return 'pesado'

    if re.search(r'\bMOTO(CICLETA)?S?\b', t):
        return 'moto'

    if re.search(
        r'UTILIT|\bSUV\b|\bVAN\b|PICK\s*UP|PICKUP|FURG|SPRINTER|MASTER|DUCATO|GUINAUTO',
        t,
    ):
        return 'utilitario'

    if re.search(
        r'\bLEVE\b|C\s*MEC|MECANICA\s*LEVE|AUTOMOVEL|REBOQUE\s*LEVE|\bPATIN',
        t,
    ):
        return 'leve'

    if re.search(r'\bPES\b', t):
        return 'pesado'

    return 'leve'


def label_tipo_veiculo_checklist(tipo):
    if tipo == 'pesado':
        return 'Veículo pesado'
    if tipo == 'utilitario':
        return 'Utilitário / SUV / Van'
    if tipo == 'moto':
        return 'Motocicleta'
    return 'Veículo leve'


def mapa_imagens_checklist_central(tipo):
    arquivos = _MAPA_ARQUIVOS_CHECKLIST.get(tipo) or _MAPA_ARQUIVOS_CHECKLIST['leve']
    return {parte: f'/static/checklist/{nome}' for parte, nome in arquivos.items()}


def imagem_checklist_parte(parte, tipo):
    mapa = mapa_imagens_checklist_central(tipo)
    return mapa.get(parte) or mapa.get('Frente')


def _protocolo_checklist_slug(protocolo, servico_id):
    base = (protocolo or servico_id or 'servico').strip()
    return re.sub(r'[^\w\-]+', '_', base) or 'servico'


def resolver_url_foto_checklist(foto, request=None):
    """
    Normaliza referência de foto do checklist para uso em img src / href.
    Aceita base64, data URI, caminhos relativos e URLs absolutas.
    """
    raw = (foto if isinstance(foto, str) else str(foto or '')).strip()
    if not raw:
        return {'url': '', 'valid': False, 'raw': raw, 'erro': 'vazio'}

    if raw.startswith('data:image'):
        return {'url': raw, 'valid': True, 'raw': raw, 'erro': ''}

    low = raw.lower()
    if low.startswith('file://') or low.startswith('content://') or low.startswith('ph://'):
        return {'url': '', 'valid': False, 'raw': raw, 'erro': 'uri_local_dispositivo'}

    if low.startswith('http://') or low.startswith('https://'):
        return {'url': raw, 'valid': True, 'raw': raw, 'erro': ''}

    if raw.startswith('/static/'):
        url = raw
        if request is not None:
            url = str(request.base_url).rstrip('/') + raw
        return {'url': url, 'valid': True, 'raw': raw, 'erro': ''}

    if raw.startswith('/uploads/'):
        path = '/static' + raw
        url = path
        if request is not None:
            url = str(request.base_url).rstrip('/') + path
        return {'url': url, 'valid': True, 'raw': raw, 'erro': ''}

    if raw.startswith('uploads/'):
        path = '/static/' + raw.lstrip('/')
        url = path
        if request is not None:
            url = str(request.base_url).rstrip('/') + path
        return {'url': url, 'valid': True, 'raw': raw, 'erro': ''}

    if raw.startswith('/'):
        path = raw if raw.startswith('/static/') else '/static' + raw
        url = path
        if request is not None:
            url = str(request.base_url).rstrip('/') + path
        return {'url': url, 'valid': True, 'raw': raw, 'erro': ''}

    if len(raw) > 200 and not raw.startswith('/'):
        try:
            base64.b64decode(raw[:200] + '==', validate=False)
            return {'url': f'data:image/jpeg;base64,{raw}', 'valid': True, 'raw': raw, 'erro': ''}
        except Exception:
            pass

    path = '/static/' + raw.lstrip('/')
    url = path
    if request is not None:
        url = str(request.base_url).rstrip('/') + path
    return {'url': url, 'valid': True, 'raw': raw, 'erro': ''}


def _bytes_foto_checklist(foto_raw):
    raw = (foto_raw or '').strip()
    if not raw:
        return None, None

    if raw.startswith('data:image'):
        match = re.match(r'^data:image/([\w+.-]+);base64,(.+)$', raw, re.DOTALL | re.IGNORECASE)
        if match:
            ext = match.group(1).lower().replace('jpeg', 'jpg')
            try:
                return base64.b64decode(match.group(2)), ext
            except Exception:
                return None, None

    if len(raw) > 500 and not raw.startswith(('file://', 'http://', 'https://', '/')):
        try:
            return base64.b64decode(raw), 'jpg'
        except Exception:
            pass

    return None, None


def _tipo_foto_checklist_recebida(raw):
    low = (raw or '').strip().lower()
    if low.startswith('data:image'):
        return 'base64'
    if low.startswith('file://') or low.startswith('ph://') or low.startswith('content://'):
        return 'file_uri'
    if low.startswith('http://') or low.startswith('https://'):
        return 'url'
    if low.startswith('/static/'):
        return 'static'
    if len(raw or '') > 200:
        return 'base64'
    return 'outro'


def persistir_fotos_checklist_lista(servico_id, parte, fotos, protocolo=''):
    """Converte base64/data URI em arquivos servidos em /static/uploads/checklist/."""
    persistidas = []
    slug = _slug_parte_checklist(parte)
    prot = _protocolo_checklist_slug(protocolo, servico_id)

    for i, foto in enumerate(fotos or []):
        raw = (foto if isinstance(foto, str) else str(foto or '')).strip()
        if not raw:
            continue

        tipo = _tipo_foto_checklist_recebida(raw)
        print(f'[CHECKLIST FOTO] recebida tipo={tipo} parte={parte} index={i + 1}')

        if raw.startswith('/static/uploads/checklist/'):
            persistidas.append(raw)
            print(f'[CHECKLIST FOTO] salva_em={raw}')
            continue

        low = raw.lower()
        if low.startswith('file://') or low.startswith('ph://') or low.startswith('content://'):
            print(f'[CHECKLIST FOTO] indisponivel=file_uri parte={parte} index={i + 1}')
            continue

        resolved = resolver_url_foto_checklist(raw)
        if resolved['valid'] and resolved['url'].startswith(('http://', 'https://')):
            persistidas.append(resolved['url'])
            print(f'[CHECKLIST FOTO] salva_em={resolved["url"]}')
            continue

        img_bytes, ext = _bytes_foto_checklist(raw)
        if img_bytes:
            dir_path = os.path.join(CHECKLIST_UPLOAD_DIR, str(servico_id))
            os.makedirs(dir_path, exist_ok=True)
            ext = (ext or 'jpg').replace('jpeg', 'jpg')
            fname = f"checklist_{prot}_{slug}_{i + 1}.{ext}"
            fpath = os.path.join(dir_path, fname)
            with open(fpath, 'wb') as fh:
                fh.write(img_bytes)
            url_salva = f"/static/uploads/checklist/{servico_id}/{fname}"
            persistidas.append(url_salva)
            print(f'[CHECKLIST FOTO] salva_em={url_salva}')
            continue

        if resolved['valid'] and resolved['url']:
            if resolved['url'].startswith('data:image'):
                img_bytes2, ext2 = _bytes_foto_checklist(resolved['url'])
                if img_bytes2:
                    dir_path = os.path.join(CHECKLIST_UPLOAD_DIR, str(servico_id))
                    os.makedirs(dir_path, exist_ok=True)
                    ext2 = (ext2 or 'jpg').replace('jpeg', 'jpg')
                    fname = f"checklist_{prot}_{slug}_{i + 1}.{ext2}"
                    fpath = os.path.join(dir_path, fname)
                    with open(fpath, 'wb') as fh:
                        fh.write(img_bytes2)
                    url_salva = f"/static/uploads/checklist/{servico_id}/{fname}"
                    persistidas.append(url_salva)
                    print(f'[CHECKLIST FOTO] salva_em={url_salva}')
                    continue
            if resolved['raw'].startswith('/static/'):
                persistidas.append(resolved['raw'])
                print(f'[CHECKLIST FOTO] salva_em={resolved["raw"]}')
                continue

        print(f'[CHECKLIST FOTO] ignorada tipo={tipo} parte={parte} index={i + 1}')

    return persistidas


def partes_checklist_dict(servico_id):
    """Retorna { parte: { marcacoes, fotos, observacao } } para um serviço."""
    rows = q(
        """
        select parte, marcacoes, fotos, observacao
          from checklist_avarias
         where servico_id = %s
         order by parte
        """,
        (str(servico_id),),
        True,
    )
    partes = {}
    for row in rows or []:
        dados_row = dict(row)
        partes[dados_row["parte"]] = {
            "marcacoes": dados_row.get("marcacoes") or [],
            "fotos": dados_row.get("fotos") or [],
            "observacao": extrair_observacao_checklist_parte(dados_row),
        }
    return partes


def assinatura_base64_valida(valor):
    texto = (valor or "").strip()
    return len(texto) > 80


def obter_assinaturas_checklist(servico_id):
    row = one(
        """
        select assinatura_origem_base64,
               assinatura_destino_base64,
               assinatura_base64,
               origem_atualizada_em,
               destino_atualizada_em,
               updated_at
          from checklist_assinaturas
         where servico_id=%s
        """,
        (str(servico_id),),
    )
    if not row:
        return {
            "assinatura_origem": None,
            "assinatura_origem_salva": False,
            "assinatura_destino": None,
            "assinatura_destino_salva": False,
        }

    origem = (row.get("assinatura_origem_base64") or "").strip()
    destino = (row.get("assinatura_destino_base64") or "").strip()
    legado = (row.get("assinatura_base64") or "").strip()

    if not origem and legado:
        origem = legado

    return {
        "assinatura_origem": origem or None,
        "assinatura_origem_salva": assinatura_base64_valida(origem),
        "assinatura_origem_em": dt_str(row.get("origem_atualizada_em")),
        "assinatura_destino": destino or None,
        "assinatura_destino_salva": assinatura_base64_valida(destino),
        "assinatura_destino_em": dt_str(row.get("destino_atualizada_em")),
        # Compatibilidade com checklists antigos / clientes legados
        "assinatura": origem or None,
        "assinatura_salva": assinatura_base64_valida(origem),
        "assinatura_atualizada_em": dt_str(row.get("origem_atualizada_em") or row.get("updated_at")),
    }


def salvar_assinatura_checklist(servico_id, tipo, assinatura_base64):
    """
    Persiste assinatura de origem ou destino.
    Nunca substitui assinatura já existente do mesmo tipo.
    """
    tipo = (tipo or "").strip().lower()
    if tipo not in ("origem", "destino"):
        return False, "tipo_invalido"

    assinatura = (assinatura_base64 or "").strip()
    if not assinatura_base64_valida(assinatura):
        return False, "vazio"

    row = one(
        "select assinatura_origem_base64, assinatura_destino_base64 from checklist_assinaturas where servico_id=%s",
        (str(servico_id),),
    )

    col = "assinatura_origem_base64" if tipo == "origem" else "assinatura_destino_base64"
    ts_col = "origem_atualizada_em" if tipo == "origem" else "destino_atualizada_em"

    if row and assinatura_base64_valida(row.get(col)):
        return False, "ja_existe"

    if not row:
        dados = {
            "servico_id": str(servico_id),
            "origem": assinatura if tipo == "origem" else None,
            "destino": assinatura if tipo == "destino" else None,
        }
        q(
            f"""
            insert into checklist_assinaturas
              (servico_id, {col}, {ts_col}, assinatura_base64, created_at, updated_at)
            values (%s, %s, now(), %s, now(), now())
            """,
            (
                dados["servico_id"],
                assinatura,
                assinatura if tipo == "origem" else None,
            ),
        )
        return True, "ok"

    q(
        f"""
        update checklist_assinaturas
           set {col}=%s,
               {ts_col}=now(),
               updated_at=now()
         where servico_id=%s
           and coalesce({col}, '') = ''
        """,
        (assinatura, str(servico_id)),
    )
    return True, "ok"


def processar_assinaturas_payload(servico_id, dados):
    """Salva assinaturas enviadas no payload sem apagar as existentes."""
    resultados = []
    if "assinatura_origem" in dados:
        ok, status = salvar_assinatura_checklist(servico_id, "origem", dados.get("assinatura_origem"))
        resultados.append(("origem", ok, status))
    if "assinatura_destino" in dados:
        ok, status = salvar_assinatura_checklist(servico_id, "destino", dados.get("assinatura_destino"))
        resultados.append(("destino", ok, status))
    if "assinatura" in dados and "assinatura_origem" not in dados:
        ok, status = salvar_assinatura_checklist(servico_id, "origem", dados.get("assinatura"))
        resultados.append(("origem", ok, status))
    return resultados


def tem_assinatura_origem(servico_id, dados=None):
    dados = dados or {}
    if assinatura_base64_valida(dados.get("assinatura_origem")):
        return True
    if assinatura_base64_valida(dados.get("assinatura")):
        return True
    info = obter_assinaturas_checklist(servico_id)
    return bool(info.get("assinatura_origem_salva"))


def resposta_checklist_dados(servico_id):
    partes = partes_checklist_dict(servico_id)
    assinaturas = obter_assinaturas_checklist(servico_id)
    return {
        "partes": partes,
        **assinaturas,
    }


@app.get("/servicos/{servico_id}/checklist", response_class=HTMLResponse)
async def pagina_checklist(servico_id: str, request: Request):
    s = servico_by_id(servico_id) or {}
    chk = checklist_resumo_servico(servico_id, request, s)
    assinaturas = chk.get('assinaturas') or obter_assinaturas_checklist(servico_id)

    return templates.TemplateResponse(
        'checklist_visualizar.html',
        {
            'request': request,
            'checklist': chk.get('checklist') or [],
            'checklist_resumo': chk,
            'servico_id': servico_id,
            'servico': s,
            'assinatura': assinaturas.get('assinatura_origem'),
            'assinatura_origem': assinaturas.get('assinatura_origem'),
            'assinatura_destino': assinaturas.get('assinatura_destino'),
        },
    )


@app.get("/api/checklist-dados/{servico_id}")
async def checklist_dados(servico_id: str):
    return resposta_checklist_dados(servico_id)


@app.get("/api/checklist-json/{servico_id}")
async def checklist_json(servico_id: str):
    return resposta_checklist_dados(servico_id)


@app.post("/api/checklist/salvar")
async def salvar_checklist(request: Request):
    dados = await request.json()

    servico_id = dados.get("servico_id")
    checklist = dados.get("checklist", {})

    if not servico_id:
        return JSONResponse({"ok": False, "erro": "servico_id não informado"}, status_code=400)

    if not isinstance(checklist, dict):
        return JSONResponse({"ok": False, "erro": "checklist deve ser um objeto"}, status_code=400)

    servico = servico_by_id(servico_id) or {}
    protocolo = servico.get('protocolo') or servico_id

    with get_conn() as conn:
        with conn.cursor() as cur:
            for parte, info in checklist.items():
                info = info or {}

                marcacoes = info.get("marcacoes", [])
                observacao = extrair_observacao_checklist_parte(info)
                fotos = persistir_fotos_checklist_lista(
                    servico_id, parte, info.get("fotos", []), protocolo
                )

                cur.execute("""
                    delete from checklist_avarias
                    where servico_id = %s and parte = %s
                """, (servico_id, parte))

                cur.execute("""
                    insert into checklist_avarias
                    (servico_id, parte, marcacoes, fotos, observacao)
                    values (%s, %s, %s::jsonb, %s::jsonb, %s)
                """, (
                    servico_id,
                    parte,
                    json.dumps(marcacoes),
                    json.dumps(fotos),
                    observacao or None,
                ))

        conn.commit()

    processar_assinaturas_payload(servico_id, dados)

    assinaturas = obter_assinaturas_checklist(servico_id)

    if not tem_assinatura_origem(servico_id, dados):
        return JSONResponse(
            {
                "ok": False,
                "erro": "Assinatura de origem é obrigatória para finalizar o checklist.",
            },
            status_code=400,
        )

    return {
        "ok": True,
        "assinatura_origem_salva": bool(assinaturas.get("assinatura_origem_salva")),
        "assinatura_destino_salva": bool(assinaturas.get("assinatura_destino_salva")),
        "assinatura_salva": bool(assinaturas.get("assinatura_origem_salva")),
    }

@app.get('/motoristas', response_class=HTMLResponse)
def pagina_motoristas(request: Request, online: str = ""):
    motoristas = q("""
        select *
          from motoristas
         order by coalesce(ativo,true) desc, nome asc
    """, fetch=True)
    if online in ("1", "true", "sim"):
        motoristas = [m for m in motoristas if m.get("online")]
    return templates.TemplateResponse(
        'motoristas_lista.html',
        {
            'request': request,
            'motoristas': motoristas,
            'filtro_online': online in ("1", "true", "sim"),
            'nav_ativo': 'motoristas',
            'nav_som': False,
        },
    )

@app.post('/motoristas/{mid}/editar')
async def editar_motorista(mid: str, request: Request):
    form = await request.form()
    ativo = True if form.get('ativo') == 'on' else False
    online = True if form.get('online') == 'on' else False

    q("""
        update motoristas
           set nome=%s,
               telefone=%s,
               veiculo=%s,
               placa=%s,
               placa_atual=%s,
               tipo=%s,
               cpf=%s,
               cnh=%s,
               vencimento_cnh=%s,
               nascimento=%s,
               estado_civil=%s,
               endereco=%s,
               login=%s,
               senha=%s,
               ativo=%s,
               online=%s,
               ultima_atualizacao=now()
         where id=%s
    """, (
        form.get('nome','').strip(),
        form.get('telefone','').strip(),
        form.get('veiculo','').strip(),
        form.get('placa','').strip().upper(),
        form.get('placa_atual','').strip().upper() or form.get('placa','').strip().upper(),
        form.get('tipo','').strip(),
        form.get('cpf','').strip(),
        form.get('cnh','').strip(),
        form.get('vencimento_cnh','').strip(),
        form.get('nascimento','').strip(),
        form.get('estado_civil','').strip(),
        form.get('endereco','').strip(),
        form.get('login','').strip(),
        form.get('senha','').strip(),
        ativo,
        online,
        str(mid)
    ))
    return RedirectResponse('/motoristas', 303)

@app.post('/motoristas/{mid}/offline')
def forcar_motorista_offline(mid: str):
    q("update motoristas set online=false, ultima_atualizacao=now() where id=%s", (str(mid),))
    return RedirectResponse('/motoristas', 303)

@app.post('/motoristas/{mid}/excluir')
def excluir_motorista(mid: str):
    # Não apaga histórico; apenas inativa o motorista.
    q("update motoristas set ativo=false, online=false, ultima_atualizacao=now() where id=%s", (str(mid),))
    return RedirectResponse('/motoristas', 303)


@app.get('/historico', response_class=HTMLResponse)
def historico(request: Request, data_ini: str="", data_fim: str="", seguradora: str="", tipo: str="", status: str="", motorista: str=""):
    filtros={k:(v or None) for k,v in dict(data_ini=data_ini,data_fim=data_fim,seguradora=seguradora,tipo=tipo,status=status,motorista=motorista).items()}
    servs=lista_servicos(filtros=filtros); total=len(servs); finalizados=len([s for s in servs if s.get("status")=="finalizado"]); ativos=len([s for s in servs if s.get("status") not in ["finalizado","recusado"]]); com_placa=len([s for s in servs if s.get("placa_veiculo_removido") or s.get("placa_removida")])
    df=pd.DataFrame(servs); por_seg=[]; por_tipo=[]; por_mot=[]
    if not df.empty:
        por_seg=df.groupby(df["seguradora"].fillna("").replace("", "Sem seguradora")).size().reset_index(name="total").rename(columns={"seguradora":"nome"}).to_dict("records")
        por_tipo=df.groupby(df["tipo"].fillna("").replace("", "Sem tipo")).size().reset_index(name="total").rename(columns={"tipo":"nome"}).to_dict("records")
        por_mot=df.groupby(df["motorista_nome"].fillna("").replace("", "Sem motorista")).size().reset_index(name="total").rename(columns={"motorista_nome":"nome"}).to_dict("records")
    return templates.TemplateResponse('historico.html', {'request':request,'servicos':servs,'filtros':filtros,'kpis':{"total":total,"finalizados":finalizados,"ativos":ativos,"com_placa":com_placa},'por_seg':por_seg,'por_tipo':por_tipo,'por_mot':por_mot,'nav_ativo':'relatorios','nav_som':False})

@app.get('/operacao', response_class=HTMLResponse)
def operacao(request: Request):
    ms=lista_motoristas(); servs=lista_servicos(ativos=True,limit=150); ultimos=lista_servicos(limit=20); online=len([m for m in ms if m.get("online")])
    return templates.TemplateResponse('operacao.html', {'request':request,'motoristas':ms,'servicos':servs,'ultimos':ultimos,'online':online,'total_motoristas':len(ms),'nav_ativo':'operacao','nav_som':False})

@app.post('/motoristas')
async def criar_motorista(request: Request):
    form = await request.form()
    q("""insert into motoristas (nome,telefone,veiculo,placa,placa_atual,tipo,cpf,cnh,vencimento_cnh,nascimento,estado_civil,endereco,login,senha,online,ultima_atualizacao)
         values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,false,now())""", (
        form.get('nome','').strip(), form.get('telefone','').strip(), form.get('veiculo','').strip(), form.get('placa','').strip(),
        form.get('placa','').strip().upper(), form.get('tipo','').strip(), form.get('cpf','').strip(), form.get('cnh','').strip(), form.get('vencimento_cnh','').strip(),
        form.get('nascimento','').strip(), form.get('estado_civil','').strip(), form.get('endereco','').strip(),
        form.get('login','').strip(), form.get('senha','').strip()
    ))
    return RedirectResponse('/central', 303)
@app.post('/servicos')
async def criar_servico(request: Request):
    form = await request.form()
    protocolo = form.get('protocolo','').strip()
    seguradora = form.get('seguradora','').strip()
    tipo = normalizar_tipo_importado(form.get('tipo','').strip())
    origem = form.get('origem','').strip()
    destino = form.get('destino','').strip()
    observacao = form.get('observacao','').strip()
    nomes = form.getlist('item_nome')
    qtds = form.getlist('item_qtd')
    valores = form.getlist('item_valor')
    row=one("insert into servicos (protocolo,seguradora,tipo,origem,destino,observacao,status,status_faturamento,criado_em,atualizado_em) values (%s,%s,%s,%s,%s,%s,'novo','para_conferir',now(),now()) returning id",(protocolo,seguradora,tipo,origem,destino,observacao))
    criar_itens_para_servico(row["id"], tipo, nomes, qtds, valores)
    registrar_evento_db(row["id"],"novo","Serviço criado com valores")
    return RedirectResponse('/central', 303)

@app.post('/servicos/importar')
async def importar_servicos(file: UploadFile=File(...)):
    content = await file.read()

    raw = pd.read_excel(io.BytesIO(content), header=None, dtype=object)

    header_idx = None
    for i in range(min(15, len(raw))):
        valores = [str(v).strip().lower() for v in raw.iloc[i].tolist() if pd.notna(v)]
        linha = " | ".join(valores)
        if "protocolo" in linha and ("o. logradouro" in linha or "origem" in linha or "d. logradouro" in linha):
            header_idx = i
            break

    if header_idx is None:
        df = pd.read_excel(io.BytesIO(content), dtype=object)
    else:
        headers = []
        usados = {}
        for v in raw.iloc[header_idx].tolist():
            h = str(v).strip() if pd.notna(v) and str(v).strip() else "Coluna"
            if h in usados:
                usados[h] += 1
                h = f"{h}_{usados[h]}"
            else:
                usados[h] = 1
            headers.append(h)
        df = raw.iloc[header_idx + 1:].copy()
        df.columns = headers

    df = df.dropna(how="all")
    cols = {str(c).lower().strip(): c for c in df.columns}

    def pick(names):
        for n in names:
            alvo = n.lower().strip()
            for k, v in cols.items():
                if alvo == k or alvo in k:
                    return v
        return None

    def val(row, col):
        if not col:
            return ""
        v = row.get(col, "")
        if pd.isna(v):
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    def juntar_endereco(row, prefixo):
        partes = []
        for col in [f"{prefixo}. Logradouro", f"{prefixo}. Bairro", f"{prefixo}. Cidade"]:
            v = val(row, col)
            if v and v.lower() != "nan":
                partes.append(v)
        return ", ".join(partes)

    c_prot = pick(["protocolo", "assistencia", "assistência"])
    c_seg = pick(["seguradora", "empresa", "produto"])
    c_tipo = pick(["tipo de serviço", "tipo servico", "tipo"])
    c_ori = pick(["origem completa", "origem"])
    c_des = pick(["destino completa", "destino completo", "destino"])
    c_obs = pick(["comentários", "comentarios", "observação", "observacao", "beneficiário", "beneficiario"])
    c_beneficiario = pick(["beneficiário", "beneficiario", "cliente", "nome cliente"])
    c_telefone = pick(["telefone", "telefone 01", "telefone 1", "celular"])
    c_veiculo = pick(["veículo / objeto", "veiculo / objeto", "veículo", "veiculo", "modelo"])
    c_placa_cliente = pick(["placa"])
    c_cor = pick(["cor"])
    c_cnpj = pick(["cnpj", "cnpj / filial"])
    c_ref_origem = pick(["o. referência", "o. referencia", "referência origem", "referencia origem"])
    c_ref_destino = pick(["d. referência", "d. referencia", "referência destino", "referencia destino"])

    importados = 0

    for _, row in df.iterrows():
        origem = val(row, c_ori)
        destino = val(row, c_des)

        if not origem:
            origem = juntar_endereco(row, "O")
        if not destino:
            destino = juntar_endereco(row, "D")

        protocolo = val(row, c_prot) or f"IMP-{uuid.uuid4().hex[:6]}"
        seguradora = val(row, c_seg)
        tipo = normalizar_tipo_importado(val(row, c_tipo))

        obs_partes = []
        obs_base = val(row, c_obs)
        if obs_base:
            obs_partes.append(obs_base)
        for label in ["Veículo / Objeto", "Placa", "Cor", "Beneficiário", "Telefone", "Senha", "CNPJ", "Status"]:
            v = val(row, label)
            if v:
                obs_partes.append(f"{label}: {v}")
        observacao = " | ".join(obs_partes) or "Importado do Excel AutEM"

        if origem and destino and origem.lower() != "nan" and destino.lower() != "nan":
            beneficiario = val(row, c_beneficiario)
            telefone_cliente = val(row, c_telefone)
            veiculo_cliente = val(row, c_veiculo)
            placa_cliente = val(row, c_placa_cliente)
            cor_cliente = val(row, c_cor)
            cnpj_cliente = val(row, c_cnpj)
            referencia_origem = val(row, c_ref_origem)
            referencia_destino = val(row, c_ref_destino)

            new = one(
                """insert into servicos (
                    protocolo, seguradora, tipo, origem, destino, observacao,
                    status, status_faturamento, placa_veiculo_removido, placa_removida,
                    beneficiario, telefone_cliente, veiculo_cliente, cor_cliente, cnpj_cliente,
                    referencia_origem, referencia_destino, criado_em, atualizado_em
                ) values (%s,%s,%s,%s,%s,%s,'novo','para_conferir',%s,%s,%s,%s,%s,%s,%s,%s,%s,now(),now()) returning id""",
                (
                    protocolo, seguradora, tipo, origem, destino, observacao,
                    placa_cliente, placa_cliente, beneficiario, telefone_cliente, veiculo_cliente,
                    cor_cliente, cnpj_cliente, referencia_origem, referencia_destino
                )
            )
            # CRÍTICO: serviços importados também ganham itens financeiros editáveis.
            criar_itens_para_servico(new["id"], tipo)
            registrar_evento_db(new["id"], "novo", "Importado do Excel AutEM com itens financeiros")
            importados += 1

    return RedirectResponse('/central?importados=' + str(importados), 303)




def google_api_key():
    return (
        os.environ.get("GOOGLE_MAPS_API_KEY")
        or os.environ.get("GOOGLE_API_KEY")
        or os.environ.get("CENTRAL_GOOGLE_MAPS_API_KEY")
        or ""
    ).strip()

def google_geocode_endereco(endereco):
    """
    Retorna (lat, lng) usando Google Geocoding, se GOOGLE_MAPS_API_KEY estiver configurada.
    """
    key = google_api_key()
    if not key or not endereco:
        return None, None

    try:
        params = urllib.parse.urlencode({
            "address": endereco,
            "key": key,
            "region": "br",
            "language": "pt-BR"
        })
        url = "https://maps.googleapis.com/maps/api/geocode/json?" + params
        with urllib.request.urlopen(url, timeout=6) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("status") == "OK" and data.get("results"):
            loc = data["results"][0]["geometry"]["location"]
            return loc.get("lat"), loc.get("lng")
    except Exception:
        pass

    return None, None

def google_distance_matrix(origem_lat, origem_lng, destino_lat, destino_lng):
    """
    Retorna (distancia_texto, duracao_texto, distancia_valor_metros).
    Usa Google Distance Matrix quando GOOGLE_MAPS_API_KEY estiver configurada.
    """
    key = google_api_key()
    if not key:
        return None, None, None

    try:
        params = urllib.parse.urlencode({
            "origins": f"{origem_lat},{origem_lng}",
            "destinations": f"{destino_lat},{destino_lng}",
            "mode": "driving",
            "language": "pt-BR",
            "region": "br",
            "key": key
        })
        url = "https://maps.googleapis.com/maps/api/distancematrix/json?" + params
        with urllib.request.urlopen(url, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        rows = data.get("rows") or []
        if rows and rows[0].get("elements"):
            el = rows[0]["elements"][0]
            if el.get("status") == "OK":
                dist = el.get("distance", {})
                dur = el.get("duration", {})
                return dist.get("text"), dur.get("text"), dist.get("value")
    except Exception:
        pass

    return None, None, None

    try:
        params = urllib.parse.urlencode({
            "origins": f"{origem_lat},{origem_lng}",
            "destinations": f"{destino_lat},{destino_lng}",
            "mode": "driving",
            "language": "pt-BR",
            "key": key
        })
        url = "https://maps.googleapis.com/maps/api/distancematrix/json?" + params
        with urllib.request.urlopen(url, timeout=6) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        rows = data.get("rows") or []
        if rows and rows[0].get("elements"):
            el = rows[0]["elements"][0]
            if el.get("status") == "OK":
                dist = el.get("distance", {})
                dur = el.get("duration", {})
                return dist.get("text"), dur.get("text"), dist.get("value")
    except Exception:
        pass

    return None, None, None

def obter_origem_lat_lng(servico):
    """
    Busca/coleta latitude e longitude da origem do serviço.
    Salva no banco para não ficar chamando o Google toda hora.
    """
    lat = servico.get("origem_lat")
    lng = servico.get("origem_lng")

    if lat and lng:
        return lat, lng

    endereco = servico.get("origem") or ""
    lat, lng = google_geocode_endereco(endereco)

    if lat and lng:
        try:
            q("update servicos set origem_lat=%s, origem_lng=%s where id=%s", (lat, lng, str(servico.get("id"))))
        except Exception:
            pass

    return lat, lng



def normalizar_endereco_google(endereco):
    endereco = (endereco or "").strip()
    if not endereco:
        return ""
    if "brasil" not in endereco.lower():
        endereco = f"{endereco}, Rio de Janeiro, RJ, Brasil"
    return endereco

def google_distance_matrix_endereco(origem_lat, origem_lng, destino_endereco):
    """
    Calcula rota usando origem em lat/lng do motorista e destino como endereço textual.
    Evita depender primeiro do Geocoding da origem do serviço.
    """
    key = google_api_key()
    destino_endereco = normalizar_endereco_google(destino_endereco)

    if not key:
        return None, None, None, "GOOGLE_MAPS_API_KEY não configurada"
    if not origem_lat or not origem_lng:
        return None, None, None, "motorista sem GPS"
    if not destino_endereco:
        return None, None, None, "origem do serviço vazia"

    try:
        params = urllib.parse.urlencode({
            "origins": f"{origem_lat},{origem_lng}",
            "destinations": destino_endereco,
            "mode": "driving",
            "language": "pt-BR",
            "region": "br",
            "key": key
        })
        url = "https://maps.googleapis.com/maps/api/distancematrix/json?" + params
        with urllib.request.urlopen(url, timeout=8) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        api_status = data.get("status")
        if api_status != "OK":
            return None, None, None, f"Google: {api_status}"

        rows = data.get("rows") or []
        if rows and rows[0].get("elements"):
            el = rows[0]["elements"][0]
            el_status = el.get("status")
            if el_status == "OK":
                dist = el.get("distance", {})
                dur = el.get("duration", {})
                return dist.get("text"), dur.get("text"), dist.get("value"), "rota Google"
            return None, None, None, f"Google rota: {el_status}"

        return None, None, None, "Google sem resultado"
    except Exception as e:
        return None, None, None, f"erro Google: {str(e)[:60]}"


def calcular_info_distancia_motorista(motorista, servico):
    """
    Textos amigáveis para a modal Enviar/trocar motorista na Central.
    Nunca expõe mensagens técnicas (ex.: nome de variável de ambiente).
    """
    mlat = motorista.get("lat")
    mlng = motorista.get("lng")
    origem_endereco = ((servico or {}).get("origem") or "").strip()

    if not mlat or not mlng:
        return {
            "distancia_texto": "Motorista sem GPS",
            "duracao_texto": "",
            "distancia_valor": None,
            "rota_disponivel": False,
            "calculo_status": "sem_gps",
        }

    if not origem_endereco:
        return {
            "distancia_texto": "Origem não informada",
            "duracao_texto": "",
            "distancia_valor": None,
            "rota_disponivel": False,
            "calculo_status": "sem_origem",
        }

    if not google_api_key():
        return {
            "distancia_texto": "Distância indisponível",
            "duracao_texto": "",
            "distancia_valor": None,
            "rota_disponivel": False,
            "calculo_status": "sem_chave",
        }

    dist_txt, dur_txt, dist_metros, calculo_status = google_distance_matrix_endereco(
        mlat, mlng, origem_endereco
    )

    if dist_txt and dur_txt:
        return {
            "distancia_texto": dist_txt,
            "duracao_texto": dur_txt,
            "distancia_valor": dist_metros,
            "rota_disponivel": True,
            "calculo_status": calculo_status or "rota_google",
        }

    if dist_txt:
        return {
            "distancia_texto": dist_txt,
            "duracao_texto": dur_txt or "",
            "distancia_valor": dist_metros,
            "rota_disponivel": False,
            "calculo_status": calculo_status or "parcial",
        }

    origem_lat, origem_lng = obter_origem_lat_lng(servico) if servico else (None, None)
    if origem_lat and origem_lng:
        dk = distancia_km(mlat, mlng, origem_lat, origem_lng)
        if dk is not None:
            return {
                "distancia_texto": f"{dk} km aprox.",
                "duracao_texto": "",
                "distancia_valor": int(dk * 1000),
                "rota_disponivel": False,
                "calculo_status": "aproximado",
            }

    return {
        "distancia_texto": "Distância indisponível",
        "duracao_texto": "",
        "distancia_valor": None,
        "rota_disponivel": False,
        "calculo_status": "indisponivel",
    }


def _norm_seguradora(txt):
    import re
    t = (txt or "").upper().strip()
    t = re.sub(r"[^A-Z0-9]", "", t.replace("Ç", "C").replace("Ã", "A").replace("Õ", "O").replace("É", "E"))
    return t


def _alerta_personalizacao_viatura(seguradora_servico, personalizacao_viatura):
    seg = _norm_seguradora(seguradora_servico)
    pers = _norm_seguradora(personalizacao_viatura)
    if not pers or not seg:
        return False, ""
    if seg in pers or pers in seg:
        return False, ""
    return True, (
        f"ATENÇÃO: viatura personalizada para {personalizacao_viatura.strip()} "
        f"e o serviço é da seguradora {seguradora_servico.strip()}."
    )


@app.get('/api/servicos/{sid}/motoristas')
def api_motoristas_para_servico(sid: str):
    s = servico_by_id(sid)
    ms = lista_motoristas()
    seguradora_srv = (s.get("seguradora") or "").strip() if s else ""

    saida = []
    for m in ms:
        ocupado = motorista_ocupado(m.get("id"))
        info = calcular_info_distancia_motorista(m, s)
        placa_trab = (m.get("placa_atual") or m.get("placa") or "").strip()
        viatura = viatura_cadastro_por_placa(placa_trab)
        alerta_pers = False
        alerta_msg = ""
        if viatura:
            alerta_pers, alerta_msg = _alerta_personalizacao_viatura(
                seguradora_srv, viatura.get("personalizacao")
            )
        prof = profissional_cadastro_por_motorista_id(m.get("id"))
        if not prof:
            cpf_m = normalizar_cpf_profissional(m.get("cpf"))
            if cpf_m:
                prof = profissional_cadastro_por_cpf(cpf_m)
        nome_exib = (prof.get("nome_exibicao") if prof else None) or m.get("nome")
        nome_compl = (prof.get("nome_completo") if prof else None) or m.get("nome")

        saida.append({
            **m,
            "nome": nome_exib,
            "nome_exibicao": nome_exib,
            "nome_completo": nome_compl,
            "funcao": prof.get("funcao") if prof else (m.get("tipo") or ""),
            "tipo_profissional": prof.get("tipo_profissional") if prof else "motorista",
            "tipo_profissional_label": prof.get("tipo_profissional_label") if prof else "Motorista",
            "classificacao_vinculo": prof.get("classificacao_vinculo") if prof else "",
            "vinculo_label": prof.get("vinculo_label") if prof else "",
            "ocupado": bool(ocupado),
            "servico_ocupado": ocupado.get("protocolo") if ocupado else None,
            "viatura_exibicao": viatura.get("nome_exibicao") if viatura else (placa_trab or "-"),
            "viatura_placa": viatura.get("placa") if viatura else placa_trab,
            "viatura_origem": viatura.get("origem") if viatura else "",
            "viatura_origem_label": viatura.get("origem_label") if viatura else "",
            "viatura_personalizacao": viatura.get("personalizacao") if viatura else "",
            "viatura_personalizada": bool(viatura.get("personalizada")) if viatura else False,
            "viatura_classificacao": viatura.get("classificacao_original") if viatura else "",
            "alerta_personalizacao": alerta_pers,
            "alerta_personalizacao_msg": alerta_msg,
            "seguradora_servico": seguradora_srv,
            **info,
        })

    saida.sort(key=lambda x: (
        0 if x.get("online") and not x.get("ocupado") else 1 if x.get("online") else 2,
        999999999 if x.get("distancia_valor") is None else x.get("distancia_valor"),
        x.get("nome") or ""
    ))
    return saida



@app.get('/api/servicos/{sid}/debug-distancia')
def api_debug_distancia(sid: str):
    s = servico_by_id(sid)
    ms = lista_motoristas()
    online = [m for m in ms if m.get("online") and m.get("lat") and m.get("lng")]
    if not s:
        return {"ok": False, "erro": "serviço não encontrado"}
    if not online:
        return {"ok": False, "erro": "nenhum motorista online com GPS", "origem": s.get("origem")}
    m = online[0]
    d, t, v, st = google_distance_matrix_endereco(m.get("lat"), m.get("lng"), s.get("origem"))
    return {
        "ok": bool(d),
        "origem": s.get("origem"),
        "motorista": m.get("nome"),
        "lat": m.get("lat"),
        "lng": m.get("lng"),
        "distancia": d,
        "tempo": t,
        "status": st
    }

@app.post('/servicos/{sid}/enviar')
def enviar_servico(sid: str, motorista_id: str=Form(...)):
    s = servico_by_id(sid)
    if not s:
        return RedirectResponse('/central', 303)

    # Só bloqueia troca de motorista quando o serviço já foi finalizado.
    if (s.get("status") or "").lower() == "finalizado":
        registrar_evento_db(sid, 'troca bloqueada', 'Serviço finalizado não permite troca de motorista')
        return RedirectResponse('/central', 303)

    m = motorista_by_id(motorista_id)
    if m:
        placa_trabalho = m.get("placa_atual") or m.get("placa") or ""
        q("""
            update servicos
               set motorista_id=%s,
                   motorista_nome=%s,
                   status='enviado',
                   atualizado_em=now()
             where id=%s
        """, (str(motorista_id), m["nome"], str(sid)))
        registrar_evento_db(sid, 'enviado', f"Enviado/reencaminhado para {m['nome']} - placa atual: {placa_trabalho}")
        try:
            s_atualizado = servico_by_id(sid)
            if s_atualizado:
                enviar_push_expo_novo_servico(str(motorista_id), s_atualizado)
        except Exception as exc:
            print(f"[PUSH] enviar_servico: {exc}")
    return RedirectResponse('/central', 303)


class AppLoginPayload(BaseModel):
    login: str
    senha: str
    placa: str

class AppLocationPayload(BaseModel):
    lat: float
    lng: float

class AppStatusPayload(BaseModel):
    status: str
    detalhe: str = ""

class AppPlacaPayload(BaseModel):
    placa: str = ""
    placa_veiculo: str = ""

class AppFcmPayload(BaseModel):
    token: str

class AppPushTokenPayload(BaseModel):
    motorista_id: str
    push_token: str
    platform: str = ""

EXPO_PUSH_URL = "https://exp.host/--/api/v2/push/send"
# Som personalizado (até 20s) — requer Development Build; Expo Go usa som padrão do sistema.
PUSH_SOUND_NOVO_SERVICO = os.environ.get("PUSH_SOUND", "ambulance_siren.wav")


def motorista_push_token(mid):
    m = motorista_by_id(mid)
    if not m:
        return None
    token = (m.get("push_token") or m.get("fcm_token") or "").strip()
    return token or None


def _limpar_push_token_invalido(motorista_id, erro_expo):
    if erro_expo in ("DeviceNotRegistered", "InvalidCredentials", "InvalidToken"):
        print(f"[PUSH] token inválido ({erro_expo}) — limpando motorista {motorista_id}")
        q("update motoristas set push_token=null, push_platform=null where id=%s", (str(motorista_id),))


def enviar_push_expo(motorista_id, token, titulo, corpo, data_extra=None, protocolo_log="", sound=None, channel_id=None):
    """Envia push via Expo Push Service e registra payload, status e resposta."""
    if not token:
        print(f"[PUSH] motorista {motorista_id} sem push_token cadastrado")
        return False, {"erro": "sem_token"}

    mensagem = [{
        "to": token,
        "title": titulo,
        "body": corpo,
        "sound": sound if sound is not None else PUSH_SOUND_NOVO_SERVICO,
        "priority": "high",
        "channelId": channel_id if channel_id is not None else "novo_servico",
        "data": data_extra or {},
    }]

    print(f"[PUSH] enviando para motorista {motorista_id}")
    print(f"[PUSH] token {token}")
    if protocolo_log:
        print(f"[PUSH] protocolo {protocolo_log}")
    print(f"[PUSH] payload {json.dumps(mensagem, ensure_ascii=False)}")

    try:
        req = urllib.request.Request(
            EXPO_PUSH_URL,
            data=json.dumps(mensagem).encode("utf-8"),
            headers={
                "Accept": "application/json",
                "Accept-encoding": "gzip, deflate",
                "Content-Type": "application/json",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=12) as resp:
            status = resp.status
            body = resp.read().decode("utf-8", errors="replace")
        print(f"[PUSH] status {status}")
        print(f"[PUSH] resposta {body}")

        resultado = json.loads(body) if body else {}
        tickets = resultado.get("data") or []
        if tickets:
            ticket = tickets[0]
            if ticket.get("status") == "error":
                detalhe = (ticket.get("details") or {})
                erro = detalhe.get("error") or ticket.get("message") or "erro_desconhecido"
                print(f"[PUSH] erro Expo: {erro}")
                _limpar_push_token_invalido(motorista_id, erro)
                return False, {"status": status, "expo": resultado, "erro": erro}
            ticket_id = ticket.get("id")
            if ticket_id:
                print(f"[PUSH] ticket {ticket_id}")

        return True, {"status": status, "expo": resultado}
    except Exception as exc:
        print(f"[PUSH] falha ao enviar para motorista {motorista_id}: {exc}")
        return False, {"erro": str(exc)}


def enviar_push_expo_novo_servico(motorista_id, servico):
    """Envia push quando a Central vincula um serviço ao motorista."""
    if not servico:
        print(f"[PUSH] servico ausente para motorista {motorista_id}")
        return False

    token = motorista_push_token(motorista_id)
    if not token:
        print(f"[PUSH] motorista {motorista_id} sem push_token — app precisa registrar token após login")
        return False

    protocolo = (servico.get("protocolo") or "").strip()
    tipo = (servico.get("tipo") or "SERVIÇO").strip().upper()
    sid = str(servico.get("id") or "")

    ok, _ = enviar_push_expo(
        motorista_id,
        token,
        "Novo serviço recebido",
        f"{protocolo} — {tipo}" if protocolo else tipo,
        {
            "servico_id": sid,
            "protocolo": protocolo,
            "type": "novo_servico",
        },
        protocolo_log=protocolo,
    )
    return ok

    return ok


def enviar_push_expo_mensagem_central(motorista_id, servico, texto_mensagem):
    """Push de nova mensagem da Central — som padrão (não usa sirene de novo serviço)."""
    token = motorista_push_token(motorista_id)
    if not token or not servico:
        return False

    protocolo = (servico.get("protocolo") or "").strip()
    sid = str(servico.get("id") or "")
    resumo = (texto_mensagem or "").replace("\n", " ").strip()
    if len(resumo) > 160:
        resumo = resumo[:157] + "..."

    ok, _ = enviar_push_expo(
        motorista_id,
        token,
        "Nova mensagem da Central",
        resumo or "Você tem uma nova mensagem",
        {
            "type": "mobile_message",
            "servico_id": sid,
            "protocolo": protocolo,
        },
        protocolo_log=protocolo,
        sound="default",
        channel_id="novo_servico",
    )
    return ok


class MobileMensagemPayload(BaseModel):
    remetente_tipo: str
    remetente_id: str = ""
    remetente_nome: str = ""
    mensagem: str


class MobileMarcarLidasPayload(BaseModel):
    leitor: str


@app.get("/api/servicos/{sid}/mobile/mensagens")
def api_servico_mobile_mensagens_listar(sid: str):
    if not servico_by_id(sid):
        return JSONResponse({"ok": False, "erro": "Serviço não encontrado."}, status_code=404)
    return {"ok": True, "mensagens": listar_mensagens_mobile_servico(sid)}


@app.post("/api/servicos/{sid}/mobile/mensagens")
def api_servico_mobile_mensagens_criar(sid: str, payload: MobileMensagemPayload):
    msg, erro = criar_mensagem_mobile_servico(
        sid,
        payload.remetente_tipo,
        payload.remetente_id,
        payload.remetente_nome,
        payload.mensagem,
    )
    if erro:
        status = 404 if "não encontrado" in erro.lower() else 400
        return JSONResponse({"ok": False, "erro": erro}, status_code=status)
    return {"ok": True, "mensagem": msg}


@app.post("/api/servicos/{sid}/mobile/marcar-lidas")
def api_servico_mobile_marcar_lidas(sid: str, payload: MobileMarcarLidasPayload):
    if not servico_by_id(sid):
        return JSONResponse({"ok": False, "erro": "Serviço não encontrado."}, status_code=404)
    marcar_mensagens_mobile_lidas(sid, payload.leitor)
    return {"ok": True}


@app.get("/api/servicos/{sid}/mobile/unread")
def api_servico_mobile_unread(sid: str, leitor: str = "central"):
    if not servico_by_id(sid):
        return JSONResponse({"ok": False, "erro": "Serviço não encontrado."}, status_code=404)
    return {
        "ok": True,
        "qtd": contar_mensagens_mobile_nao_lidas(sid, leitor),
    }


@app.get("/api/central/mobile/unread-summary")
def api_central_mobile_unread_summary():
    resumo = resumo_mensagens_mobile_central()
    return {"ok": True, **resumo}


@app.get("/api/central/alertas")
def api_central_alertas():
    return {"ok": True, **resumo_alertas_central()}


@app.post("/api/central/alertas/{tipo}/{sid}/marcar-visto")
def api_central_alerta_marcar_visto(tipo: str, sid: str):
    if not servico_by_id(sid):
        return JSONResponse({"ok": False, "erro": "Serviço não encontrado."}, status_code=404)
    marcar_alerta_operacional_visto(sid, tipo)
    return {"ok": True}


@app.get("/api/servicos/{sid}/localizacao-motorista")
def api_servico_localizacao_motorista(sid: str):
    s = servico_by_id(sid)
    if not s:
        return JSONResponse({"ok": False, "erro": "Serviço não encontrado."}, status_code=404)

    mid = s.get("motorista_id")
    m = motorista_by_id(mid) if mid else None
    origem_lat, origem_lng = obter_origem_lat_lng(s)
    destino_lat, destino_lng = None, None
    destino_end = (s.get("destino") or "").strip()
    if destino_end:
        destino_lat, destino_lng = google_geocode_endereco(destino_end)

    dist_txt = ""
    dur_txt = ""
    if m and m.get("lat") and m.get("lng"):
        info = calcular_info_distancia_motorista(m, s)
        dist_txt = info.get("distancia_texto") or ""
        dur_txt = info.get("duracao_texto") or ""

    return {
        "ok": True,
        "servico_id": str(sid),
        "protocolo": s.get("protocolo") or "",
        "motorista": {
            "id": str(mid) if mid else "",
            "nome": (m.get("nome") if m else s.get("motorista_nome")) or "",
            "lat": m.get("lat") if m else None,
            "lng": m.get("lng") if m else None,
            "online": bool(m.get("online")) if m else False,
            "atualizado_em": formatar_data_hora_central(m.get("ultima_atualizacao")) if m else "",
        },
        "origem": {
            "endereco": s.get("origem") or "",
            "lat": origem_lat,
            "lng": origem_lng,
        },
        "destino": {
            "endereco": destino_end,
            "lat": destino_lat,
            "lng": destino_lng,
        },
        "distancia_texto": dist_txt,
        "duracao_texto": dur_txt,
    }


@app.get("/api/central/config/maps")
def api_central_config_maps():
    key = google_api_key()
    return {
        "ok": True,
        "google_maps_api_key": key,
        "googleMapsApiKey": key,
        "configurado": bool(key),
    }


class LocationPayload(BaseModel):
    lat: float; lng: float; online: bool=True
@app.post('/api/motoristas/{mid}/localizacao')
def atualizar_localizacao(mid: str, payload: LocationPayload):
    q("""
        update motoristas
           set lat=%s,
               lng=%s,
               online=true,
               ultima_atualizacao=now()
         where id=%s
    """, (payload.lat, payload.lng, str(mid)))
    return {'ok': True, 'motorista': motorista_by_id(mid)}
@app.post('/api/motoristas/{mid}/offline')
def motorista_offline(mid: str):
    q("update motoristas set online=false,ultima_atualizacao=now() where id=%s",(str(mid),)); return {'ok':True}


def _vianet_txt(*valores):
    for v in valores:
        s = (v or "").strip()
        if s and s.lower() not in ("-", "nan", "none", "null"):
            return s
    return ""


_VIANET_LABELS_PROTOCOLO = {
    "solicitacao", "solicitação", "assistencia", "assistência",
    "protocolo", "protocolo externo", "local", "cidade", "uf", "estado",
    "bairro", "cep", "selecionar prestador",
}


def _vianet_norm_label(s):
    return (s or "").strip().lower().rstrip(":")


def _vianet_eh_label_invalido(valor):
    n = _vianet_norm_label(valor)
    if not n:
        return True
    if n in _VIANET_LABELS_PROTOCOLO:
        return True
    if n.endswith(":") or valor.strip().endswith(":"):
        return True
    return False


def _vianet_extrair_digitos(valor, min_len=1, max_len=12):
    if _vianet_eh_label_invalido(valor):
        return ""
    digits = re.sub(r"\D", "", str(valor or ""))
    if min_len <= len(digits) <= max_len:
        return digits
    return ""


def _vianet_protocolo_valido(protocolo):
    p = (protocolo or "").strip()
    if not p or _vianet_eh_label_invalido(p):
        return False
    if not re.search(r"\d", p):
        return False
    return True


def montar_protocolo_vianet(dados: dict):
    """Monta protocolo final: assistência-solicitação, ou assistência, ou protocolo externo."""
    assist = _vianet_extrair_digitos(
        _vianet_txt(dados.get("assistencia")), min_len=5, max_len=12
    )
    solicit = _vianet_extrair_digitos(
        _vianet_txt(dados.get("solicitacao")), min_len=1, max_len=4
    )
    externo_raw = _vianet_txt(dados.get("protocolo_externo"))
    externo = externo_raw if _vianet_protocolo_valido(externo_raw) else ""

    proto_capturado = _vianet_txt(dados.get("protocolo"))
    if _vianet_protocolo_valido(proto_capturado):
        if assist and solicit and proto_capturado == f"{assist}-{solicit}":
            return proto_capturado
        if assist and proto_capturado == assist:
            return proto_capturado
        if not assist and not solicit:
            return proto_capturado

    if assist and solicit:
        return f"{assist}-{solicit}"
    if assist:
        return assist
    if externo:
        return externo
    if _vianet_protocolo_valido(proto_capturado):
        return proto_capturado
    return ""


def sanitizar_campos_endereco_vianet(bairro, cidade, estado, cep):
    """Evita UF em bairro e CEP em cidade."""
    bairro = (bairro or "").strip()
    cidade = (cidade or "").strip()
    estado = (estado or "").strip().upper()
    cep = (cep or "").strip()

    if bairro and len(bairro) == 2 and bairro.isalpha():
        if not estado:
            estado = bairro.upper()
        bairro = ""

    if cidade:
        cep_digits = re.sub(r"\D", "", cidade)
        if re.fullmatch(r"\d{5,8}", cep_digits):
            if not cep:
                cep = cep_digits
            cidade = ""
        elif re.search(r"\(ref\s*:", cidade, re.I):
            cidade = re.sub(r"\s*\(ref\s*:.*\)\s*$", "", cidade, flags=re.I).strip()

    if estado and len(estado) > 2:
        uf_match = re.search(r"\b([A-Z]{2})\b", estado.upper())
        if uf_match:
            estado = uf_match.group(1)

    if cep:
        cep = re.sub(r"\D", "", cep)
        if len(cep) == 8:
            cep = f"{cep[:5]}-{cep[5:]}"

    return bairro, cidade, estado, cep


def montar_endereco_vianet(local, complemento, bairro, cidade, estado, cep, referencias=""):
    """Formato compatível com extrair_bairro_cidade da Central: local - bairro - cidade/UF."""
    bairro, cidade, estado, cep = sanitizar_campos_endereco_vianet(bairro, cidade, estado, cep)
    local = (local or "").strip()
    complemento = (complemento or "").strip()
    if complemento and complemento.lower() not in ("-", "nan"):
        local = f"{local}, {complemento}" if local else complemento

    partes = []
    if local:
        partes.append(local)
    if bairro:
        partes.append(bairro)
    if cidade:
        if estado and len(estado) == 2:
            partes.append(f"{cidade}/{estado}")
        else:
            partes.append(cidade)
    elif estado and len(estado) == 2:
        partes.append(estado)

    endereco = " - ".join(partes) if partes else ""
    if not endereco and cep:
        endereco = f"CEP {cep}"
    return endereco or "-"


def servico_existe_por_protocolo(protocolo):
    prot = (protocolo or "").strip()
    if not prot:
        return None
    return one(
        """
        select id, protocolo
          from servicos
         where lower(trim(coalesce(protocolo, ''))) = lower(%s)
         limit 1
        """,
        (prot,),
    )


def importar_servico_vianet(dados: dict):
    """Cria serviço a partir do payload da extensão Vianet/Mondial."""
    protocolo = montar_protocolo_vianet(dados)
    if not _vianet_protocolo_valido(protocolo):
        return {
            "ok": False,
            "erro": "Protocolo inválido. Abra a tela do serviço no Vianet.",
        }

    existente = servico_existe_por_protocolo(protocolo)
    if existente:
        return {
            "ok": False,
            "erro": "serviço já cadastrado",
            "id": str(existente.get("id")),
            "protocolo": existente.get("protocolo") or protocolo,
        }

    seguradora = _vianet_txt(dados.get("seguradora"), dados.get("produto"))
    tipo = normalizar_tipo_importado(
        _vianet_txt(dados.get("tipo_servico"), dados.get("servico"), dados.get("serviço"))
    )

    origem = montar_endereco_vianet(
        dados.get("local_origem"),
        dados.get("complemento_origem"),
        dados.get("bairro_origem"),
        dados.get("cidade_origem"),
        dados.get("estado_origem"),
        dados.get("cep_origem"),
        dados.get("referencias_origem"),
    )
    destino = montar_endereco_vianet(
        dados.get("local_destino"),
        dados.get("complemento_destino"),
        dados.get("bairro_destino"),
        dados.get("cidade_destino"),
        dados.get("estado_destino"),
        dados.get("cep_destino"),
        _vianet_txt(dados.get("referencia_destino"), dados.get("referencias_destino")),
    )

    placa_raw = _vianet_txt(dados.get("placa"))
    placa = normalizar_placa_cliente(placa_raw) if placa_raw else ""

    veiculo_partes = [
        _vianet_txt(dados.get("veiculo")),
        _vianet_txt(dados.get("ano")),
        _vianet_txt(dados.get("cor")),
        _vianet_txt(dados.get("combustivel")),
    ]
    veiculo_cliente = " · ".join([p for p in veiculo_partes if p])

    beneficiario = _vianet_txt(dados.get("segurado"), dados.get("beneficiario"))
    telefone_cliente = _vianet_txt(dados.get("telefone"))
    solicitante = _vianet_txt(dados.get("solicitante"))
    problema = _vianet_txt(dados.get("problema"))
    cor_cliente = _vianet_txt(dados.get("cor"))
    referencia_origem = _vianet_txt(dados.get("referencias_origem"))
    referencia_destino = _vianet_txt(
        dados.get("referencia_destino"), dados.get("referencias_destino")
    )

    obs_partes = ["Importado via extensão Essência Exportador (Vianet/Mondial)"]
    for rotulo, chave in [
        ("Aceito em", "aceito_em"),
        ("Prazo", "prazo"),
        ("Problema", "problema"),
        ("Solicitante", "solicitante"),
        ("Combustível", "combustivel"),
    ]:
        val = _vianet_txt(dados.get(chave))
        if val:
            obs_partes.append(f"{rotulo}: {val}")
    observacao = " | ".join(obs_partes)

    ts_importacao = agora_dt()

    row = one(
        """
        insert into servicos (
            protocolo, seguradora, tipo, origem, destino, observacao,
            status, status_faturamento,
            placa_veiculo_removido, placa_removida,
            beneficiario, telefone_cliente, veiculo_cliente, cor_cliente,
            referencia_origem, referencia_destino,
            solicitante, problema, fonte_importacao,
            criado_em, atualizado_em, created_at
        ) values (
            %s,%s,%s,%s,%s,%s,
            'novo','para_conferir',
            %s,%s,
            %s,%s,%s,%s,
            %s,%s,
            %s,%s,'vianet',
            %s,%s,%s
        ) returning id, protocolo
        """,
        (
            protocolo,
            seguradora,
            tipo,
            origem,
            destino,
            observacao,
            placa or None,
            placa or None,
            beneficiario or None,
            telefone_cliente or None,
            veiculo_cliente or None,
            cor_cliente or None,
            referencia_origem or None,
            referencia_destino or None,
            solicitante or None,
            problema or None,
            ts_importacao,
            ts_importacao,
            ts_importacao,
        ),
    )

    sid = str(row["id"])
    criar_itens_para_servico(row["id"], tipo)
    registrar_evento_db(sid, "novo", "Importado da extensão Vianet/Mondial")

    return {
        "ok": True,
        "id": sid,
        "protocolo": row.get("protocolo") or protocolo,
        "mensagem": "Serviço importado com sucesso.",
    }


class VianetImportPayload(BaseModel):
    protocolo: str = ""
    assistencia: str = ""
    solicitacao: str = ""
    protocolo_externo: str = ""
    seguradora: str = ""
    produto: str = ""
    tipo_servico: str = ""
    servico: str = ""
    aceito_em: str = ""
    prazo: str = ""
    local_origem: str = ""
    complemento_origem: str = ""
    bairro_origem: str = ""
    cidade_origem: str = ""
    estado_origem: str = ""
    cep_origem: str = ""
    referencias_origem: str = ""
    problema: str = ""
    segurado: str = ""
    solicitante: str = ""
    telefone: str = ""
    veiculo: str = ""
    placa: str = ""
    ano: str = ""
    cor: str = ""
    combustivel: str = ""
    local_destino: str = ""
    complemento_destino: str = ""
    referencia_destino: str = ""
    bairro_destino: str = ""
    cidade_destino: str = ""
    estado_destino: str = ""
    cep_destino: str = ""
    raw: dict | None = None


@app.post("/api/importar/vianet")
async def api_importar_vianet(payload: VianetImportPayload):
    dados = payload.model_dump()
    if payload.raw and isinstance(payload.raw, dict):
        for k, v in payload.raw.items():
            if k not in dados or not str(dados.get(k) or "").strip():
                dados[k] = v
    resultado = importar_servico_vianet(dados)
    status = 200 if resultado.get("ok") else (409 if resultado.get("erro") == "serviço já cadastrado" else 400)
    return JSONResponse(resultado, status_code=status)


@app.get('/api/motoristas')
def api_motoristas(): return lista_motoristas()
@app.get('/api/servicos')
def api_servicos(
    statuses: str = "",
    status: str = "",
    data_ini: str = "",
    data_fim: str = "",
    protocolo: str = "",
    placa: str = "",
    seguradora: str = "",
    motorista: str = "",
    tipo: str = "",
    origem: str = "",
    destino: str = "",
    checklist: str = "",
    status_faturamento: str = "",
):
    return lista_servicos_central_filtrada(
        statuses=statuses,
        status=status,
        data_ini=data_ini,
        data_fim=data_fim,
        protocolo=protocolo,
        placa=placa,
        seguradora=seguradora,
        motorista=motorista,
        tipo=tipo,
        origem=origem,
        destino=destino,
        checklist=checklist,
        status_faturamento=status_faturamento,
        limit=200,
    )


# ============================================================
# API DO APP ANDROID - ESSÊNCIA LOGÍSTICA / MOTORISTA
# ============================================================

@app.post("/api/app/motorista/login")
def api_app_motorista_login(payload: AppLoginPayload):
    login = (payload.login or "").strip()
    senha = (payload.senha or "").strip()
    placa = (payload.placa or "").upper().replace("-", "").replace(" ", "").strip()

    if not login or not senha or not placa:
        return JSONResponse({"ok": False, "erro": "Preencha login, senha e placa."}, status_code=400)

    m = one("""
        select * from motoristas
        where coalesce(ativo,true)=true
          and (
            lower(coalesce(login,''))=lower(%s)
            or regexp_replace(coalesce(cpf,''),'[^0-9]','','g')=regexp_replace(%s,'[^0-9]','','g')
          )
          and coalesce(senha,'')=%s
        limit 1
    """, (login, login, senha))

    if not m:
        return JSONResponse({"ok": False, "erro": "Login ou senha inválidos."}, status_code=401)

    mid = str(m["id"])
    q("""
        update motoristas
           set placa_atual=%s,
               placa=%s,
               online=true,
               ultima_atualizacao=now(),
               ultimo_login=now()
         where id=%s
    """, (placa, placa, mid))

    m = motorista_by_id(mid)
    return {
        "ok": True,
        "motorista": {
            "id": str(m.get("id")),
            "nome": m.get("nome"),
            "placa_atual": m.get("placa_atual") or m.get("placa"),
            "veiculo": m.get("veiculo"),
            "online": m.get("online")
        }
    }

@app.get("/api/app/motorista/{mid}/servicos")
def api_app_motorista_servicos(mid: str):
    rows = q("""
        select *
          from servicos
         where motorista_id=%s
           and coalesce(status,'novo') not in ('finalizado')
         order by created_at desc
         limit 50
    """, (str(mid),), fetch=True)
    return {"ok": True, "servicos": [normalizar_servico(r) for r in rows]}

@app.get("/api/app/motorista/{mid}/historico")
def api_app_motorista_historico(mid: str):
    rows = q("""
        select *
          from servicos
         where motorista_id=%s
         order by created_at desc nulls last
         limit 150
    """, (str(mid),), fetch=True)
    return {"ok": True, "servicos": [normalizar_servico(r) for r in rows]}

@app.post("/api/app/motorista/{mid}/localizacao")
def api_app_motorista_localizacao(mid: str, payload: AppLocationPayload):
    q("""
        update motoristas
           set lat=%s,
               lng=%s,
               online=true,
               ultima_atualizacao=now()
         where id=%s
    """, (payload.lat, payload.lng, str(mid)))
    return {"ok": True}

@app.post("/api/app/motorista/{mid}/offline")
def api_app_motorista_offline(mid: str):
    q("update motoristas set online=false, ultima_atualizacao=now() where id=%s", (str(mid),))
    return {"ok": True}

@app.post("/api/app/servicos/{sid}/status")
def api_app_servico_status(sid: str, payload: AppStatusPayload):
    status = (payload.status or "").strip().lower()
    permitidos = ["aceito", "recusado", "a caminho", "na origem", "em transporte", "finalizado"]
    if status not in permitidos:
        return JSONResponse({"ok": False, "erro": "Status inválido."}, status_code=400)

    if status == "na origem":
        servico = servico_by_id(sid)
        if not servico:
            return JSONResponse({"ok": False, "erro": "Serviço não encontrado."}, status_code=404)
        if not placa_cliente_valida(placa_cliente_do_servico(servico)):
            return JSONResponse(
                {
                    "ok": False,
                    "erro": "Informe a placa do veículo antes de marcar chegada na origem.",
                },
                status_code=400,
            )

    extra = ""
    if status == "finalizado":
        extra = ", finalizado_em=now()"

    q(f"update servicos set status=%s, atualizado_em=now(){extra} where id=%s", (status, str(sid)))
    registrar_evento_db(sid, status, payload.detalhe or f"Status atualizado pelo App do Motorista: {status}")
    if status == "recusado":
        s_atual = servico_by_id(sid)
        protocolo = (s_atual.get("protocolo") or "").strip() if s_atual else ""
        registrar_alerta_operacional(
            sid,
            "recusa",
            "Serviço recusado",
            f"Motorista recusou protocolo {protocolo}" if protocolo else "Motorista recusou o serviço",
        )
    return {"ok": True, "servico": servico_by_id(sid)}


@app.post("/api/app/servicos/{sid}/placa")
def api_app_servico_placa(sid: str, payload: AppPlacaPayload):
    placa = payload.placa_veiculo or payload.placa
    ok, erro = salvar_placa_cliente_servico(sid, placa)
    if not ok:
        return JSONResponse({"ok": False, "erro": erro}, status_code=400)
    return {"ok": True, "servico": servico_by_id(sid)}

@app.post("/api/app/motorista/push-token")
def api_app_motorista_push_token(payload: AppPushTokenPayload):
    mid = (payload.motorista_id or "").strip()
    token = (payload.push_token or "").strip()
    platform = (payload.platform or "").strip().lower()

    if not mid or not token:
        return JSONResponse({"ok": False, "erro": "motorista_id e push_token são obrigatórios."}, status_code=400)

    if not motorista_by_id(mid):
        return JSONResponse({"ok": False, "erro": "Motorista não encontrado."}, status_code=404)

    q(
        "update motoristas set push_token=%s, push_platform=%s where id=%s",
        (token, platform or None, str(mid)),
    )
    print(f"[PUSH] token salvo motorista={mid} platform={platform or '?'} token={token}")
    return {"ok": True, "motorista_id": mid, "platform": platform or None}


@app.get("/api/app/motorista/{mid}/push-status")
def api_app_motorista_push_status(mid: str):
    m = motorista_by_id(mid)
    if not m:
        return JSONResponse({"ok": False, "erro": "Motorista não encontrado."}, status_code=404)
    token = motorista_push_token(mid)
    preview = None
    if token:
        preview = token if len(token) <= 60 else f"{token[:57]}..."
    return {
        "ok": True,
        "motorista_id": mid,
        "tem_push_token": bool(token),
        "push_platform": m.get("push_platform"),
        "push_token": preview,
    }


@app.post("/api/app/motorista/{mid}/push-test")
def api_app_motorista_push_test(mid: str):
    """Dispara notificação de teste para validar cadeia Expo Push (tela bloqueada)."""
    m = motorista_by_id(mid)
    if not m:
        return JSONResponse({"ok": False, "erro": "Motorista não encontrado."}, status_code=404)
    token = motorista_push_token(mid)
    if not token:
        return JSONResponse(
            {"ok": False, "erro": "Sem push_token. Faça login no app e aceite notificações."},
            status_code=400,
        )
    ok, detalhe = enviar_push_expo(
        mid,
        token,
        "Teste — Essência Motorista",
        "Push de teste. Se viu isso com tela bloqueada, o push está OK.",
        {"type": "push_test"},
        protocolo_log="TESTE",
    )
    status_code = 200 if ok else 502
    return JSONResponse({"ok": ok, "detalhe": detalhe}, status_code=status_code)

@app.post("/api/app/motorista/{mid}/fcm-token")
def api_app_motorista_fcm_token(mid: str, payload: AppFcmPayload):
    # Nesta V1 apenas registra o token no cadastro do motorista.
    # A próxima etapa conecta com Firebase Admin para push real.
    token = (payload.token or "").strip()
    if not token:
        return {"ok": False, "erro": "Token vazio"}
    try:
        q("alter table motoristas add column if not exists fcm_token text")
    except Exception:
        pass
    q("update motoristas set fcm_token=%s, push_token=%s where id=%s", (token, token, str(mid)))
    return {"ok": True}

@app.get('/motorista/login', response_class=HTMLResponse)
def motorista_login_page(request: Request):
    return templates.TemplateResponse('motorista_login.html', {'request': request, 'erro': None})

@app.post('/motorista/login')
async def motorista_login(request: Request):
    form = await request.form()
    login = (form.get('login') or '').strip()
    senha = (form.get('senha') or '').strip()
    placa = (form.get('placa') or '').upper().replace('-', '').replace(' ', '').strip()

    if not login or not senha or not placa:
        return templates.TemplateResponse('motorista_login.html', {'request': request, 'erro': 'Preencha login, senha e placa do veículo do dia.'})

    m = one("""
        select * from motoristas
        where coalesce(ativo,true)=true
          and (lower(coalesce(login,''))=lower(%s) or regexp_replace(coalesce(cpf,''),'[^0-9]','','g')=regexp_replace(%s,'[^0-9]','','g'))
          and coalesce(senha,'')=%s
        limit 1
    """, (login, login, senha))

    if not m:
        return templates.TemplateResponse('motorista_login.html', {'request': request, 'erro': 'Login ou senha inválidos.'})

    mid = str(m['id'])
    q("update motoristas set placa_atual=%s, placa=%s, online=true, ultima_atualizacao=now(), ultimo_login=now() where id=%s", (placa, placa, mid))
    return RedirectResponse(f'/motorista/{mid}', 303)


@app.get('/motorista/{mid}', response_class=HTMLResponse)
def tela_motorista(mid: str, request: Request):
    return templates.TemplateResponse('motorista.html', {'request':request,'m':motorista_by_id(mid)})
@app.get('/api/motorista/{mid}/servicos')
def api_servicos_motorista(mid: str):
    return [normalizar_servico(r) for r in q("select * from servicos where motorista_id=%s and status not in ('finalizado','recusado') order by created_at desc",(str(mid),),True)]

@app.post('/api/servicos/{sid}/status')
def atualizar_status(sid: str, status: str=Form(...)):
    if not servico_by_id(sid): return JSONResponse({'ok':False,'erro':'Serviço não encontrado'},404)
    if status=="finalizado": q("update servicos set status=%s,atualizado_em=now(),finalizado_em=now() where id=%s",(status,str(sid)))
    else: q("update servicos set status=%s,atualizado_em=now() where id=%s",(status,str(sid)))
    registrar_evento_db(sid,status)
    if (status or "").strip().lower() == "recusado":
        s_atual = servico_by_id(sid)
        protocolo = (s_atual.get("protocolo") or "").strip() if s_atual else ""
        registrar_alerta_operacional(
            sid,
            "recusa",
            "Serviço recusado",
            f"Motorista recusou protocolo {protocolo}" if protocolo else "Motorista recusou o serviço",
        )
    return {'ok':True,'servico':servico_by_id(sid)}
@app.post('/api/servicos/{sid}/placa')
def enviar_placa(sid: str, placa_veiculo: str=Form(...)):
    ok, erro = salvar_placa_cliente_servico(sid, placa_veiculo)
    if not ok:
        status = 404 if erro == "Serviço não encontrado" else 400
        return JSONResponse({"ok": False, "erro": erro}, status_code=status)
    return {"ok": True, "servico": servico_by_id(sid)}
@app.post('/api/servicos/{sid}/fotos')
async def enviar_fotos(sid: str, fotos: list[UploadFile]=File(default=[])):
    if not servico_by_id(sid): return JSONResponse({'ok':False,'erro':'Serviço não encontrado'},404)
    salvas=[]
    for foto in fotos:
        if not foto or not foto.filename: continue
        ext=os.path.splitext(foto.filename)[1].lower() or '.jpg'
        if ext not in ['.jpg','.jpeg','.png','.webp','.gif','.heic']: ext='.jpg'
        nome=f"servico_{sid}_{uuid.uuid4().hex}{ext}"; caminho=os.path.join(UPLOAD_DIR,nome)
        with open(caminho,'wb') as buffer: shutil.copyfileobj(foto.file,buffer)
        url=f"/static/uploads/{nome}"; q("insert into fotos (servico_id,url,filename,created_at) values (%s,%s,%s,now())",(str(sid),url,foto.filename)); salvas.append(url)
    registrar_evento_db(sid,'fotos/checklist',f'Fotos adicionadas: {len(salvas)}'); return {'ok':True,'fotos':salvas,'servico':servico_by_id(sid)}
@app.post('/api/servicos/{sid}/finalizar')
def finalizar_servico(sid: str):
    if not servico_by_id(sid): return JSONResponse({'ok':False,'erro':'Serviço não encontrado'},404)
    q("update servicos set status='finalizado',atualizado_em=now(),finalizado_em=now() where id=%s",(str(sid),)); registrar_evento_db(sid,'finalizado','Serviço finalizado pelo motorista'); return {'ok':True,'servico':servico_by_id(sid)}

@app.get('/faturamento', response_class=HTMLResponse)
def faturamento(request: Request, data_ini: str="", data_fim: str="", seguradora: str="", status_faturamento: str="", motorista: str=""):
    filtros={k:(v or None) for k,v in dict(data_ini=data_ini,data_fim=data_fim,seguradora=seguradora,motorista=motorista).items()}
    where=[]; params=[]
    if filtros.get("data_ini"): where.append("date(created_at) >= %s"); params.append(filtros["data_ini"])
    if filtros.get("data_fim"): where.append("date(created_at) <= %s"); params.append(filtros["data_fim"])
    if filtros.get("seguradora"): where.append("seguradora ilike %s"); params.append(f"%{filtros['seguradora']}%")
    if filtros.get("motorista"): where.append("coalesce(motorista_nome,'') ilike %s"); params.append(f"%{filtros['motorista']}%")
    if status_faturamento: where.append("coalesce(status_faturamento,'para_conferir')=%s"); params.append(status_faturamento)
    sql="select * from servicos" + ((" where " + " and ".join(where)) if where else "") + " order by created_at desc"
    servs=[normalizar_servico(r) for r in q(sql, tuple(params), True)]
    total=sum(float(s.get("valor_total") or 0) for s in servs)
    kpis={
      "total_servicos": len(servs),
      "valor_total": fmt_moeda(total),
      "para_conferir": len([s for s in servs if s.get("status_faturamento")=="para_conferir"]),
      "para_faturar": len([s for s in servs if s.get("status_faturamento")=="para_faturar"]),
      "negociacao": len([s for s in servs if s.get("status_faturamento")=="negociacao"]),
      "faturado": len([s for s in servs if s.get("status_faturamento")=="faturado"]),
    }
    return templates.TemplateResponse('faturamento.html', {'request':request,'servicos':servs,'filtros':dict(data_ini=data_ini,data_fim=data_fim,seguradora=seguradora,status_faturamento=status_faturamento,motorista=motorista),'kpis':kpis,'nav_ativo':'faturamento','nav_som':False})


def _faturamento_redirect_com_filtros(form_or_dict):
    """Monta URL /faturamento preservando filtros da query string."""
    if hasattr(form_or_dict, "get"):
        get = form_or_dict.get
    else:
        get = form_or_dict.get
    params = {}
    for key, form_key in [
        ("data_ini", "data_ini"),
        ("data_fim", "data_fim"),
        ("seguradora", "seguradora"),
        ("motorista", "motorista"),
        ("status_faturamento", "status_faturamento_filtro"),
    ]:
        val = (get(form_key) or get(key) or "").strip()
        if val:
            params[key] = val
    qs = urllib.parse.urlencode(params)
    return f"/faturamento?{qs}" if qs else "/faturamento"


@app.post('/faturamento/acao-massa')
async def faturamento_acao_massa(request: Request):
    form = await request.form()
    ids = form.getlist("servico_ids")
    novo_status = (form.get("status_faturamento") or "").strip()
    permitidos = {"para_conferir", "para_faturar", "negociacao", "faturado"}
    if novo_status not in permitidos:
        return RedirectResponse(_faturamento_redirect_com_filtros(form), 303)

    atualizados = 0
    for sid in ids:
        sid = str(sid).strip()
        if not sid:
            continue
        if not one("select id from servicos where id=%s", (sid,)):
            continue
        q(
            "update servicos set status_faturamento=%s, atualizado_em=now() where id=%s",
            (novo_status, sid),
        )
        registrar_evento_db(sid, "faturamento", f"Status faturamento (massa): {novo_status}")
        atualizados += 1

    destino = _faturamento_redirect_com_filtros(form)
    if atualizados:
        sep = "&" if "?" in destino else "?"
        destino = f"{destino}{sep}massa_ok={atualizados}"
    return RedirectResponse(destino, 303)


_FAT_STATUS_EXPORT = {
    "para_conferir": "Para conferir",
    "para_faturar": "Para faturar",
    "negociacao": "Negociar",
    "faturado": "Faturado",
}


def _formatar_planilha_faturamento_export(ws, total_linhas: int):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    for cell in ws[1]:
        cell.font = Font(bold=True)

    ultima = total_linhas + 1
    for row in range(2, ultima + 1):
        ws.cell(row=row, column=2).number_format = "DD/MM/YYYY HH:MM"
        ws.cell(row=row, column=9).number_format = 'R$ #,##0.00'

    for col_idx in range(1, 10):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ultima + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(val if val is not None else "")))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 48)


@app.post("/faturamento/exportar-excel")
async def faturamento_exportar_excel(request: Request):
    form = await request.form()
    ids = [str(x).strip() for x in form.getlist("servico_ids") if str(x).strip()]
    if not ids:
        return JSONResponse(
            {"ok": False, "erro": "Selecione ao menos um serviço para exportar."},
            status_code=400,
        )

    linhas = []
    for sid in ids:
        row = one("select * from servicos where id=%s", (sid,))
        if not row:
            continue
        fat_st = (row.get("status_faturamento") or "para_conferir").strip()
        dt_raw = _parse_dt_valor(row.get("created_at"))
        data_cell = _dt_para_local_br(dt_raw) if dt_raw else None
        placa = row.get("placa_veiculo_removido") or row.get("placa_removida") or "-"
        linhas.append(
            {
                "Protocolo": row.get("protocolo") or "-",
                "Data": data_cell,
                "Seguradora": row.get("seguradora") or "-",
                "Tipo": row.get("tipo") or "-",
                "Status Operacional": row.get("status") or "-",
                "Status Faturamento": _FAT_STATUS_EXPORT.get(fat_st, fat_st),
                "Motorista": row.get("motorista_nome") or "-",
                "Placa": placa,
                "Valor": float(row.get("valor_total") or 0),
            }
        )

    if not linhas:
        return JSONResponse(
            {"ok": False, "erro": "Selecione ao menos um serviço para exportar."},
            status_code=400,
        )

    df = pd.DataFrame(linhas)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Serviços")
        _formatar_planilha_faturamento_export(writer.sheets["Serviços"], len(linhas))
    output.seek(0)

    nome = f"faturamento_servicos_selecionados_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome}"'},
    )


@app.get('/faturamento/{sid}', response_class=HTMLResponse)
def faturamento_detalhe(sid: str, request: Request, aba: str = "editar", ok: str = ""):
    s = servico_by_id(sid)
    if not s:
        return RedirectResponse('/faturamento', 303)
    itens = itens_do_servico(sid)
    if not itens:
        criar_itens_para_servico(sid, s.get("tipo"))
        s = servico_by_id(sid)
        itens = itens_do_servico(sid)
    end_origem = desmontar_endereco_servico(s.get("origem"))
    end_destino = desmontar_endereco_servico(s.get("destino"))
    chk = checklist_resumo_servico(sid, request, s)
    abas_validas = {"editar", "financeiro", "comentarios", "mobile", "arquivos", "checklist", "patio"}
    aba_ativa = aba if aba in abas_validas else "editar"
    return templates.TemplateResponse(
        'servico_financeiro.html',
        {
            'request': request,
            's': s,
            'itens': itens,
            'tipos_servico': lista_tipos_servico(),
            'precos_por_tipo': {t: itens_padrao_tipo(t) for t in lista_tipos_servico()},
            'motoristas': lista_motoristas(),
            'status_operacionais': [
                'novo', 'enviado', 'aceito', 'a caminho', 'na origem',
                'em transporte', 'finalizado', 'recusado', 'cancelado',
            ],
            'end_origem': end_origem,
            'end_destino': end_destino,
            'comentarios': listar_comentarios_servico(sid),
            'arquivos': coletar_arquivos_servico(sid, s),
            'checklist_resumo': chk,
            'aba_ativa': aba_ativa,
            'ok': ok,
            'nav_ativo': 'faturamento',
            'nav_som': False,
        },
    )


@app.post('/faturamento/{sid}/editar')
async def faturamento_editar(sid: str, request: Request):
    form = await request.form()
    motorista_id = (form.get("motorista_id") or "").strip() or None
    motorista_nome = (form.get("motorista_nome") or "").strip()
    if motorista_id:
        m = one("select nome from motoristas where id=%s", (motorista_id,))
        if m:
            motorista_nome = m.get("nome") or motorista_nome

    origem = montar_endereco_vianet(
        form.get("local_origem"),
        form.get("complemento_origem"),
        form.get("bairro_origem"),
        form.get("cidade_origem"),
        form.get("uf_origem"),
        form.get("cep_origem"),
        form.get("referencia_origem"),
    )
    destino = montar_endereco_vianet(
        form.get("local_destino"),
        form.get("complemento_destino"),
        form.get("bairro_destino"),
        form.get("cidade_destino"),
        form.get("uf_destino"),
        form.get("cep_destino"),
        form.get("referencia_destino"),
    )
    placa = normalizar_placa_cliente(form.get("placa") or "")
    q(
        """
        update servicos set
            protocolo=%s, seguradora=%s, tipo=%s, status=%s,
            motorista_id=%s, motorista_nome=%s,
            placa_veiculo_removido=%s, placa_removida=%s,
            beneficiario=%s, solicitante=%s, telefone_cliente=%s,
            veiculo_cliente=%s, cor_cliente=%s,
            origem=%s, destino=%s,
            referencia_origem=%s, referencia_destino=%s,
            observacao=%s, problema=%s,
            atualizado_em=%s
         where id=%s
        """,
        (
            (form.get("protocolo") or "").strip(),
            (form.get("seguradora") or "").strip(),
            (form.get("tipo") or "").strip(),
            (form.get("status") or "novo").strip(),
            motorista_id,
            motorista_nome or None,
            placa or None,
            placa or None,
            (form.get("beneficiario") or "").strip() or None,
            (form.get("solicitante") or "").strip() or None,
            (form.get("telefone") or "").strip() or None,
            (form.get("veiculo") or "").strip() or None,
            (form.get("cor") or "").strip() or None,
            origem,
            destino,
            (form.get("referencia_origem") or "").strip() or None,
            (form.get("referencia_destino") or "").strip() or None,
            (form.get("observacao") or "").strip() or None,
            (form.get("problema") or "").strip() or None,
            agora_dt(),
            str(sid),
        ),
    )
    registrar_evento_db(sid, "edição", "Dados do serviço atualizados na auditoria")
    return RedirectResponse(f"/faturamento/{sid}?aba=editar&ok=1", 303)


@app.post('/faturamento/{sid}/comentario')
async def faturamento_comentario(sid: str, request: Request):
    form = await request.form()
    texto = (form.get("texto") or "").strip()
    if not servico_by_id(sid):
        return RedirectResponse("/faturamento", 303)
    if texto:
        q(
            "insert into comentarios_servico (servico_id, texto, criado_em) values (%s, %s, %s)",
            (str(sid), texto, agora_dt()),
        )
        registrar_evento_db(sid, "comentário", texto[:120])
    return RedirectResponse(f"/faturamento/{sid}?aba=comentarios&ok=1", 303)

@app.post('/faturamento/{sid}/salvar')
async def faturamento_salvar(sid: str, request: Request):
    form=await request.form()
    status_faturamento=form.get('status_faturamento','para_conferir')
    nomes=form.getlist('item_nome')
    qtds=form.getlist('item_qtd')
    valores=form.getlist('item_valor')
    tipo=form.get('tipo','')
    criar_itens_para_servico(sid, tipo, nomes, qtds, valores)
    total=atualizar_total_servico(sid)
    q("update servicos set tipo=%s,status_faturamento=%s,valor_total=%s,atualizado_em=now() where id=%s",(tipo,status_faturamento,total,str(sid)))
    registrar_evento_db(sid,'faturamento',f'Status: {status_faturamento} - Total: {fmt_moeda(total)}')
    return RedirectResponse(f'/faturamento/{sid}?aba=financeiro&ok=1',303)

@app.post('/faturamento/{sid}/status')
def faturamento_status(sid: str, status_faturamento: str=Form(...)):
    q("update servicos set status_faturamento=%s,atualizado_em=now() where id=%s",(status_faturamento,str(sid)))
    registrar_evento_db(sid,'faturamento',f'Status faturamento: {status_faturamento}')
    return RedirectResponse('/faturamento',303)


@app.get('/exportar')
def exportar(data_ini: str="", data_fim: str="", seguradora: str="", tipo: str="", status: str="", motorista: str=""):
    filtros={k:(v or None) for k,v in dict(data_ini=data_ini,data_fim=data_fim,seguradora=seguradora,tipo=tipo,status=status,motorista=motorista).items()}
    servs=lista_servicos(filtros=filtros); registros=[]
    for s in servs:
        linha=dict(s); linha['fotos']=', '.join([f.get("url","") for f in s.get('fotos',[])]); linha['historico']=' | '.join([f"{h.get('data_hora')} - {h.get('status')} - {h.get('detalhe','')}" for h in s.get('historico',[])]); registros.append(linha)
    output=io.BytesIO()
    with pd.ExcelWriter(output,engine='openpyxl') as writer:
        pd.DataFrame(registros).to_excel(writer,index=False,sheet_name='Servicos')
        itens=[]
        for s in servs:
            for it in itens_do_servico(s["id"]):
                itens.append({"protocolo":s.get("protocolo"),"seguradora":s.get("seguradora"),"tipo":s.get("tipo"),"item":it.get("nome_item"),"quantidade":it.get("quantidade"),"valor_unitario":it.get("valor_unitario"),"valor_total":it.get("valor_total"),"status_faturamento":s.get("status_faturamento")})
        pd.DataFrame(itens).to_excel(writer,index=False,sheet_name='Itens_Faturamento')
        pd.DataFrame(lista_motoristas()).to_excel(writer,index=False,sheet_name='Motoristas')
    output.seek(0)
    return StreamingResponse(output,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':'attachment; filename=sistema_reboque_relatorio.xlsx'})

import financeiro_routes
import financeiro_notas_entrada
import financeiro_nfse
import financeiro_config_fiscal
import cadastros_import_clientes
import relatorios_faturamento
import dashboard_api
financeiro_routes.register(app, templates)
financeiro_notas_entrada.register(app, templates)
financeiro_config_fiscal.register(app, templates)
financeiro_nfse.register(app, templates)
cadastros_import_clientes.register(app)
relatorios_faturamento.register(app, templates)
dashboard_api.register(app, templates)
